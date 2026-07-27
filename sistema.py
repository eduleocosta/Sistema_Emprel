import os
import tempfile
import subprocess
from PIL import Image, ImageTk, ImageDraw, ImageGrab, ImageOps
import ctypes
from ctypes import wintypes
import json
import re
import hashlib
import secrets
import uuid
from tkinter import ttk
from datetime import datetime
import tkinter as tk
from tkinter import messagebox, simpledialog
import ttkbootstrap as tb
from ttkbootstrap.constants import *
import customtkinter as ctk
import sys
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
import pandas as pd
from openpyxl.styles import Font
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import unicodedata
import zipfile
import shutil
import socket
import platform
import time
from datetime import datetime, timedelta


from config import (
    PASTA_LISTAS, APP_DIR, PASTA_DOC, PASTA_IMG, ARQUIVO_ACAO_ATIVA,
    PASTA_BACKUP_LOCAL, PASTA_BACKUP_NUVEM, DB_FILE, DB_USUARIOS, DB_ENTREGAS,
    DB_EXCLUIDOS_CADASTROS, DB_EXCLUIDOS_ACOES, HISTORICO_BACKUP, APP_TITLE, THEME,
    DB_VANS
)
from database.core import load_db, save_db, load_usuarios, save_usuarios, load_organizacoes, save_organizacoes, load_vans, save_vans, carregar_lista_tree, salvar_lista_tree, salvar_excluido, registrar_entrega_json_separado
from database.models import (
    migrar_usuarios, migrar_telefones_antigos, migrar_servicos_antigos,
    migrar_cadastros_acoes_orfas, migrar_organizacoes, migrar_vans, hoje_str, nome_arquivo_seguro,
    caminho_pdf, get_nome_servico_por_id, get_acao_por_id, get_nome_acao_por_id, hash_password, verify_password,
    senha_expirada_60_dias
)
from services.pdf_service import gerar_pdf_relatorio, gerar_pdf_relatorio_acao, gerar_pdf_ranking
from utils.validators import validar_email, validar_data_nascimento, validar_cpf, somente_digs, formatar_telefone_novo
from utils.formatters import mascara_cpf_entrada, mascara_tel_entrada, mascara_data_entrada, mask_cpf_from_clean, mascarar_cpf_pdf, normalizar_texto_ordenacao
from utils.backup import verificar_backup_diario, criar_backup_zip, restaurar_backup_zip, registrar_backup_realizado, registrar_restore_realizado

ARQUIVO_SESSAO = os.path.join(tempfile.gettempdir(), "sistema_emprel_sessao.json")
ARQUIVO_MODO_TESTE = os.path.join(APP_DIR, "modo_teste.json")

# Garantir pasta backup
def garantir_pasta_backup():

    os.makedirs(PASTA_BACKUP_LOCAL, exist_ok=True)

    os.makedirs(PASTA_BACKUP_NUVEM, exist_ok=True)


# Carrega histórico backup
def carregar_historico_backup():
    if not os.path.exists(HISTORICO_BACKUP):
        return {
            "ultimo_backup": None,
            "ultimo_restore": None
        }

    try:
        with open(HISTORICO_BACKUP, "r", encoding="utf-8") as f:
            return json.load(f)
    except:
        return {
            "ultimo_backup": None,
            "ultimo_restore": None
        }

# Salva histórico backup
def salvar_historico_backup(historico):
    with open(HISTORICO_BACKUP, "w", encoding="utf-8") as f:
        json.dump(historico, f, indent=4, ensure_ascii=False)


# Registra backup realizado
def registrar_backup_realizado(tipo, caminho):
    historico = carregar_historico_backup()
    historico["ultimo_backup"] = {
        "tipo": tipo,
        "data": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "caminho": caminho
    }
    salvar_historico_backup(historico)

# Registra restaura realizado
def registrar_restore_realizado(caminho):
    historico = carregar_historico_backup()
    historico["ultimo_restore"] = {
        "data": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "caminho": caminho
    }
    salvar_historico_backup(historico)


# Pasta backup deve pular
def _pasta_backup_deve_pular(raiz_abs, pasta_backup_nuvem, pasta_backup_local):
    return (
        raiz_abs == pasta_backup_nuvem
        or raiz_abs.startswith(pasta_backup_nuvem + os.sep)
        or raiz_abs == pasta_backup_local
        or raiz_abs.startswith(pasta_backup_local + os.sep)
    )

# Arquivos backup
def arquivos_backup():
    pasta_backup_nuvem = os.path.normcase(os.path.normpath(os.path.abspath(PASTA_BACKUP_NUVEM)))
    pasta_backup_local = os.path.normcase(os.path.normpath(os.path.abspath(PASTA_BACKUP_LOCAL)))

    arquivos = []

    for raiz, _, nomes in os.walk(APP_DIR):
        raiz_abs = os.path.normcase(os.path.normpath(os.path.abspath(raiz)))

        if _pasta_backup_deve_pular(raiz_abs, pasta_backup_nuvem, pasta_backup_local):
            nomes[:] = []
            continue

        for nome in nomes:
            caminho = os.path.join(raiz, nome)
            arcname = os.path.relpath(caminho, APP_DIR)
            arquivos.append((caminho, arcname))

    return arquivos

# Arquivos backup por tipo
def arquivos_backup_por_tipo(tipo):
    pasta_backup_nuvem = os.path.normcase(os.path.normpath(os.path.abspath(PASTA_BACKUP_NUVEM)))
    pasta_backup_local = os.path.normcase(os.path.normpath(os.path.abspath(PASTA_BACKUP_LOCAL)))

    if tipo == "sistema":
        return arquivos_backup()

    arquivos = []

    if tipo == "db":
        pastas = [APP_DIR]
        filtro = lambda nome: nome.lower().endswith(".json")
        pasta_base = APP_DIR
    elif tipo == "doc_img":
        pastas = [PASTA_DOC, PASTA_IMG]
        filtro = lambda nome: True
        pasta_base = APP_DIR
    else:
        raise ValueError(f"Tipo de backup inválido: {tipo}")

    for pasta in pastas:
        if not os.path.exists(pasta):
            continue

        for raiz, _, nomes in os.walk(pasta):
            raiz_abs = os.path.normcase(os.path.normpath(os.path.abspath(raiz)))

            if _pasta_backup_deve_pular(raiz_abs, pasta_backup_nuvem, pasta_backup_local):
                continue

            for nome in nomes:
                if not filtro(nome):
                    continue

                caminho = os.path.join(raiz, nome)
                arcname = os.path.relpath(caminho, pasta_base)
                arquivos.append((caminho, arcname))

    return arquivos

# Nome usuário logado
def _nome_usuario_logado(self):
    dados = self.db.get("usuarios", {}).get(self.usuario or "", {})
    return dados.get("nome") or mask_cpf_from_clean(self.usuario or "")

# Carrega estado modo teste
def _carregar_modo_teste():
    if not os.path.exists(ARQUIVO_MODO_TESTE):
        return None
    try:
        with open(ARQUIVO_MODO_TESTE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None

# Salva estado modo teste
def _salvar_modo_teste(dados):
    try:
        with open(ARQUIVO_MODO_TESTE, "w", encoding="utf-8") as f:
            json.dump(dados, f, ensure_ascii=False, indent=4)
    except Exception:
        pass

# Verifica se modo teste está ativo e dentro do prazo
def _verificar_modo_teste_ativo():
    dados = _carregar_modo_teste()
    if not dados or not dados.get("ativo"):
        return False
    try:
        expira_em = datetime.strptime(dados.get("expira_em", ""), "%d/%m/%Y %H:%M:%S")
        if datetime.now() > expira_em:
            _salvar_modo_teste({"ativo": False})
            return False
        return True
    except Exception:
        return False

# Ativa modo teste
def _ativar_modo_teste(admin_cpf="", duracao_minutos=10, max_prorrogacoes=5):
    expira_em = datetime.now() + timedelta(minutes=duracao_minutos)
    dados = {
        "ativo": True,
        "expira_em": expira_em.strftime("%d/%m/%Y %H:%M:%S"),
        "prorrogacoes": 0,
        "max_prorrogacoes": max_prorrogacoes,
        "duracao_minutos": duracao_minutos,
        "admin_ativou": admin_cpf
    }
    _salvar_modo_teste(dados)
    return dados

# Prorroga modo teste
def _prorrogar_modo_teste():
    dados = _carregar_modo_teste()
    if not dados or not dados.get("ativo"):
        return None
    try:
        expira_em = datetime.strptime(dados.get("expira_em", ""), "%d/%m/%Y %H:%M:%S")
        prorrogacoes = dados.get("prorrogacoes", 0)
        max_prorrogacoes = dados.get("max_prorrogacoes", 5)
        if prorrogacoes >= max_prorrogacoes:
            _salvar_modo_teste({"ativo": False})
            return None
        nova_expiracao = expira_em + timedelta(minutes=dados.get("duracao_minutos", 10))
        dados["expira_em"] = nova_expiracao.strftime("%d/%m/%Y %H:%M:%S")
        dados["prorrogacoes"] = prorrogacoes + 1
        _salvar_modo_teste(dados)
        return dados
    except Exception:
        return None

# Desativa modo teste
def _desativar_modo_teste():
    _salvar_modo_teste({"ativo": False})

# Mascara telefone entrada
def mascara_tel_entrada(entry: tk.Entry):
    t = somente_digs(entry.get())[:11]

    if len(t) == 0:
        out = ""
    elif len(t) <= 2:
        out = f"({t}"
    elif len(t) <= 6:
        out = f"({t[:2]}) {t[2:]}"
    elif len(t) <= 10:
        out = f"({t[:2]}) {t[2:6]}-{t[6:]}"
    else:
        out = f"({t[:2]}) {t[2]}.{t[3:7]}-{t[7:]}"

    if entry.get() != out:
        entry.delete(0, tk.END)
        entry.insert(0, out)
# Mascara data entrada
def mascara_data_entrada(entry: tk.Entry):
    t = somente_digs(entry.get())[:8]
    out = ""
    if len(t) >= 2:
        out = t[:2] + "/"
    else:
        out = t
    if len(t) >= 4:
        out += t[2:4] + "/"
    elif len(t) > 2:
        out += t[2:]
    if len(t) >= 5:
        out += t[4:]
    if entry.get() != out:
        entry.delete(0, tk.END)
        entry.insert(0, out)

# Mascara CPF a partir de limpo
def mask_cpf_from_clean(cpf: str) -> str:
    if len(cpf) == 11:
        return f"{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:]}"
    return cpf

# Mascara CPF PDF
def mascarar_cpf_pdf(cpf: str) -> str:
    cpf = somente_digs(cpf)

    if len(cpf) != 11:
        return cpf

    return f"xxx.xx{cpf[5]}.{cpf[6:9]}-{cpf[9:]}"

# Obtém nome ação por id
def get_nome_acao_por_id(db, acao_id):
    acao_id = str(acao_id or "").strip()
    acao = next((a for a in db.get("acoes", []) if str(a.get("id", "")).strip() == acao_id), None)
    if not acao:
        return ""

    return acao.get("local", "")

# Obtém ação por id
def get_acao_por_id(db, acao_id):
    return next((a for a in db.get("acoes", []) if str(a.get("id")) == str(acao_id)), None)

class App(tb.Window):
    # Inicializa
    def __init__(self):
        super().__init__(title=APP_TITLE, themename=THEME, size=(1100, 700))
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")
        self.protocol("WM_DELETE_WINDOW", self._fechar_sistema)
        self._largura_padrao = 1100
        self._altura_padrao = 940
        self._janela_posicionada = False
        migrar_usuarios()

        self.lista_atual_nome = "lista_temp.json"
        self._cadastro_ordenacao_reversa = {}
        self.db = load_db()

        # Corrigir entregas antigas
        def corrigir_entregas_antigas(self):

            protocolos = self._protocolos_entregues_confirmados()

            alterou = False

            for r in self.db.get("cadastros", []):

                protocolo = str(r.get("protocolo", "")).strip()

                if protocolo in protocolos and not r.get("entregue", False):

                    r["entregue"] = True
                    alterou = True

            if alterou:
                save_db(self.db)

        self.usuarios = load_usuarios()
        self.vans = load_vans()
        migrar_telefones_antigos(self.db)
        migrar_servicos_antigos(self.db)
        migrar_cadastros_acoes_orfas(self.db)
        migrar_organizacoes()
        migrar_vans()
        self.usuario = None
        self.perfil = None
        self._home_ativa = False
        self._botoes_home = []
        self._bind_home_ids = {}
        self._comando_voltar_atual = None
        for seq in ("<Left>", "<Right>", "<Up>", "<Down>", "<Return>"):
            self._bind_home_ids[seq] = self.bind(seq, self._on_home_key, add="+")
        self.bind("<Escape>", self._on_escape_global)
        if "servicos" not in self.db:
            self.db["servicos"] = []
        self.lista_servicos = self.db["servicos"]
        self._selected_acao = None
        self._selected_servico = None
        self._limpar_acao_ativa_expirada()
        self._carregar_acao_ativa()
        self.corrigir_acoes_maiusculas()
        verificar_backup_diario()
        self._build_login_screen()

    # Corrigir acoes maiusculas
    def corrigir_acoes_maiusculas(self):
        alterado = False

        for a in self.db.get("acoes", []):
            local = a.get("local", "")
            if local != local.upper():
                a["local"] = local.upper()
                alterado = True

        if alterado:
            save_db(self.db)

    # Salva ação ativa
    def _salvar_acao_ativa(self):
        if not self._selected_acao or not self._selected_servico:
            return

        dados = {
            "acao": self._selected_acao,
            "servico": self._selected_servico
        }

        with open(ARQUIVO_ACAO_ATIVA, "w", encoding="utf-8") as f:
            json.dump(dados, f, ensure_ascii=False, indent=4)

    # Limpa ação ativa expirada
    def _limpar_acao_ativa_expirada(self):
        pass

    # Carrega ação ativa
    def _carregar_acao_ativa(self):
        if not os.path.exists(ARQUIVO_ACAO_ATIVA):
            return

        try:
            with open(ARQUIVO_ACAO_ATIVA, "r", encoding="utf-8") as f:
                dados = json.load(f)

            acao = dados.get("acao")
            servico = dados.get("servico")

            if not acao or not servico:
                return

            data_acao_str = acao.get("data")
            if not data_acao_str:
                return

            data_acao = datetime.strptime(data_acao_str, "%d/%m/%Y").date()
            hoje = datetime.now().date()

            if data_acao != hoje:
                self._selected_acao = None
                self._selected_servico = None

                try:
                    os.remove(ARQUIVO_ACAO_ATIVA)
                except Exception:
                    pass

                return

            self._selected_acao = acao
            self._selected_servico = servico

        except Exception:
            self._selected_acao = None
            self._selected_servico = None

    # Sincronizar ação ativa banco de dados
    def _sincronizar_acao_ativa_db(self):
        acao_ativa = self.db.get("acao_ativa") or {}
        acao_id = str(acao_ativa.get("id", "")).strip()
        if not acao_id:
            return False

        acao = next(
            (a for a in self.db.get("acoes", []) if str(a.get("id", "")).strip() == acao_id),
            None
        )
        if not acao:
            return False

        self._selected_acao = {
            "id": acao.get("id", ""),
            "data": acao.get("data", acao_ativa.get("data", "")),
            "local": acao.get("local", ""),
            "vans": acao.get("vans", [])
        }

        servico_id = str(acao_ativa.get("servico_id", "")).strip()
        if not servico_id:
            servicos_acao = acao.get("servicos", [])
            if servicos_acao:
                servico_id = str(servicos_acao[0]).strip()

        if not servico_id:
            cadastro = next(
                (
                    r for r in self.db.get("cadastros", [])
                    if str(r.get("acao_id", "")).strip() == acao_id
                    and str(r.get("servico_id", "")).strip()
                ),
                None
            )
            if cadastro:
                servico_id = str(cadastro.get("servico_id", "")).strip()

        servico = next(
            (s for s in self.db.get("servicos", []) if str(s.get("id", "")).strip() == servico_id),
            None
        )
        self._selected_servico = {
            "id": servico.get("id", ""),
            "nome": servico.get("nome", "")
        } if servico else None

        return True

    # Preencher campos ação ativa
    def _preencher_campos_acao_ativa(self):
        if self._selected_acao:
            self.ent_acao_info.configure(state="normal")
            self.ent_acao_info.delete(0, tk.END)
            self.ent_acao_info.insert(0, f"{self._selected_acao.get('data', '')} - {self._selected_acao.get('local', '')}")
            self.ent_acao_info.configure(state="disabled")
            self.ent_data.configure(state="normal")
            self.ent_data.delete(0, tk.END)
            self.ent_data.insert(0, self._selected_acao.get("data") or hoje_str())
            self.ent_data.configure(state="disabled")
        else:
            self.ent_acao_info.configure(state="normal")
            self.ent_acao_info.delete(0, tk.END)
            self.ent_acao_info.insert(0, "Nenhuma ação selecionada")
            self.ent_acao_info.configure(state="disabled")
            self.ent_data.configure(state="normal")
            self.ent_data.delete(0, tk.END)
            self.ent_data.insert(0, "")
            self.ent_data.configure(state="disabled")

        if self._selected_servico:
            nome_servico = self._selected_servico.get("nome", "") if isinstance(self._selected_servico, dict) else get_nome_servico_por_id(self.db, self._selected_servico)
            self.ent_servico_info.configure(state="normal")
            self.ent_servico_info.delete(0, tk.END)
            self.ent_servico_info.insert(0, nome_servico)
            self.ent_servico_info.configure(state="disabled")
        else:
            self.ent_servico_info.configure(state="normal")
            self.ent_servico_info.delete(0, tk.END)
            self.ent_servico_info.insert(0, "Nenhum serviço selecionado")
            self.ent_servico_info.configure(state="disabled")

    # Volta para login
    def _voltar_para_login(self):
      if self.usuario == "00000000000" or _verificar_modo_teste_ativo():
          _desativar_modo_teste()
          self._garantir_usuario_teste()
          self.usuarios["00000000000"]["perguntas_usadas"] = []
          self.usuarios["00000000000"]["pergunta_atual_id"] = None
          save_usuarios(self.usuarios)
      self.usuario = None
      self.perfil = None
      if os.path.exists(ARQUIVO_SESSAO):
          try:
              os.remove(ARQUIVO_SESSAO)
          except Exception:
              pass
      self._build_login_screen()

    # Fazer backup manual
    def _fazer_backup_manual(self):
        self._open_backup_screen()

    # Abre backup tela
    def _open_backup_screen(self):
        top = self._open_inline_container("Backup", voltar=self._build_home_screen)

        ctk.CTkLabel(
            top,
            text="Escolha uma opção de backup",
            font=("Segoe UI", 15, "bold"),
            text_color=("#0F172A", "#E2E8F0")
        ).pack(pady=(0, 8))

        botoes = ctk.CTkFrame(top, fg_color="transparent")
        botoes.pack(fill="x", pady=(0, 12))
        botoes.columnconfigure(0, weight=1)
        botoes.columnconfigure(1, weight=1)
        botoes.rowconfigure(0, weight=1)
        botoes.rowconfigure(1, weight=1)

        botoes_backup = [
            {
                "icone": "DB",
                "texto": "db",
                "subtexto": "Arquivos JSON",
                "tipo": "db",
                "descricao": "arquivos JSON",
                "bootstyle": "primary",
                "linha": 0,
                "coluna": 0
            },
            {
                "icone": "DI",
                "texto": "doc e img",
                "subtexto": "Docs + imagens",
                "tipo": "doc_img",
                "descricao": "pastas doc e img",
                "bootstyle": "warning",
                "linha": 0,
                "coluna": 1
            },
            {
                "icone": "Sistema",
                "texto": "Completo",
                "subtexto": "sistema",
                "tipo": "sistema",
                "descricao": "toda a pasta do sistema",
                "bootstyle": "success",
                "linha": 1,
                "coluna": 0
            },
            {
                "icone": "↩",
                "texto": "Restaurar",
                "subtexto": "Backup",
                "tipo": None,
                "descricao": "backup",
                "bootstyle": "danger",
                "linha": 1,
                "coluna": 1
            }
        ]

        for cfg in botoes_backup:
            if cfg["tipo"]:
                comando = lambda cfg=cfg: self._criar_backup_com_feedback(top, cfg["tipo"], cfg["descricao"])
            else:
                comando = self._restaurar_backup_manual

            cores_card = {
                "primary": ("#2563EB", "#1D4ED8"),
                "warning": ("#D97706", "#B45309"),
                "success": ("#16A34A", "#15803D"),
                "danger": ("#E53935", "#C62828"),
            }
            fg_card, hv_card = cores_card.get(cfg["bootstyle"], cores_card["primary"])
            card = ctk.CTkButton(
                botoes,
                text=f"{cfg['icone']}\n{cfg['texto']}",
                command=comando,
                fg_color=fg_card,
                hover_color=hv_card,
                text_color="white",
                corner_radius=10,
                height=90,
                font=("Segoe UI", 13, "bold")
            )
            card.grid(row=cfg["linha"], column=cfg["coluna"], padx=12, pady=12, sticky="ew")

            card.configure(cursor="hand2")

        self._render_backup_historico(top)

    # Renderiza backup histórico
    def _render_backup_historico(self, parent):
        historico = carregar_historico_backup()
        frame = ctk.CTkFrame(parent, corner_radius=10, border_width=1, border_color=("#CBD5E1", "#334155"), fg_color=("#F8FAFC", "#0F172A"))
        frame.pack(fill="x", pady=(0, 8))

        ctk.CTkLabel(
            frame,
            text="Histórico",
            font=("Segoe UI", 11, "bold"),
            text_color=("#334155", "#CBD5E1"),
            anchor="w"
        ).pack(fill="x", padx=10, pady=(8, 4))

        backup = historico.get("ultimo_backup") or {}
        restore = historico.get("ultimo_restore") or {}

        linhas = [
            f"Último backup: {backup.get('data', 'Nunca')}",
            f"Tipo: {backup.get('tipo_desc', self._nome_tipo_backup(backup.get('tipo')))}",
            f"Arquivo: {os.path.basename(backup.get('caminho', '')) if backup.get('caminho') else 'Nenhum'}",
            f"Sistema restaurado: {restore.get('data', 'Nunca')}",
            f"Arquivo restaurado: {os.path.basename(restore.get('caminho', '')) if restore.get('caminho') else 'Nenhum'}"
        ]

        for linha in linhas:
            ctk.CTkLabel(
                frame,
                text=linha,
                anchor="w",
                justify="left",
                text_color=("#475569", "#94A3B8")
            ).pack(fill="x", padx=10, pady=(0, 2))

    # Nome tipo backup
    def _nome_tipo_backup(self, tipo):
        nomes = {
            "db": "Arquivos JSON",
            "doc_img": "Pastas doc e img",
            "sistema": "Pasta do sistema"
        }
        return nomes.get(tipo, tipo or "-")

    # Cria backup com feedback
    def _criar_backup_com_feedback(self, parent, tipo, descricao):
        if not self._autorizar_admin():
            return

        try:
            caminho = criar_backup_zip(tipo)
            registrar_backup_realizado(tipo, caminho)
            self._render_backup_historico(parent)
            messagebox.showinfo(
                "Sucesso",
                f"Backup de {descricao} realizado com sucesso.\n\n"
                f"Arquivo: {os.path.basename(caminho)}"
            )
        except Exception as e:
            messagebox.showerror(
                "Erro",
                f"Erro ao criar backup:\n\n{e}"
            )

    # Restaura backup manual
    def _restaurar_backup_manual(self):

        from tkinter import filedialog

        if not self._autorizar_admin():
            return

        caminho = filedialog.askopenfilename(
            initialdir=PASTA_BACKUP_LOCAL,
            title="Selecionar Backup",
            filetypes=[("Arquivo ZIP", "*.zip")]
        )

        if not caminho:
            return

        confirmar = messagebox.askyesno(
            "Confirmação",
            "Deseja restaurar este backup?\n\n"
            "Os arquivos atuais serão substituídos."
        )

        if not confirmar:
            return

        try:

            restaurar_backup_zip(caminho)
            registrar_restore_realizado(caminho)

            self.db = load_db()

            if self.usuario:
                try:
                    with open(ARQUIVO_SESSAO, "w", encoding="utf-8") as f:
                        json.dump({"usuario": self.usuario}, f)
                except Exception:
                    pass

            messagebox.showinfo(
                "Sucesso",
                "Backup restaurado com sucesso.\n\n"
                "O sistema será reiniciado."
            )

            self.destroy()
            subprocess.Popen([sys.executable, os.path.join(APP_DIR, "sistema.py")])

        except Exception as e:

            messagebox.showerror(
                "Erro",
                f"Erro ao restaurar backup:\n\n{e}"
            )

    # Fecha sistema
    def _fechar_sistema(self):
        try:
            self.quit()
        finally:
            tb.Window.destroy(self)

    # Reinicia sistema
    def _reiniciar_sistema(self):
        if not self.usuario:
            messagebox.showinfo("Aviso", "Nenhum usuário logado.")
            return

        confirmar = messagebox.askyesno(
            "Reiniciar Sistema",
            "Deseja reiniciar o sistema?\n\nVocê será reconectado automaticamente."
        )

        if not confirmar:
            return

        try:
            if _verificar_modo_teste_ativo():
                with open(ARQUIVO_SESSAO, "w", encoding="utf-8") as f:
                    json.dump({"usuario": "00000000000"}, f)
            else:
                with open(ARQUIVO_SESSAO, "w", encoding="utf-8") as f:
                    json.dump({"usuario": self.usuario}, f)
        except Exception:
            pass

        self.destroy()
        subprocess.Popen([sys.executable, os.path.join(APP_DIR, "sistema.py")])

    # Limpa tela
    def _clear_screen(self):
        self._home_ativa = False
        self._comando_voltar_atual = None
        for widget in self.winfo_children():
            widget.destroy()

    # Ao início tecla
    def _on_home_key(self, event):
        if not getattr(self, "_home_ativa", False):
            return

        tecla = event.keysym
        botoes = getattr(self, "_botoes_home", [])
        botoes_validos = [b for b in botoes if b.winfo_exists()]
        if not botoes_validos:
            return

        focado = self.focus_get()
        if focado not in botoes_validos:
            if tecla == "Return":
                return
            return

        if tecla == "Return":
            try:
                focado.invoke()
            except Exception:
                pass
            return

        idx = botoes_validos.index(focado)
        total = len(botoes_validos)
        cols = 4

        if tecla == "Left":
            novo = (idx - 1) % total
        elif tecla == "Right":
            novo = (idx + 1) % total
        elif tecla == "Up":
            novo = (idx - cols) % total
        elif tecla == "Down":
            novo = (idx + cols) % total
        else:
            return

        botoes_validos[novo].focus_set()

    # Ao escape global
    def _on_escape_global(self, event):
        cmd = getattr(self, "_comando_voltar_atual", None)
        if callable(cmd):
            try:
                cmd()
            except Exception:
                pass

    # Ajusta janela responsiva
    def _ajustar_janela_responsiva(self, largura=1100, altura=850):
        screen_w = self.winfo_screenwidth()
        screen_h = self.winfo_screenheight()
        largura = min(largura, max(760, screen_w - 80))
        altura = min(altura, max(520, screen_h - 120))

        if not self._janela_posicionada:
            x = max(0, (screen_w // 2) - (largura // 2))
            y = max(0, (screen_h // 2) - (altura // 2))
            self._janela_posicionada = True
        else:
            x = self.winfo_x()
            y = self.winfo_y()

        self.geometry(f"{largura}x{altura}+{x}+{y}")
        self.minsize(760, 520)

    # Ajusta janela principal
    def _ajustar_janela_principal(self):
        self._ajustar_janela_responsiva(self._largura_padrao, self._altura_padrao)

    # Inicia monitoramento modo teste
    def _iniciar_monitoramento_modo_teste(self):
        if getattr(self, "_monitoramento_modo_teste_id", None):
            try:
                self.after_cancel(self._monitoramento_modo_teste_id)
            except Exception:
                pass
        self._tentativas_prorrogacao = 0
        self._verificar_expiracao_modo_teste()

    # Verifica expiração modo teste
    def _verificar_expiracao_modo_teste(self):
        if not _verificar_modo_teste_ativo():
            self._monitoramento_modo_teste_id = None
            return

        dados_mt = _carregar_modo_teste()
        if not dados_mt:
            self._monitoramento_modo_teste_id = None
            return

        try:
            expira_em = datetime.strptime(dados_mt.get("expira_em", ""), "%d/%m/%Y %H:%M:%S")
            restante = expira_em - datetime.now()
            minutos = restante.total_seconds() / 60

            if minutos <= 2 and minutos > 0:
                if not getattr(self, "_aviso_expiracao_mostrado", False):
                    self._aviso_expiracao_mostrado = True
                    self._mostrar_dialogo_prorrogacao()
                    return
        except Exception:
            pass

        self._monitoramento_modo_teste_id = self.after(30000, self._verificar_expiracao_modo_teste)

    # Mostra diálogo prorrogação
    def _mostrar_dialogo_prorrogacao(self):
        confirmar = messagebox.askyesno(
            "Modo Teste",
            "O Modo Teste está prestes a expirar!\n\n"
            "Deseja prorrogar por mais 10 minutos?"
        )

        if not confirmar:
            _desativar_modo_teste()
            self._build_home_screen()
            return

        if not self._verificar_pergunta_modo_teste():
            self._tentativas_prorrogacao = getattr(self, "_tentativas_prorrogacao", 0) + 1
            if self._tentativas_prorrogacao >= 3:
                messagebox.showerror(
                    "Modo Teste",
                    "3 tentativas incorretas. Modo Teste encerrado imediatamente."
                )
                _desativar_modo_teste()
                self._build_home_screen()
                return

            messagebox.showwarning(
                "Aviso",
                f"Resposta incorreta. Tentativa {self._tentativas_prorrogacao}/3.\n"
                f"Você tem mais {3 - self._tentativas_prorrogacao} tentativa(s)."
            )
            self._mostrar_dialogo_prorrogacao()
            return

        dados = _prorrogar_modo_teste()
        if dados:
            self._tentativas_prorrogacao = 0
            self._aviso_expiracao_mostrado = False
            messagebox.showinfo(
                "Modo Teste",
                f"Modo Teste prorrogado com sucesso!\n\n"
                f"Nova expiração: {dados.get('expira_em', '')}"
            )
            self._monitoramento_modo_teste_id = self.after(30000, self._verificar_expiracao_modo_teste)
        else:
            messagebox.showwarning("Aviso", "Não foi possível prorrogar. Limite atingido.")
            _desativar_modo_teste()
            self._build_home_screen()

    # Abre interna container
    def _open_inline_container(self, titulo, voltar=None, padding=12, mostrar_voltar=True):
        self._clear_screen()
        self._ajustar_janela_principal()

        wrapper = tb.Frame(self, padding=10)
        wrapper.pack(fill="both", expand=True)

        destino_voltar = voltar or (self._build_home_screen if self.usuario else self._build_login_screen)
        texto_cancelar = "Cancelar" if not self.usuario or destino_voltar != self._build_home_screen else "Voltar"
        self._criar_cabecalho(
            wrapper,
            titulo,
            voltar=destino_voltar if mostrar_voltar else None,
            mostrar_sair=True,
            texto_cancelar=texto_cancelar
        )

        conteudo = tb.Frame(wrapper, padding=padding)
        conteudo.pack(fill="both", expand=True)

        # Fecha tela
        def fechar_tela():
            destino_voltar()

        conteudo.title = lambda *args, **kwargs: None
        conteudo.geometry = lambda *args, **kwargs: self._ajustar_janela_principal()
        conteudo.grab_set = lambda *args, **kwargs: None
        conteudo.resizable = lambda *args, **kwargs: None
        conteudo.fechar_tela = fechar_tela
        return conteudo

    # Fecha tela interna
    def _fechar_tela_interna(self, tela):
        if hasattr(tela, "fechar_tela"):
            tela.fechar_tela()
        else:
            tela.destroy()

    # Renderiza ação ativa
    def _render_acao_ativa(self, parent):
        frame = tb.Frame(parent, padding=8, bootstyle="light")
        frame.pack(fill="x", pady=(0, 10))

        if self._selected_acao and self._selected_servico:
            if isinstance(self._selected_servico, dict):
                nome_servico = self._selected_servico.get("nome", "")
            else:
                nome_servico = get_nome_servico_por_id(self.db, self._selected_servico)

            texto = (
                f"Ação ativa: {self._selected_acao.get('data')} - "
                f"{self._selected_acao.get('local')}   |   "
                f"Serviço: {nome_servico}"
            )
            estilo = "success"
        else:
            texto = "Ação ativa: Nenhuma ação selecionada   |   Serviço: Nenhum serviço selecionado"
            estilo = "warning"

        tb.Label(
            frame,
            text=texto,
            bootstyle=estilo,
            font=("Segoe UI", 10, "bold"),
            anchor="w"
        ).pack(fill="x")

    # Troca ação
    def _trocar_acao(self):
        if not self._selected_acao or not self._selected_servico:
            self._open_selecao_acao_servico()
            return

        escolha = messagebox.askyesnocancel(
            "Ação ativa",
            "Já existe uma ação ativa.\n\n"
            "👉 SIM = Usar esta ação\n"
            "👉 NÃO = Trocar ação\n"
            "👉 CANCELAR = Voltar"
        )

        if escolha is True:
            self._build_cadastro_screen(preselect=True)

        elif escolha is False:
            self._open_selecao_acao_servico()

    # Registra log
    def _registrar_log(self, acao, info="", tipo="geral"):
        data_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        usuario = self.usuario if hasattr(self, "usuario") else "N/A"

        base = f"[{data_hora}] | Usuário: {usuario} | Ação: {acao} | Tipo: {tipo} | "

        if isinstance(info, dict):
            dados_formatados = []

            for chave, valor in info.items():
                dados_formatados.append(f"{chave.capitalize()}: {valor}")

            linha = base + " | ".join(dados_formatados) + "\n"
        else:
            linha = base + str(info) + "\n"

        try:
            with open("logs.txt", "a", encoding="utf-8") as f:
                f.write(linha)
        except Exception as e:
            print("Erro ao gravar log:", e)

    # Abre arquivo
    @staticmethod
    def _abrir_arquivo(caminho: str):
        try:
            os.startfile(caminho)
        except Exception as e:
            print("Erro ao abrir arquivo:", e)

    # Obtém ip
    def _get_ip():
        try:
            return socket.gethostbyname(socket.gethostname())
        except:
            return "N/A"

    # Obtém maquina
    def _get_maquina():
        try:
            return platform.node()
        except:
            return "N/A"

    # Limpa
    def _clear(self):
        for w in self.winfo_children():
            w.destroy()

    # Cria cabeçalho
    def _criar_cabecalho(self, parent, titulo, voltar=None, mostrar_sair=True, texto_cancelar="Cancelar", padding=(10, 8)):
        header = ctk.CTkFrame(parent, corner_radius=0, fg_color=("#2563EB", "#1E40AF"))
        header.pack(fill="x", pady=(0, 10))

        area_titulo = ctk.CTkFrame(header, fg_color="transparent")
        area_titulo.pack(side="left", fill="y", padx=padding[0], pady=padding[1])
        ctk.CTkLabel(
            area_titulo,
            text=titulo,
            font=("Segoe UI", 16, "bold"),
            text_color="white",
            anchor="w"
        ).pack(anchor="w")

        if self.usuario:
            ctk.CTkLabel(
                area_titulo,
                text=f"Usuário: {self._nome_usuario_logado()}",
                font=("Segoe UI", 9),
                text_color="white",
                anchor="w"
            ).pack(anchor="w")

        area_acoes = ctk.CTkFrame(header, fg_color="transparent")
        area_acoes.pack(side="right", fill="y", padx=padding[0], pady=padding[1])

        if voltar:
            self._comando_voltar_atual = voltar
            ctk.CTkButton(
                area_acoes,
                text=texto_cancelar,
                width=110,
                height=34,
                fg_color="#F59E0B",
                hover_color="#D97706",
                text_color="white",
                corner_radius=8,
                font=("Segoe UI", 11, "bold"),
                command=voltar
            ).pack(side="left", padx=(6, 0))
        else:
            self._comando_voltar_atual = None

        if mostrar_sair and self.usuario:
            ctk.CTkButton(
                area_acoes,
                text="Reiniciar",
                width=110,
                height=34,
                fg_color="#64748B",
                hover_color="#475569",
                text_color="white",
                corner_radius=8,
                font=("Segoe UI", 11, "bold"),
                command=self._reiniciar_sistema
            ).pack(side="left", padx=(0, 6))
            ctk.CTkButton(
                area_acoes,
                text="Sair",
                width=90,
                height=34,
                fg_color="#E53935",
                hover_color="#C62828",
                text_color="white",
                corner_radius=8,
                font=("Segoe UI", 11, "bold"),
                command=self._voltar_para_login
            ).pack(side="left", padx=6)

        return header

    # Cria cabeçalho login
    def _criar_cabecalho_login(self, parent, titulo, subtitulo=None):
        header = ctk.CTkFrame(parent, corner_radius=0, fg_color=("#2563EB", "#1E40AF"), height=72)
        header.pack(fill="x", pady=(0, 18))
        header.pack_propagate(False)

        area_titulo = ctk.CTkFrame(header, fg_color="transparent")
        area_titulo.pack(side="left", fill="y", padx=18)
        ctk.CTkLabel(
            area_titulo,
            text=titulo,
            font=("Segoe UI", 18, "bold"),
            text_color="white",
            anchor="w"
        ).pack(anchor="w", pady=(12, 0))
        if subtitulo:
            ctk.CTkLabel(
                area_titulo,
                text=subtitulo,
                font=("Segoe UI", 10),
                text_color="white",
                anchor="w"
            ).pack(anchor="w")

        ctk.CTkButton(
            header,
            text="Sair",
            width=90,
            height=34,
            fg_color="#E53935",
            hover_color="#C62828",
            text_color="white",
            corner_radius=8,
            font=("Segoe UI", 11, "bold"),
            command=self.destroy
        ).pack(side="right", padx=18, pady=12)

        return header

    # Cria título pagina
    def _criar_titulo_pagina(self, parent, titulo, subtitulo=None, pady=(0, 12)):
        ctk.CTkLabel(
            parent,
            text=titulo,
            font=("Segoe UI", 16, "bold"),
            text_color=("#0F172A", "#E2E8F0")
        ).pack(anchor="w", pady=pady)
        if subtitulo:
            sub_pady = (0, pady[1]) if isinstance(pady, tuple) else (0, 12)
            ctk.CTkLabel(
                parent,
                text=subtitulo,
                font=("Segoe UI", 10),
                text_color=("#64748B", "#94A3B8")
            ).pack(anchor="w", pady=sub_pady)

    # Cria card
    def _criar_card(self, parent, padding=12, bootstyle="light"):
        card = tb.Frame(parent, padding=padding, bootstyle=bootstyle)
        card.pack(fill="x", pady=8)
        return card

    # Lista documentos externos
    def _listar_documentos_externos(self):
        if not os.path.exists(PASTA_DOC):
            return []

        return [
            os.path.join(PASTA_DOC, nome)
            for nome in os.listdir(PASTA_DOC)
            if os.path.isfile(os.path.join(PASTA_DOC, nome))
        ]

    # Ícone documento
    def _icone_documento(self, caminho):
        ext = os.path.splitext(caminho)[1].lower().replace(".", "") or "arq"
        ext = ext[:3].upper()
        cores = {
            "PDF": "#dc3545",
            "DOC": "#286fb3",
            "XLS": "#198754",
            "TXT": "#6c757d",
            "RTF": "#6c757d",
            "PNG": "#6f42c1",
            "JPG": "#6f42c1",
            "JPE": "#6f42c1",
            "JPEG": "#6f42c1",
            "GIF": "#6f42c1",
            "BMP": "#6f42c1",
        }
        cor = cores.get(ext, "#0d6efd")

        imagem = Image.new("RGBA", (96, 96), (255, 255, 255, 0))
        desenho = ImageDraw.Draw(imagem)
        desenho.rounded_rectangle((8, 8, 88, 88), radius=18, fill=cor)
        desenho.line((68, 8, 88, 28), fill=(255, 255, 255, 220), width=5)
        desenho.line((68, 28, 88, 28), fill=(255, 255, 255, 220), width=5)
        texto = Image.new("RGBA", (96, 96), (255, 255, 255, 0))
        desenho_texto = ImageDraw.Draw(texto)
        desenho_texto.text((12, 48), ext, fill="white")
        imagem = Image.alpha_composite(imagem, texto)
        imagem = imagem.resize((72, 72), Image.LANCZOS)
        return ImageTk.PhotoImage(imagem)

    # Abre documento externo
    def _abrir_documento_externo(self, caminho):
        try:
            os.startfile(caminho)
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível abrir o documento.\n\n{e}")

    # Imprime documento externo
    def _imprimir_documento_externo(self, caminho):
        try:
            os.startfile(caminho, "print")
            messagebox.showinfo("Impressão", "Documento enviado para a impressora padrão.")
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível imprimir o documento.\n\n{e}")

    # Abre documentos externos
    def _open_documentos_externos(self):
        top = self._open_inline_container("Documentos Externos", voltar=self._build_home_screen)
        top.title("Documentos Externos")
        top.geometry("900x620")
        top.grab_set()

        wrapper = tb.Frame(top, padding=12)
        wrapper.pack(fill="both", expand=True)

        self._criar_titulo_pagina(
            wrapper,
            "Documentos Externos",
            subtitulo="Arquivos em pdf."
        )

        canvas = tk.Canvas(wrapper, highlightthickness=0)
        scrollbar = tb.Scrollbar(wrapper, orient="vertical", command=canvas.yview)
        conteudo = tb.Frame(canvas)

        conteudo.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=conteudo, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        arquivos = self._listar_documentos_externos()
        self._icones_documentos = []

        if not arquivos:
            tb.Label(
                conteudo,
                text="Nenhum documento encontrado na pasta DOC.",
                font=("Segoe UI", 11)
            ).pack(pady=20)
            return

        for coluna in range(4):
            conteudo.columnconfigure(coluna, weight=1)

        for indice, caminho in enumerate(sorted(arquivos, key=lambda p: os.path.basename(p).casefold())):
            linha = indice // 4
            coluna = indice % 4

            card = tb.Frame(conteudo, padding=12, bootstyle="light")
            card.grid(row=linha, column=coluna, padx=8, pady=8, sticky="nsew")

            try:
                foto = self._icone_documento(caminho)
                self._icones_documentos.append(foto)
                tb.Label(card, image=foto).pack(anchor="center")
            except Exception:
                tb.Label(card, text="DOC", font=("Segoe UI", 18, "bold"), bootstyle="secondary").pack(anchor="center", pady=8)

            nome = os.path.basename(caminho)
            tb.Label(
                card,
                text=nome,
                font=("Segoe UI", 10, "bold"),
                wraplength=150,
                justify="center"
            ).pack(pady=(8, 10))

            botoes = tb.Frame(card)
            botoes.pack(fill="x")
            botoes.columnconfigure(0, weight=1)
            botoes.columnconfigure(1, weight=1)

            tb.Button(
                botoes,
                text="Abrir",
                bootstyle=INFO,
                command=lambda p=caminho: self._abrir_documento_externo(p)
            ).grid(row=0, column=0, padx=(0, 4), sticky="ew")

            tb.Button(
                botoes,
                text="Imprimir",
                bootstyle=SUCCESS,
                command=lambda p=caminho: self._imprimir_documento_externo(p)
            ).grid(row=0, column=1, padx=(4, 0), sticky="ew")

    # Nome usuário logado
    def _nome_usuario_logado(self):
        dados = self.db.get("usuarios", {}).get(self.usuario or "", {})
        return dados.get("nome") or mask_cpf_from_clean(self.usuario or "")

    # Forcar maiusculo campo
    def _forcar_maiusculo_entry(self, entry):
        texto = entry.get()
        texto_maiusculo = texto.upper()
        if texto != texto_maiusculo:
            pos = entry.index(tk.INSERT)
            entry.delete(0, tk.END)
            entry.insert(0, texto_maiusculo)
            entry.icursor(min(pos, len(texto_maiusculo)))

    # Configura rolagem tabela
    def _configurar_rolagem_tabela(self, tree):
        tree.configure(takefocus=True)

        scroll_y = tb.Scrollbar(
            tree.master,
            orient="vertical",
            command=tree.yview
        )

        scroll_x = tb.Scrollbar(
            tree.master,
            orient="horizontal",
            command=tree.xview
        )

        tree.configure(
            yscrollcommand=scroll_y.set,
            xscrollcommand=scroll_x.set
        )

        scroll_y.place(
            in_=tree,
            relx=1.0,
            rely=0,
            relheight=1.0,
            anchor="ne"
        )

        scroll_x.place(
            in_=tree.master,
            relx=0,
            rely=1.0,
            relwidth=1.0,
            y=0,
            height=12
        )

        tree._scrollbars = (scroll_y, scroll_x)

    # Automático ajusta colunas tabela
    def _auto_ajustar_colunas_tabela(self, tree, minimo=70, maximo=1000):
        try:
            import tkinter.font as tkfont
            fonte = tkfont.nametofont("TkDefaultFont")
            colunas = list(tree["columns"])
            itens = tree.get_children("")

            for indice, coluna in enumerate(colunas):
                titulo = tree.heading(coluna, "text") or coluna
                largura = fonte.measure(str(titulo)) + 32

                for item in itens:
                    valores = tree.item(item, "values")
                    if indice < len(valores):
                        largura = max(largura, fonte.measure(str(valores[indice])) + 32)

                largura = max(minimo, min(maximo, largura))
                tree.column(coluna, width=largura, minwidth=minimo, stretch=False)
        except Exception:
            pass

    # Ativar automático ajuste tabela
    def _ativar_auto_ajuste_tabela(self, tree, minimo=70, maximo=520):
        if getattr(tree, "_auto_ajuste_ativo", False):
            return

        tree._auto_ajuste_ativo = True
        tree._auto_ajuste_minimo = minimo
        tree._auto_ajuste_maximo = maximo
        original_insert = tree.insert
        original_delete = tree.delete
        original_item = tree.item

        # Agenda ajuste
        def agendar_ajuste():
            if getattr(tree, "_auto_ajuste_pausado", False):
                return
            tree.after_idle(
                lambda: self._auto_ajustar_colunas_tabela(
                    tree,
                    getattr(tree, "_auto_ajuste_minimo", minimo),
                    getattr(tree, "_auto_ajuste_maximo", maximo)
                )
            )

        # Insere automático
        def insert_auto(*args, **kwargs):
            item = original_insert(*args, **kwargs)
            agendar_ajuste()
            return item

        # Exclui automático
        def delete_auto(*args, **kwargs):
            resultado = original_delete(*args, **kwargs)
            agendar_ajuste()
            return resultado

        # Item automático
        def item_auto(*args, **kwargs):
            resultado = original_item(*args, **kwargs)
            if kwargs:
                agendar_ajuste()
            return resultado

        tree.insert = insert_auto
        tree.delete = delete_auto
        tree.item = item_auto
        tree.bind("<Configure>", lambda e: agendar_ajuste(), add="+")
        agendar_ajuste()

    # Normaliza texto ordenação
    def _normalizar_texto_ordenacao(self, valor):
        return normalizar_texto_ordenacao(valor)

    # Valor ordenação cadastro
    def _valor_ordenacao_cadastro(self, coluna, valor):
        texto = str(valor or "").strip()

        if coluna == "ordem":
            try:
                return (0, int(somente_digs(texto) or "0"))
            except ValueError:
                return (1, self._normalizar_texto_ordenacao(texto))

        if coluna in ("protocolo", "cpf", "telefone"):
            digitos = somente_digs(texto)
            if digitos:
                return (0, int(digitos), self._normalizar_texto_ordenacao(texto))
            return (1, self._normalizar_texto_ordenacao(texto))

        if coluna == "data":
            try:
                return (0, datetime.strptime(texto, "%d/%m/%Y"))
            except ValueError:
                return (1, self._normalizar_texto_ordenacao(texto))

        return (0, self._normalizar_texto_ordenacao(texto))

    # Ordena tabela cadastro
    def _ordenar_tabela_cadastro(self, coluna):
        if not hasattr(self, "tree") or not self.tree.winfo_exists():
            return

        colunas = list(self.tree["columns"])
        if coluna not in colunas:
            return

        indice = colunas.index(coluna)
        reverso = self._cadastro_ordenacao_reversa.get(coluna, False)
        itens = []

        for item in self.tree.get_children(""):
            valores = self.tree.item(item, "values")
            valor = valores[indice] if indice < len(valores) else ""
            itens.append((self._valor_ordenacao_cadastro(coluna, valor), item))

        itens.sort(key=lambda item: item[0], reverse=reverso)

        for posicao, item in enumerate([item for _, item in itens]):
            self.tree.move(item, "", posicao)

        self._cadastro_ordenacao_reversa[coluna] = not reverso

    # Prepara tabela
    def _preparar_tabela(self, tree, minimo=70, maximo=520, auto_ajuste=True):
        self._configurar_rolagem_tabela(tree)
        if auto_ajuste:
            self._ativar_auto_ajuste_tabela(tree, minimo=minimo, maximo=maximo)

    # Pausa automático ajuste tabela
    def _pausar_auto_ajuste_tabela(self, tree):
        tree._auto_ajuste_pausado = True

    # Retoma automático ajuste tabela
    def _retomar_auto_ajuste_tabela(self, tree):
        tree._auto_ajuste_pausado = False
        if getattr(tree, "_auto_ajuste_ativo", False):
            self._auto_ajustar_colunas_tabela(
                tree,
                getattr(tree, "_auto_ajuste_minimo", 70),
                getattr(tree, "_auto_ajuste_maximo", 520)
            )

    # Garante usuário teste
    def _garantir_usuario_teste(self):
        usuario_teste_cpf = "00000000000"
        if usuario_teste_cpf not in self.usuarios:
            senha_hash = hash_password("123456")
            self.usuarios[usuario_teste_cpf] = {
                "nome": "Usuário Teste",
                "cpf": usuario_teste_cpf,
                "perfil": "admin",
                "ativo": True,
                "senha_expirada": False,
                "ultima_troca_senha": datetime.now().strftime("%d/%m/%Y"),
                "tentativas_login": 0,
                "tentativas_recuperacao": 0,
                "bloqueado_ate": "",
                "vans_permitidas": [],
                "van_ativa": "",
                **senha_hash
            }
            save_usuarios(self.usuarios)

    # Constrói a tela login tela
    def _build_login_screen(self):
        self._clear_screen()
        self._clear()

        modo_teste_ativo = _verificar_modo_teste_ativo()
        aviso_expiracao = None
        if modo_teste_ativo:
            dados_mt = _carregar_modo_teste()
            try:
                expira_em = datetime.strptime(dados_mt.get("expira_em", ""), "%d/%m/%Y %H:%M:%S")
                restante = expira_em - datetime.now()
                minutos = restante.total_seconds() / 60
                if minutos <= 2:
                    aviso_expiracao = f"Modo Teste expira em {max(0, int(minutos))} minuto(s)!"
            except Exception:
                pass

        if os.path.exists(ARQUIVO_SESSAO):
            try:
                with open(ARQUIVO_SESSAO, "r", encoding="utf-8") as f:
                    sessao = json.load(f)
                usuario_sessao = sessao.get("usuario", "")
                if usuario_sessao:
                    try:
                        os.remove(ARQUIVO_SESSAO)
                    except Exception:
                        pass
            except Exception:
                usuario_sessao = ""
        else:
            usuario_sessao = ""

        container = ctk.CTkFrame(self, fg_color="transparent")
        container.pack(fill="both", expand=True, padx=24, pady=24)

        self._criar_cabecalho_login(container, "VAN CONECTA ITINERANTE", "EMPREL")

        card = ctk.CTkFrame(
            container,
            corner_radius=16,
            border_width=1,
            border_color=("#E2E8F0", "#334155"),
            fg_color=("#FFFFFF", "#1E293B")
        )
        card.pack(padx=20, pady=10)

        ctk.CTkLabel(
            card,
            text="Faça login para acessar o sistema",
            font=("Segoe UI", 14, "bold"),
            text_color=("#0F172A", "#E2E8F0")
        ).pack(pady=(18, 12), padx=24)

        if aviso_expiracao:
            ctk.CTkLabel(
                card,
                text=aviso_expiracao,
                font=("Segoe UI", 11, "bold"),
                text_color="#DC2626",
                anchor="w"
            ).pack(anchor="w", padx=24, pady=(0, 8))

        if modo_teste_ativo:
            self._garantir_usuario_teste()
            self.usuario = "00000000000"
            self.perfil = "admin"
            self._registrar_log("Login modo teste automático")
            self.after(100, self._iniciar_monitoramento_modo_teste)
            self.after(150, self._build_home_screen)
            return

        ctk.CTkLabel(
            card,
            text="Login (CPF):",
            font=("Segoe UI", 11, "bold"),
            anchor="w",
            text_color=("#334155", "#CBD5E1")
        ).pack(anchor="w", padx=24)

        self.login_user = ctk.CTkEntry(
            card,
            width=300,
            height=38,
            placeholder_text="000.000.000-00",
            font=("Segoe UI", 12)
        )
        self.login_user.pack(fill="x", padx=24, pady=(6, 12))
        if usuario_sessao:
            self.login_user.insert(0, usuario_sessao)

        # Mascara login e move foco
        def _mascara_login_e_move_foco(event=None):
            mascara_cpf_entrada(self.login_user)
            if len(somente_digs(self.login_user.get())) >= 11:
                self.login_pass.focus_set()

        self.login_user.bind("<KeyRelease>", lambda e: _mascara_login_e_move_foco())
        self.login_user.bind("<Return>", lambda e: _mascara_login_e_move_foco())

        ctk.CTkLabel(
            card,
            text="Senha:",
            font=("Segoe UI", 11, "bold"),
            anchor="w",
            text_color=("#334155", "#CBD5E1")
        ).pack(anchor="w", padx=24)

        self.login_pass = ctk.CTkEntry(
            card,
            width=300,
            height=38,
            show="*",
            font=("Segoe UI", 12)
        )
        self.login_pass.pack(fill="x", padx=24, pady=(6, 10))
        self.login_pass.bind("<Return>", lambda e: self._entrar())

        ctk.CTkButton(
            card,
            text="Recuperar senha",
            command=self._recuperar_senha_offline,
            fg_color="transparent",
            text_color=("#2563EB", "#60A5FA"),
            hover_color=("#EFF6FF", "#1E3A8A"),
            anchor="w",
            width=200,
            height=28,
            font=("Segoe UI", 11)
        ).pack(anchor="w", padx=24, pady=(0, 12))

        ctk.CTkButton(
            card,
            text="Entrar",
            command=self._entrar,
            fg_color=("#2563EB", "#2563EB"),
            hover_color=("#1D4ED8", "#1D4ED8"),
            text_color="white",
            corner_radius=10,
            height=42,
            font=("Segoe UI", 13, "bold")
        ).pack(fill="x", padx=24, pady=(0, 20))

        self._ajustar_janela_responsiva(360, 420)

        self.after_idle(self.login_user.focus_set)

    # Autoriza administrador
    def _autorizar_admin(self):

        if self.perfil == "admin":
            return True

        senha_admin = simpledialog.askstring(
            "Autorização",
            "Digite a senha do administrador:",
            show="*"
        )

        if not senha_admin:
            return False

        for cpf, dados in self.usuarios.items():

            if (
                dados.get("perfil") == "admin"
                and verify_password(dados, senha_admin)
            ):
                return True

        messagebox.showerror(
            "Acesso Negado",
            "Senha de administrador inválida."
        )

        return False

    # Recupera senha offline
    def _recuperar_senha_offline(self):
        from tkinter import simpledialog

        top = self._open_inline_container("Recuperar Senha", voltar=self._build_login_screen)
        top.title("Recuperar Senha")
        top.geometry("400x280")
        top.grab_set()

        frm = ctk.CTkFrame(top, fg_color="transparent")
        frm.pack(fill="both", expand=True, padx=15, pady=15)

        ctk.CTkLabel(frm, text="CPF do usuário:", text_color=("#334155", "#CBD5E1")).pack(anchor="w")
        ent_cpf = ctk.CTkEntry(frm)
        ent_cpf.pack(fill="x", pady=5)
        ent_cpf.bind("<KeyRelease>", lambda e: mascara_cpf_entrada(ent_cpf))

        ctk.CTkLabel(frm, text="Data de Nascimento (DD/MM/AAAA):", text_color=("#334155", "#CBD5E1")).pack(anchor="w")
        ent_dn = ctk.CTkEntry(frm)
        ent_dn.pack(fill="x", pady=5)
        ent_dn.bind("<KeyRelease>", lambda e: mascara_data_entrada(ent_dn))

        ctk.CTkLabel(frm, text="Nova senha:", text_color=("#334155", "#CBD5E1")).pack(anchor="w")
        ent_senha = ctk.CTkEntry(frm, show="*")
        ent_senha.pack(fill="x", pady=5)

        # Redefine
        def redefinir():
            cpf = somente_digs(ent_cpf.get())
            dn = ent_dn.get().strip()
            nova = ent_senha.get().strip()

            if not validar_cpf(cpf):
                messagebox.showerror("Erro", "CPF inválido.")
                return

            if not validar_data_nascimento(dn):
                messagebox.showerror("Erro", "Data de nascimento inválida (DD/MM/AAAA).")
                return

            if cpf not in self.usuarios:
                messagebox.showerror("Erro", "Usuário não encontrado.")
                return

            dados = self.usuarios[cpf]
            if dados.get("perfil") == "admin":
                messagebox.showerror("Erro", "Recuperação de senha de administrador não é permitida por este meio. Procure outro administrador.")
                return
            if dados.get("data_nascimento") != dn:
                dados["tentativas_recuperacao"] = dados.get("tentativas_recuperacao", 0) + 1
                save_usuarios(self.usuarios)
                messagebox.showerror("Erro", "Data de nascimento não confere.")
                return

            if dados.get("bloqueado_ate"):
                try:
                    bloqueio = datetime.strptime(dados["bloqueado_ate"], "%d/%m/%Y %H:%M:%S")
                    if datetime.now() < bloqueio:
                        messagebox.showerror("Erro", "Recuperação bloqueada. Procure o administrador.")
                        return
                except Exception:
                    pass

            tentativas = dados.get("tentativas_recuperacao", 0)
            if tentativas >= 3:
                dados["bloqueado_ate"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                save_usuarios(self.usuarios)
                messagebox.showerror("Erro", "Número máximo de tentativas atingido. Recuperação bloqueada. Procure o administrador.")
                return

            if not nova:
                messagebox.showerror("Erro", "Digite a nova senha.")
                return

            self.usuarios[cpf].update(hash_password(nova))
            self.usuarios[cpf]["ultima_troca_senha"] = datetime.now().strftime("%d/%m/%Y")
            self.usuarios[cpf]["senha_expirada"] = False
            self.usuarios[cpf]["tentativas_recuperacao"] = 0
            self.usuarios[cpf]["bloqueado_ate"] = ""
            save_usuarios(self.usuarios)

            messagebox.showinfo("Sucesso", "Senha redefinida com sucesso.")
            self._fechar_tela_interna(top)

        ctk.CTkButton(frm, text="Redefinir Senha", command=redefinir, fg_color="#16A34A", hover_color="#15803D", text_color="white", corner_radius=8, height=38, font=("Segoe UI", 12, "bold")).pack(pady=10)

    # Realiza o login
    def _entrar(self):
        user_mask = self.login_user.get().strip()
        pwd = self.login_pass.get()

        if not validar_cpf(user_mask):
            messagebox.showerror("Erro", "O login deve ser um CPF válido.")
            return

        user = somente_digs(user_mask)

        if not self.usuarios:
            self._open_criar_usuario()
            return

    # Realiza o login
    def _entrar(self):
        user_mask = self.login_user.get().strip()
        pwd = self.login_pass.get()

        if not validar_cpf(user_mask):
            messagebox.showerror("Erro", "O login deve ser um CPF válido.")
            return

        user = somente_digs(user_mask)

        if not self.usuarios:
            self._open_criar_usuario()
            return

        stored = self.usuarios.get(user)
        if stored:
            if stored.get("perfil") == "admin":
                if verify_password(stored, pwd):
                    self.usuario = user
                    self.perfil = "admin"
                    self._registrar_log("Login admin")
                    self._build_home_screen()
                else:
                    messagebox.showerror("Erro", "Usuário ou senha inválidos.")
                return
            if stored.get("bloqueado_ate"):
                try:
                    bloqueio = datetime.strptime(stored["bloqueado_ate"], "%d/%m/%Y %H:%M:%S")
                    if datetime.now() < bloqueio:
                        messagebox.showerror("Erro", "Login bloqueado. Procure o administrador.")
                        return
                except Exception:
                    pass
            if verify_password(stored, pwd):
                if not stored.get("ativo", True):
                    messagebox.showerror("Erro", "Usuário desativado. Procure o administrador.")
                    return
                if stored.get("senha_expirada", True) or senha_expirada_60_dias(stored):
                    if stored.get("senha_expirada", True):
                        pass
                    else:
                        stored["senha_expirada"] = True
                        save_usuarios(self.usuarios)
                    self._abrir_troca_senha_obrigatoria(user)
                    return
                self.usuario = user
                self.perfil = stored.get("perfil", "usuario")
                self._registrar_log("Login")
                self._build_home_screen()
            else:
                tentativas = stored.get("tentativas_login", 0) + 1
                stored["tentativas_login"] = tentativas
                if tentativas >= 3:
                    stored["bloqueado_ate"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                    save_usuarios(self.usuarios)
                    messagebox.showerror("Erro", "Número máximo de tentativas atingido. Login bloqueado. Procure o administrador.")
                    return
                save_usuarios(self.usuarios)
                messagebox.showerror("Erro", f"Usuário ou senha inválidos. Tentativa {tentativas}/3.")
        else:
            messagebox.showerror("Erro", "Usuário ou senha inválidos.")

        if os.path.exists(ARQUIVO_SESSAO):
            try:
                os.remove(ARQUIVO_SESSAO)
            except Exception:
                pass

    # Abre troca senha obrigatoria
    def _abrir_troca_senha_obrigatoria(self, cpf):
        dados = self.usuarios.get(cpf)
        if not dados:
            return
        top = self._open_inline_container("Cadastro Obrigatório", voltar=self._build_login_screen)
        top.title("Cadastro Obrigatório")
        top.geometry("480x360")
        top.grab_set()
        frm = ctk.CTkFrame(top, fg_color="transparent")
        frm.pack(fill="both", expand=True, padx=15, pady=15)

        if dados.get("senha_expirada", True):
            ctk.CTkLabel(frm, text="Primeiro acesso ou senha expirada.\nComplete seus dados e cadastre uma nova senha.", font=("Segoe UI", 11, "bold"), text_color=("#334155", "#CBD5E1")).pack(pady=(0, 10))
        else:
            ctk.CTkLabel(frm, text="Complete seus dados e cadastre uma nova senha.", font=("Segoe UI", 11, "bold"), text_color=("#334155", "#CBD5E1")).pack(pady=(0, 10))

        ctk.CTkLabel(frm, text="Data de Nascimento (DD/MM/AAAA):", text_color=("#334155", "#CBD5E1")).pack(anchor="w")
        ent_dn = ctk.CTkEntry(frm)
        ent_dn.pack(fill="x", pady=5)
        ent_dn.bind("<KeyRelease>", lambda e: mascara_data_entrada(ent_dn))

        ctk.CTkLabel(frm, text="E-mail:", text_color=("#334155", "#CBD5E1")).pack(anchor="w")
        ent_email = ctk.CTkEntry(frm)
        ent_email.pack(fill="x", pady=5)

        ctk.CTkLabel(frm, text="Nova senha:", text_color=("#334155", "#CBD5E1")).pack(anchor="w")
        ent1 = ctk.CTkEntry(frm, show="*")
        ent1.pack(fill="x", pady=5)
        ctk.CTkLabel(frm, text="Confirmar nova senha:", text_color=("#334155", "#CBD5E1")).pack(anchor="w")
        ent2 = ctk.CTkEntry(frm, show="*")
        ent2.pack(fill="x", pady=5)

        # Confirma
        def confirmar():
            dn = ent_dn.get().strip()
            email = ent_email.get().strip()
            p1 = ent1.get().strip()
            p2 = ent2.get().strip()
            if not validar_data_nascimento(dn):
                messagebox.showerror("Erro", "Informe uma data de nascimento válida (DD/MM/AAAA).")
                return
            if not validar_email(email):
                messagebox.showerror("Erro", "Informe um e-mail válido.")
                return
            if not p1 or len(p1) < 4:
                messagebox.showerror("Erro", "A senha deve ter ao menos 4 caracteres.")
                return
            if p1 != p2:
                messagebox.showerror("Erro", "As senhas não coincidem.")
                return
            self.usuarios[cpf]["data_nascimento"] = dn
            self.usuarios[cpf]["email"] = email
            self.usuarios[cpf].update(hash_password(p1))
            self.usuarios[cpf]["ultima_troca_senha"] = datetime.now().strftime("%d/%m/%Y")
            self.usuarios[cpf]["senha_expirada"] = False
            self.usuarios[cpf]["tentativas_recuperacao"] = 0
            self.usuarios[cpf]["tentativas_login"] = 0
            self.usuarios[cpf]["bloqueado_ate"] = ""
            save_usuarios(self.usuarios)
            self._registrar_log("Forçou troca de senha + cadastro", cpf)
            messagebox.showinfo("OK", "Dados cadastrados e senha alterada com sucesso.")
            self._fechar_tela_interna(top)
            self.login_pass.delete(0, tk.END)
            self._entrar()

        ctk.CTkButton(frm, text="Salvar e Entrar", command=confirmar, fg_color="#16A34A", hover_color="#15803D", text_color="white", corner_radius=8, height=38, font=("Segoe UI", 12, "bold")).pack(pady=10)

    # Abre cria usuário
    def _open_criar_usuario(self):
        usuarios_existem = bool(self.usuarios)
        if usuarios_existem and self.perfil != "admin":
            messagebox.showerror("Acesso negado", "Somente administradores podem criar usuários.")
            return

        top = self._open_inline_container("Criar usuário", voltar=self._build_login_screen if not self.usuario else self._open_acesso_screen)
        top.title("Criar usuário")
        top.geometry("480x420")
        top.grab_set()

        frm = ctk.CTkFrame(top, fg_color="transparent")
        frm.pack(fill="both", expand=True, padx=12, pady=12)

        ctk.CTkLabel(frm, text="Criar novo usuário", font=("Segoe UI", 12, "bold"), text_color=("#0F172A", "#E2E8F0")).pack(pady=8)

        ctk.CTkLabel(frm, text="Usuário (CPF):", text_color=("#334155", "#CBD5E1")).pack(anchor="w")
        ent_user = ctk.CTkEntry(frm)
        ent_user.pack(fill="x", pady=4)
        ent_user.bind("<KeyRelease>", lambda e: mascara_cpf_entrada(ent_user))

        ctk.CTkLabel(frm, text="Nome:", text_color=("#334155", "#CBD5E1")).pack(anchor="w")
        ent_nome = ctk.CTkEntry(frm)
        ent_nome.pack(fill="x", pady=4)
        ent_nome.bind("<KeyRelease>", lambda e: self._forcar_maiusculo_entry(ent_nome))

        ctk.CTkLabel(frm, text="Data de Nascimento (DD/MM/AAAA):", text_color=("#334155", "#CBD5E1")).pack(anchor="w")
        ent_dn = ctk.CTkEntry(frm)
        ent_dn.pack(fill="x", pady=4)
        ent_dn.bind("<KeyRelease>", lambda e: mascara_data_entrada(ent_dn))

        ctk.CTkLabel(frm, text="E-mail:", text_color=("#334155", "#CBD5E1")).pack(anchor="w")
        ent_email = ctk.CTkEntry(frm)
        ent_email.pack(fill="x", pady=4)

        ctk.CTkLabel(frm, text="Senha:", text_color=("#334155", "#CBD5E1")).pack(anchor="w")
        ent_pwd = ctk.CTkEntry(frm, show="*")
        ent_pwd.pack(fill="x", pady=4)

        ctk.CTkLabel(frm, text="Confirmar Senha:", text_color=("#334155", "#CBD5E1")).pack(anchor="w")
        ent_pwd2 = ctk.CTkEntry(frm, show="*")
        ent_pwd2.pack(fill="x", pady=4)

        if not usuarios_existem:
            ctk.CTkLabel(frm, text="Primeiro usuário será ADMIN", text_color="#D97706", font=("Segoe UI", 11, "bold")).pack(pady=6)
            perfil_combo = None
            listbox_vans = None
            combo_van_ativa = None
        else:
            ctk.CTkLabel(frm, text="Perfil:", text_color=("#334155", "#CBD5E1")).pack(anchor="w")
            perfil_combo = tb.Combobox(frm, values=["admin", "usuario"])
            perfil_combo.current(1)
            perfil_combo.pack(fill="x", pady=4)

            ctk.CTkLabel(frm, text="Vans Permitidas (Ctrl+Clique para múltiplos):", text_color=("#334155", "#CBD5E1")).pack(anchor="w", pady=(8, 0))
            listbox_vans = tk.Listbox(frm, selectmode="extended", height=4, exportselection=False)
            listbox_vans.pack(fill="x", pady=4)
            vans = load_vans()
            vans_por_nome = {}
            for v in vans:
                if v.get("ativa", True):
                    listbox_vans.insert("end", v.get("nome", ""))
                    vans_por_nome[v.get("nome", "")] = v.get("id", "")

            ctk.CTkLabel(frm, text="Van Ativa:", text_color=("#334155", "#CBD5E1")).pack(anchor="w", pady=(8, 0))
            combo_van_ativa = tb.Combobox(frm, values=[], state="readonly")
            combo_van_ativa.pack(fill="x", pady=4)

            def atualizar_combo_van_ativa(event=None):
                sel = listbox_vans.curselection()
                vans_selecionadas = [listbox_vans.get(i) for i in sel]
                combo_van_ativa.configure(values=vans_selecionadas)
                if vans_selecionadas:
                    combo_van_ativa.set(vans_selecionadas[0])
                else:
                    combo_van_ativa.set("")

            listbox_vans.bind("<<ListboxSelect>>", atualizar_combo_van_ativa)
            atualizar_combo_van_ativa()

        # Salva
        def salvar():
            u_mask = ent_user.get().strip()
            u = somente_digs(u_mask)
            if not validar_cpf(u):
                messagebox.showerror("Erro", "O usuário deve ser um CPF válido.")
                return
            nome = ent_nome.get().strip().upper()
            if not nome:
                messagebox.showerror("Erro", "Informe o nome do usuário.")
                return
            dn = ent_dn.get().strip()
            if not validar_data_nascimento(dn):
                messagebox.showerror("Erro", "Informe uma data de nascimento válida (DD/MM/AAAA).")
                return
            email = ent_email.get().strip()
            if not validar_email(email):
                messagebox.showerror("Erro", "Informe um e-mail válido.")
                return
            p1 = ent_pwd.get()
            p2 = ent_pwd2.get()
            if not u or not p1:
                messagebox.showerror("Erro", "Preencha todos os campos.")
                return
            if p1 != p2:
                messagebox.showerror("Erro", "As senhas não coincidem.")
                return
            if u in self.usuarios:
                messagebox.showerror("Erro", "Usuário já existe.")
                return
            perfil = "admin" if not usuarios_existem else perfil_combo.get()
            vans_permitidas = []
            van_ativa = ""
            if listbox_vans:
                sel = listbox_vans.curselection()
                for idx in sel:
                    nome_van = listbox_vans.get(idx)
                    vid = vans_por_nome.get(nome_van, "")
                    if vid:
                        vans_permitidas.append(vid)
            if combo_van_ativa:
                van_ativa = vans_por_nome.get(combo_van_ativa.get().strip(), "")
            self.usuarios[u] = {
                **hash_password(p1),
                "perfil": perfil,
                "nome": nome,
                "data_nascimento": dn,
                "email": email,
                "ultima_troca_senha": datetime.now().strftime("%d/%m/%Y"),
                "senha_expirada": True,
                "tentativas_recuperacao": 0,
                "bloqueado_ate": "",
                "ativo": True,
                "vans_permitidas": vans_permitidas,
                "van_ativa": van_ativa
            }
            save_usuarios(self.usuarios)
            self._registrar_log("Criou usuário", u)
            messagebox.showinfo("OK", "Usuário criado com sucesso.")
            self._fechar_tela_interna(top)
            if hasattr(self, "tree_users") and self.tree_users.winfo_exists():
                self._carregar_usuarios()

        ctk.CTkButton(frm, text="Salvar", command=salvar, fg_color="#16A34A", hover_color="#15803D", text_color="white", corner_radius=8, height=38, font=("Segoe UI", 12, "bold")).pack(pady=10)


    # Abre exclui usuário
    def _open_excluir_usuario(self):
        if self.perfil != "admin":
            messagebox.showerror("Acesso negado", "Somente administradores podem excluir usuários.")
            return
        top = self._open_inline_container("Excluir Usuário", voltar=self._build_home_screen)
        top.title("Excluir Usuário")
        top.geometry("420x260")
        top.grab_set()

        frm = ctk.CTkFrame(top, fg_color="transparent")
        frm.pack(fill="both", expand=True, padx=12, pady=12)

        ctk.CTkLabel(frm, text="Excluir Usuário", font=("Segoe UI", 12, "bold"), text_color=("#0F172A", "#E2E8F0")).pack(pady=8)
        ctk.CTkLabel(frm, text="Selecione o usuário:", text_color=("#334155", "#CBD5E1")).pack(anchor="w")

        usuarios = list(self.db["usuarios"].keys())
        if self.usuario in usuarios:
            usuarios.remove(self.usuario)
        if not usuarios:
            messagebox.showinfo("Aviso", "Não há outros usuários para excluir.")
            self._fechar_tela_interna(top)
            return

        cb_users = ctk.CTkComboBox(frm, values=usuarios, width=300)
        cb_users.pack(fill="x", pady=6)

        # Exclui
        def excluir():
            u = cb_users.get()
            if not u:
                messagebox.showwarning("Aviso", "Selecione um usuário.")
                return
            if messagebox.askyesno("Confirmar", f"Deseja excluir o usuário '{u}'?"):
                self.db["usuarios"].pop(u, None)
                save_db(self.db)
                self._registrar_log("Excluiu usuário", u)
                messagebox.showinfo("OK", f"Usuário '{u}' excluído.")
                self._fechar_tela_interna(top)

        ctk.CTkButton(frm, text="Excluir", command=excluir, fg_color="#E53935", hover_color="#C62828", text_color="white", corner_radius=8, height=38, font=("Segoe UI", 12, "bold")).pack(pady=10)

    # Constrói a tela início tela
    def _build_home_screen(self):
        self._ajustar_janela_principal()
        self._clear_screen()
        self._home_ativa = True
        self._botoes_home = []

        main = ctk.CTkFrame(self, fg_color="transparent")
        main.pack(fill="both", expand=True, padx=10, pady=10)

        self._criar_cabecalho(
            main,
            "SERVIÇOS - VAN CONECTA ITINERANTE",
            mostrar_sair=True,
            texto_cancelar="Sair"
        )

        bloco_logo = ctk.CTkFrame(
            main,
            corner_radius=12,
            border_width=1,
            border_color=("#E2E8F0", "#334155"),
            fg_color=("#FFFFFF", "#1E293B")
        )
        bloco_logo.pack(fill="x", pady=8)
        try:
            caminho_logo = r"img/conecta_recife.png"

            img = Image.open(caminho_logo).resize((100, 50))
            self._logo_img = ctk.CTkImage(light_image=img, dark_image=img, size=(100, 50))

            ctk.CTkLabel(bloco_logo, image=self._logo_img, text="").pack(anchor="w", padx=12, pady=12)
        except Exception as e:
            print("Erro ao carregar logo:", e)

        ctk.CTkLabel(
            bloco_logo,
            text="EMPRESA MUNICIPAL DE INFORMÁTICA - EMPREL",
            font=("Segoe UI", 12, "bold"),
            text_color=("#0F172A", "#E2E8F0")
        ).pack(anchor="w", padx=12, pady=(0, 2))
        ctk.CTkLabel(
            bloco_logo,
            text="VAN CONECTA ITINERANTE",
            font=("Segoe UI", 11),
            text_color=("#334155", "#CBD5E1")
        ).pack(anchor="w", padx=12, pady=(0, 12))

        bloco_acao = ctk.CTkFrame(
            main,
            corner_radius=12,
            border_width=1,
            border_color=("#BFDBFE", "#1E3A8A"),
            fg_color=("#EFF6FF", "#172554")
        )
        bloco_acao.pack(fill="x", pady=8)
        ctk.CTkLabel(
            bloco_acao,
            text="AÇÃO ATIVA",
            font=("Segoe UI", 11, "bold"),
            text_color=("#1E3A8A", "#BFDBFE")
        ).pack(anchor="w", padx=12, pady=(8, 0))
        self._render_acao_ativa(bloco_acao)

        bloco_acoes = ctk.CTkFrame(
            main,
            corner_radius=12,
            border_width=1,
            border_color=("#E2E8F0", "#334155"),
            fg_color=("#FFFFFF", "#1E293B")
        )
        bloco_acoes.pack(fill="x", pady=8)
        ctk.CTkLabel(
            bloco_acoes,
            text="OPERAÇÕES",
            font=("Segoe UI", 11, "bold"),
            text_color=("#0F172A", "#E2E8F0")
        ).pack(anchor="w", padx=12, pady=(8, 5))

        linha_botoes = ctk.CTkFrame(bloco_acoes, fg_color="transparent")
        linha_botoes.pack(fill="x", padx=6, pady=6)

        estilos_ctk = {
            "primary": ("#2563EB", "#1D4ED8"),
            "success": ("#16A34A", "#15803D"),
            "warning": ("#D97706", "#B45309"),
        }

        botoes_home = [
            {
                "text": "Login",
                "command": self._open_acesso_screen,
                "bootstyle": "primary"
            },
            {
                "text": "Ações",
                "command": self._open_gerenciar_acoes,
                "bootstyle": "primary"
            },
            {
                "text": "Serviços",
                "command": self._open_gerenciar_servicos,
                "bootstyle": "primary"
            },
            {
                "text": "Organizações",
                "command": self._open_gerenciar_organizacoes,
                "bootstyle": "primary"
            },
            {
                "text": "Vans",
                "command": self._open_gerenciar_vans,
                "bootstyle": "primary",
                "admin_only": True
            },
            {
                "text": "Ação Ativa" if self._selected_acao and self._selected_servico else "Cadastro",
                "command": (
                    (lambda: self._build_cadastro_screen(preselect=True))
                    if self._selected_acao and self._selected_servico
                    else self._build_cadastro_screen
                ),
                "bootstyle": "success" if self._selected_acao and self._selected_servico else "primary"
            },
            {
                "text": "Entrega de RG",
                "command": self.tela_entrega_identidades,
                "bootstyle": "primary"
            },
            {
                "text": "Imprimir Fichas",
                "command": self.tela_gerar_fichas,
                "bootstyle": "primary"
            },
            {
                "text": "Externo",
                "command": self._open_documentos_externos,
                "bootstyle": "primary"
            },
            {
                "text": "Relatórios",
                "command": self._open_relatorios,
                "bootstyle": "primary"
            },
            {
                "text": "Backup",
                "command": self._fazer_backup_manual,
                "bootstyle": "warning"
            }
        ]

        for coluna in range(4):
            linha_botoes.columnconfigure(coluna, weight=1, uniform="botoes_home")

        for indice, cfg in enumerate(botoes_home):
            if cfg.get("admin_only") and self.perfil != "admin":
                continue
            fg, hv = estilos_ctk.get(cfg["bootstyle"], estilos_ctk["primary"])
            btn = ctk.CTkButton(
                linha_botoes,
                text=cfg["text"],
                command=cfg["command"],
                height=46,
                corner_radius=10,
                fg_color=fg,
                hover_color=hv,
                text_color="white",
                font=("Segoe UI", 12, "bold")
            )
            btn.grid(
                row=indice // 4,
                column=indice % 4,
                padx=6,
                pady=6,
                sticky="ew"
            )
            self._botoes_home.append(btn)

        if self.perfil == "admin":
            bloco_modo_teste = ctk.CTkFrame(
                main,
                corner_radius=12,
                border_width=1,
                border_color=("#BBF7D0", "#14532D"),
                fg_color=("#F0FDF4", "#052E16")
            )
            bloco_modo_teste.pack(fill="x", pady=8)
            ctk.CTkLabel(
                bloco_modo_teste,
                text="MODO DESENVOLVEDOR",
                font=("Segoe UI", 11, "bold"),
                text_color=("#14532D", "#BBF7D0")
            ).pack(anchor="w", padx=12, pady=(8, 0))

            linha_modo_teste = ctk.CTkFrame(bloco_modo_teste, fg_color="transparent")
            linha_modo_teste.pack(fill="x", padx=6, pady=6)

            dados_mt = _carregar_modo_teste()
            ativo = _verificar_modo_teste_ativo()
            prorrogacoes = dados_mt.get("prorrogacoes", 0) if dados_mt else 0
            max_prorrogacoes = dados_mt.get("max_prorrogacoes", 5) if dados_mt else 5

            if not ativo:
                ctk.CTkButton(
                    linha_modo_teste,
                    text="Ativar Modo Desenvolvedor (10 min)",
                    command=self._ativar_modo_teste_admin,
                    fg_color="#16A34A",
                    hover_color="#15803D",
                    text_color="white",
                    corner_radius=8,
                    height=36,
                    font=("Segoe UI", 11, "bold")
                ).pack(side="left", padx=4)
            else:
                if prorrogacoes < max_prorrogacoes:
                    ctk.CTkButton(
                        linha_modo_teste,
                        text=f"Prorrogar (+10 min) ({max_prorrogacoes - prorrogacoes} restantes)",
                        command=self._prorrogar_modo_teste_admin,
                        fg_color="#D97706",
                        hover_color="#B45309",
                        text_color="white",
                        corner_radius=8,
                        height=36,
                        font=("Segoe UI", 11, "bold")
                    ).pack(side="left", padx=4)

                ctk.CTkButton(
                    linha_modo_teste,
                    text="Desativar Modo Teste",
                    command=self._desativar_modo_teste_admin,
                    fg_color="#E53935",
                    hover_color="#C62828",
                    text_color="white",
                    corner_radius=8,
                    height=36,
                    font=("Segoe UI", 11, "bold")
                ).pack(side="left", padx=4)

            lbl_info = ctk.CTkLabel(
                bloco_modo_teste,
                text="",
                font=("Segoe UI", 10),
                text_color=("#334155", "#CBD5E1")
            )
            lbl_info.pack(anchor="w", padx=12, pady=(0, 8))
            if ativo and dados_mt:
                try:
                    expira_em = datetime.strptime(dados_mt.get("expira_em", ""), "%d/%m/%Y %H:%M:%S")
                    restante = expira_em - datetime.now()
                    minutos = max(0, int(restante.total_seconds() // 60))
                    lbl_info.configure(text=f"Expira em {minutos} min | Prorrogações: {prorrogacoes}/{max_prorrogacoes}")
                except Exception:
                    pass

        if self._botoes_home and self._botoes_home[0].winfo_exists():
            self.after_idle(self._botoes_home[0].focus_set)

    # Verifica pergunta modo teste
    def _verificar_pergunta_modo_teste(self):
        usuario_teste_cpf = "00000000000"
        dados_usuario = self.usuarios.get(usuario_teste_cpf, {})
        perguntas_usadas = dados_usuario.get("perguntas_usadas", [])
        pergunta_atual_id = dados_usuario.get("pergunta_atual_id")

        caminho_perguntas = os.path.join(APP_DIR, "perguntas_modo_teste.json")
        if not os.path.exists(caminho_perguntas):
            return True

        try:
            with open(caminho_perguntas, "r", encoding="utf-8") as f:
                perguntas = json.load(f)
        except Exception:
            return True

        if not perguntas:
            return True

        if pergunta_atual_id is None:
            candidatas = []
            for p in perguntas:
                usadas = [u for u in perguntas_usadas if u.get("id") == p["id"]]
                if len(usadas) < 2:
                    candidatas.append(p)

            if not candidatas:
                messagebox.showwarning("Aviso", "Limite de prorrogações atingido. Modo Teste será desativado.")
                _desativar_modo_teste()
                self._build_home_screen()
                return False

            ultima_pergunta = perguntas_usadas[-1]["id"] if perguntas_usadas else None
            filtradas = [p for p in candidatas if p["id"] != ultima_pergunta] or candidatas
            import random
            pergunta = random.choice(filtradas)

            top = ctk.CTkToplevel(self)
            top.title("Pergunta de Segurança - Modo Teste")
            top.geometry("420x180")
            top.grab_set()
            top.transient(self)

            ctk.CTkLabel(
                top,
                text="Responda para prorrogar o Modo Teste:",
                font=("Segoe UI", 12, "bold"),
                text_color=("#334155", "#CBD5E1")
            ).pack(pady=(20, 10), padx=20)

            ctk.CTkLabel(
                top,
                text=pergunta["pergunta"],
                font=("Segoe UI", 11, "bold"),
                text_color=("#0F172A", "#E2E8F0")
            ).pack(pady=(0, 10), padx=20)

            entrada = ctk.CTkEntry(top, width=250, height=32, font=("Segoe UI", 12))
            entrada.pack(pady=(0, 15), padx=20)
            entrada.focus_set()

            resultado = {"ok": False}

            def confirmar():
                valor = entrada.get().strip()
                if valor.lower() == pergunta["resposta"].lower():
                    resultado["ok"] = True
                    top.destroy()
                else:
                    messagebox.showerror("Erro", "Resposta incorreta.")
                    entrada.delete(0, tk.END)
                    entrada.focus_set()

            ctk.CTkButton(
                top,
                text="Confirmar",
                command=confirmar,
                fg_color="#2563EB",
                hover_color="#1D4ED8",
                text_color="white",
                corner_radius=8,
                height=34,
                font=("Segoe UI", 11, "bold")
            ).pack(pady=(0, 15))

            top.wait_window()

            if not resultado["ok"]:
                return False

            dados_usuario["pergunta_atual_id"] = pergunta["id"]
            dados_usuario["perguntas_usadas"] = perguntas_usadas + [pergunta]
            save_usuarios(self.usuarios)
            return True

        return True

    # Ativa modo teste admin
    def _ativar_modo_teste_admin(self):
        dados = _ativar_modo_teste(admin_cpf=self.usuario or "", duracao_minutos=10, max_prorrogacoes=5)
        self._garantir_usuario_teste()
        messagebox.showinfo(
            "Modo Teste",
            "Modo Teste ativado com sucesso!\n\n"
            "Duração: 10 minutos\n"
            "Prorrogações permitidas: 5\n\n"
            "O sistema será reiniciado e entrará automaticamente no Modo Teste."
        )
        self._registrar_log("Ativou Modo Teste")
        try:
            with open(ARQUIVO_SESSAO, "w", encoding="utf-8") as f:
                json.dump({"usuario": "00000000000"}, f)
        except Exception:
            pass
        self.destroy()
        subprocess.Popen([sys.executable, os.path.join(APP_DIR, "sistema.py")])

    # Prorroga modo teste admin
    def _prorrogar_modo_teste_admin(self):
        if self.usuario == "00000000000":
            if not self._verificar_pergunta_modo_teste():
                return

        dados = _prorrogar_modo_teste()
        if not dados:
            messagebox.showwarning("Aviso", "Não foi possível prorrogar. Limite de prorrogações atingido ou Modo Teste expirado.")
            self._build_home_screen()
            return
        prorrogacoes = dados.get("prorrogacoes", 0)
        max_prorrogacoes = dados.get("max_prorrogacoes", 5)
        messagebox.showinfo(
            "Modo Teste",
            f"Modo Teste prorrogado com sucesso!\n\n"
            f"Prorrogação: {prorrogacoes}/{max_prorrogacoes}\n"
            f"Nova expiração: {dados.get('expira_em', '')}"
        )
        self._registrar_log(f"Prorrogou Modo Teste ({prorrogacoes}/{max_prorrogacoes})")
        self._build_home_screen()

    # Desativa modo teste admin
    def _desativar_modo_teste_admin(self):
        confirmar = messagebox.askyesno("Confirmação", "Deseja desativar o Modo Teste?")
        if not confirmar:
            return
        _desativar_modo_teste()
        messagebox.showinfo("Modo Teste", "Modo Teste desativado com sucesso.")
        self._registrar_log("Desativou Modo Teste")
        self._build_home_screen()

    # Abre relatorios
    def _open_relatorios(self):
        top = self._open_inline_container("Relatórios", voltar=self._build_home_screen)

        notebook = tb.Notebook(top)
        notebook.pack(fill="both", expand=True, padx=10, pady=10)

        aba1 = tb.Frame(notebook, padding=10)
        notebook.add(aba1, text="Cadastros por Período")

        aba2 = tb.Frame(notebook, padding=10)
        notebook.add(aba2, text="Média de Tempo por Ação")

        aba3 = tb.Frame(notebook, padding=10)
        notebook.add(aba3, text="Ações com Mais/Menos Cadastros")

        aba4 = tb.Frame(notebook, padding=10)
        notebook.add(aba4, text="Cadastros por Ação")

        cadastros = self.db.get("cadastros", [])
        acoes = self.db.get("acoes", [])
        acao_por_id = {str(a.get("id", "")): a for a in acoes}

        # Interpreta data
        def parse_data(d):
            try:
                return datetime.strptime(d, "%d/%m/%Y")
            except:
                return None

        # Aba1 content
        def aba1_content():
            frm = aba1

            tb.Label(frm, text="Período:", font=("Segoe UI", 10)).pack(anchor="w", pady=(5, 0))

            filtro_frame = tb.Frame(frm)
            filtro_frame.pack(fill="x", pady=5)

            tb.Label(filtro_frame, text="De:").pack(side="left", padx=(0, 5))
            ent_de = tb.Entry(filtro_frame, width=15)
            ent_de.pack(side="left", padx=5)
            ent_de.insert(0, "01/01/2025")

            tb.Label(filtro_frame, text="Até:").pack(side="left", padx=(15, 5))
            ent_ate = tb.Entry(filtro_frame, width=15)
            ent_ate.pack(side="left", padx=5)
            ent_ate.insert(0, datetime.now().strftime("%d/%m/%Y"))

            var_tipo = tk.StringVar(value="TODOS")
            tb.Label(filtro_frame, text="Tipo Serviço:").pack(side="left", padx=(20, 5))
            combo_tipo = tb.Combobox(filtro_frame, values=["TODOS"] + [s.get("nome", "") for s in self.db.get("servicos", [])], textvariable=var_tipo, width=20, state="readonly")
            combo_tipo.pack(side="left", padx=5)

            txt_resultado = tk.Text(frm, height=15, wrap="none")
            txt_resultado.pack(fill="both", expand=True, pady=10)

            # Calcula
            def calcular():
                de = parse_data(ent_de.get())
                ate = parse_data(ent_ate.get())
                tipo = var_tipo.get()

                if not de or not ate:
                    messagebox.showerror("Erro", "Datas inválidas", parent=frm)
                    return

                filtrados = [c for c in cadastros if c.get("data") and de <= parse_data(c.get("data", "")) <= ate]
                if tipo and tipo != "TODOS":
                    filtrados = [c for c in filtrados if get_nome_servico_por_id(self.db, c.get("servico_id", "")) == tipo]

                total = len(filtrados)
                entregues = sum(1 for c in filtrados if c.get("entregue"))

                por_data = {}
                por_mes = {}
                por_ano = {}

                for c in filtrados:
                    d = parse_data(c.get("data", ""))
                    if d:
                        por_data[c.get("data", "")] = por_data.get(c.get("data", ""), 0) + 1
                        mes = d.strftime("%m/%Y")
                        por_mes[mes] = por_mes.get(mes, 0) + 1
                        ano = d.strftime("%Y")
                        por_ano[ano] = por_ano.get(ano, 0) + 1

                txt_resultado.delete("1.0", "end")
                txt_resultado.insert("end", f"TOTAL: {total}\nENTREGUES: {entregues}\nPENDENTES: {total - entregues}\n\n")

                txt_resultado.insert("end", "=== POR DATA ===\n")
                for d, q in sorted(por_data.items(), key=lambda x: parse_data(x[0]) or datetime.strptime("01/01/1900", "%d/%m/%Y")):
                    txt_resultado.insert("end", f"  {d}: {q}\n")

                txt_resultado.insert("end", "\n=== POR MÊS ===\n")
                for m, q in sorted(por_mes.items()):
                    txt_resultado.insert("end", f"  {m}: {q}\n")

                txt_resultado.insert("end", "\n=== POR ANO ===\n")
                for a, q in sorted(por_ano.items()):
                    txt_resultado.insert("end", f"  {a}: {q}\n")

            # Exporta PDF
            def exportar_pdf():
                try:
                    de = parse_data(ent_de.get())
                    ate = parse_data(ent_ate.get())
                    tipo = var_tipo.get()

                    filtrados = [c for c in cadastros if c.get("data") and de <= parse_data(c.get("data", "")) <= ate]
                    if tipo and tipo != "TODOS":
                        filtrados = [c for c in filtrados if get_nome_servico_por_id(self.db, c.get("servico_id", "")) == tipo]

                    caminho = caminho_pdf("relatorios", f"cadastros_periodo_{ent_de.get().replace('/', '-')}_{ent_ate.get().replace('/', '-')}")
                    gerar_pdf_relatorio(filtrados, caminho, tipo, ent_de.get(), ent_ate.get())
                    messagebox.showinfo("PDF", f"Relatório salvo em:\n{caminho}", parent=frm)
                    self._abrir_arquivo(caminho)
                except Exception as e:
                    messagebox.showerror("Erro", str(e), parent=frm)

            btn_frame = tb.Frame(frm)
            btn_frame.pack(fill="x", pady=5)
            tb.Button(btn_frame, text="Calcular", bootstyle=PRIMARY, command=calcular).pack(side="left", padx=5)
            tb.Button(btn_frame, text="Exportar PDF", bootstyle=SUCCESS, command=exportar_pdf).pack(side="left", padx=5)

        # Aba2 content
        def aba2_content():

            frm = aba2

            tb.Label(frm, text="Selecione a Ação:", font=("Segoe UI", 10)).pack(anchor="w", pady=(5, 0))

            acoes_lista = [f"{a.get('id', '')} - {a.get('data', '')} - {a.get('local', '')}" for a in acoes]
            combo_acao = tb.Combobox(frm, values=acoes_lista, width=50, state="readonly")
            combo_acao.pack(fill="x", pady=5)

            intervalo_frame = tb.Labelframe(frm, text="Intervalo de Horas", padding=8)
            intervalo_frame.pack(fill="x", pady=(8, 10))

            tb.Label(intervalo_frame, text="Início Manhã:").pack(side="left", padx=(0, 4))
            ent_inicio_manha = tb.Entry(intervalo_frame, width=9)
            ent_inicio_manha.pack(side="left", padx=(0, 10))
            ent_inicio_manha.insert(0, "09:00")

            tb.Label(intervalo_frame, text="Fim Manhã:").pack(side="left", padx=(5, 4))
            ent_fim_manha = tb.Entry(intervalo_frame, width=9)
            ent_fim_manha.pack(side="left", padx=(0, 10))
            ent_fim_manha.insert(0, "12:00")

            tb.Label(intervalo_frame, text="Início Tarde:").pack(side="left", padx=(5, 4))
            ent_inicio_tarde = tb.Entry(intervalo_frame, width=9)
            ent_inicio_tarde.pack(side="left", padx=(0, 10))
            ent_inicio_tarde.insert(0, "13:00")

            tb.Label(intervalo_frame, text="Fim Tarde:").pack(side="left", padx=(5, 4))
            ent_fim_tarde = tb.Entry(intervalo_frame, width=9)
            ent_fim_tarde.pack(side="left", padx=(0, 10))
            ent_fim_tarde.insert(0, "17:00")

            # Mascara hora entrada local
            def mascara_hora_entrada_local(entry):
                digits = "".join(ch for ch in entry.get() if ch.isdigit())
                digits = digits[:4]
                if len(digits) > 2:
                    digits = digits[:2] + ":" + digits[2:]
                entry.delete(0, tk.END)
                entry.insert(0, digits)

            for campo in (ent_inicio_manha, ent_fim_manha, ent_inicio_tarde, ent_fim_tarde):
                campo.bind("<KeyRelease>", lambda e, c=campo: mascara_hora_entrada_local(c))
                campo.bind("<FocusOut>", lambda e, c=campo: mascara_hora_entrada_local(c))

            txt_resultado = tk.Text(frm, height=15, wrap="none")
            txt_resultado.pack(fill="both", expand=True, pady=10)
            txt_resultado.tag_configure("cap_alta", foreground="red", font=("Segoe UI", 10, "bold"))
            txt_resultado.tag_configure("cap_ok", foreground="blue", font=("Segoe UI", 10, "bold"))

            # Formata hora min
            def formatar_hora_min(minutos):
                h = int(minutos // 60)
                m = int(minutos % 60)
                return f"{h}h:{m:02d}min"

            # Interpreta hora
            def parse_hora(hora_str):
                if not hora_str:
                    return None
                try:
                    h, m = hora_str.split(":")
                    return int(h) * 60 + int(m)
                except Exception:
                    try:
                        return int(hora_str) * 60
                    except Exception:
                        return None

            # Calcula
            def calcular():
                acao_str = combo_acao.get()
                if not acao_str:
                    return

                acao_id = acao_str.split(" - ")[0]
                acao = acao_por_id.get(acao_id, {})
                dados_acao = [c for c in cadastros if str(c.get("acao_id", "")) == acao_id]

                total = len(dados_acao)
                entregues = sum(1 for c in dados_acao if c.get("entregue"))

                dados_ordenados = sorted(dados_acao, key=lambda x: x.get("hora", "") or "00:00")

                tempos_entre_registros = []
                for i in range(1, len(dados_ordenados)):
                    h1 = parse_hora(dados_ordenados[i-1].get("hora", "00:00"))
                    h2 = parse_hora(dados_ordenados[i].get("hora", "00:00"))
                    if h1 is not None and h2 is not None and h2 > h1:
                        tempos_entre_registros.append(h2 - h1)

                media_tempo = sum(tempos_entre_registros) / len(tempos_entre_registros) if tempos_entre_registros else 0
                total_percorrido = sum(tempos_entre_registros)

                primeira_hora = dados_ordenados[0].get("hora", "-") if dados_ordenados else "-"
                ultima_hora = dados_ordenados[-1].get("hora", "-") if dados_ordenados else "-"

                inicio_manha = parse_hora(ent_inicio_manha.get())
                fim_manha = parse_hora(ent_fim_manha.get())
                inicio_tarde = parse_hora(ent_inicio_tarde.get())
                fim_tarde = parse_hora(ent_fim_tarde.get())

                tempo_acao = 0
                if inicio_manha is not None and fim_manha is not None and fim_manha > inicio_manha:
                    tempo_acao += fim_manha - inicio_manha
                if inicio_tarde is not None and fim_tarde is not None and fim_tarde > inicio_tarde:
                    tempo_acao += fim_tarde - inicio_tarde

                media_min = int(media_tempo) if media_tempo > 0 else 0
                capacidade_estimada = round(total_percorrido / media_min) if media_min > 0 else 0

                # Média diária nos últimos 30 dias (todas as ações)
                hoje_dt = datetime.now()
                limite_30 = hoje_dt - timedelta(days=30)
                def _parse_data_rel(d):
                    try:
                        return datetime.strptime(d, "%d/%m/%Y")
                    except Exception:
                        return None
                cadastros_30d = 0
                for c in cadastros:
                    pd = _parse_data_rel(c.get("data", ""))
                    if pd and limite_30 <= pd <= hoje_dt:
                        cadastros_30d += 1
                media_por_dia = cadastros_30d / 30.0

                txt_resultado.delete("1.0", "end")
                txt_resultado.insert("end", f"AÇÃO: {acao.get('local', '')} - {acao.get('data', '')}\n")
                txt_resultado.insert("end", f"TOTAL CADASTROS: {total}\n")
                txt_resultado.insert("end", f"IDENTIDADES ENTREGUES: {entregues}\n")
                txt_resultado.insert("end", f"\n")
                txt_resultado.insert("end", f"HORÁRIO INICIAL: {primeira_hora}\n")
                txt_resultado.insert("end", f"HORÁRIO FINAL: {ultima_hora}\n")
                txt_resultado.insert("end", f"TEMPO TOTAL: {formatar_hora_min(total_percorrido)}\n")
                txt_resultado.insert("end", f"\n")
                txt_resultado.insert("end", f"MÉDIA ENTRE REGISTROS: {formatar_hora_min(media_tempo)}\n")
                tag = "cap_alta" if capacidade_estimada > total else "cap_ok"
                txt_resultado.insert("end", f"CAPACIDADE ESTIMADA: ~")
                txt_resultado.insert("end", f"{capacidade_estimada}", tag)
                txt_resultado.insert("end", f" cadastros\n")
                txt_resultado.insert("end", f"\n")
                txt_resultado.insert("end", f"MÉDIA POR DIA (30 DIAS): {media_por_dia:.1f} cadastros/dia\n")

            # Exporta PDF
            def exportar_pdf():
                acao_str = combo_acao.get()
                if not acao_str:
                    return
                acao_id = acao_str.split(" - ")[0]
                acao = acao_por_id.get(acao_id, {})
                dados_acao = [c for c in cadastros if str(c.get("acao_id", "")) == acao_id]

                total = len(dados_acao)
                entregues = sum(1 for c in dados_acao if c.get("entregue"))

                dados_ordenados = sorted(dados_acao, key=lambda x: x.get("hora", "") or "00:00")

                tempos_entre_registros = []
                for i in range(1, len(dados_ordenados)):
                    h1 = parse_hora(dados_ordenados[i-1].get("hora", "00:00"))
                    h2 = parse_hora(dados_ordenados[i].get("hora", "00:00"))
                    if h1 is not None and h2 is not None and h2 > h1:
                        tempos_entre_registros.append(h2 - h1)

                media_tempo = sum(tempos_entre_registros) / len(tempos_entre_registros) if tempos_entre_registros else 0
                total_percorrido = sum(tempos_entre_registros)

                primeira_hora = dados_ordenados[0].get("hora", "-") if dados_ordenados else "-"
                ultima_hora = dados_ordenados[-1].get("hora", "-") if dados_ordenados else "-"

                inicio_manha = parse_hora(ent_inicio_manha.get())
                fim_manha = parse_hora(ent_fim_manha.get())
                inicio_tarde = parse_hora(ent_inicio_tarde.get())
                fim_tarde = parse_hora(ent_fim_tarde.get())

                tempo_acao = 0
                if inicio_manha is not None and fim_manha is not None and fim_manha > inicio_manha:
                    tempo_acao += fim_manha - inicio_manha
                if inicio_tarde is not None and fim_tarde is not None and fim_tarde > inicio_tarde:
                    tempo_acao += fim_tarde - inicio_tarde

                media_min = int(media_tempo) if media_tempo > 0 else 0
                capacidade_estimada = round(total_percorrido / media_min) if media_min > 0 else 0

                # Média diária nos últimos 30 dias (todas as ações)
                hoje_dt = datetime.now()
                limite_30 = hoje_dt - timedelta(days=30)
                def _parse_data_rel(d):
                    try:
                        return datetime.strptime(d, "%d/%m/%Y")
                    except Exception:
                        return None
                cadastros_30d = 0
                for c in cadastros:
                    pd = _parse_data_rel(c.get("data", ""))
                    if pd and limite_30 <= pd <= hoje_dt:
                        cadastros_30d += 1
                media_por_dia = cadastros_30d / 30.0

                try:
                    caminho = caminho_pdf("relatorios", f"acao_{acao_id}")
                    gerar_pdf_relatorio_acao(dados_acao, caminho, acao, formatar_hora_min(total_percorrido), total, entregues, formatar_hora_min(media_tempo), capacidade_estimada, media_por_dia, primeira_hora, ultima_hora)
                    messagebox.showinfo("PDF", f"Relatório salvo em:\n{caminho}", parent=frm)
                    self._abrir_arquivo(caminho)
                except Exception as e:
                    messagebox.showerror("Erro", str(e), parent=frm)

            btn_frame = tb.Frame(frm)
            btn_frame.pack(fill="x", pady=5)
            tb.Button(btn_frame, text="Calcular", bootstyle=PRIMARY, command=calcular).pack(side="left", padx=5)
            tb.Button(btn_frame, text="Exportar PDF", bootstyle=SUCCESS, command=exportar_pdf).pack(side="left", padx=5)

        # Aba3 content
        def aba3_content():
            frm = aba3

            txt_resultado = tk.Text(frm, height=20, wrap="none")
            txt_resultado.pack(fill="both", expand=True, pady=10)

            # Popula aba3
            def popular_aba3():
                cadastros_atuais = self.db.get("cadastros", [])
                acoes_atuais = self.db.get("acoes", [])
                acao_por_id_atual = {str(a.get("id", "")): a for a in acoes_atuais}

                stats_acoes = {}
                for c in cadastros_atuais:
                    aid = c.get("acao_id", "")
                    if aid:
                        stats_acoes[aid] = stats_acoes.get(aid, 0) + 1

                ordenado_local = sorted(stats_acoes.items(), key=lambda x: x[1], reverse=True)
                maior = ordenado_local[0] if ordenado_local else (None, 0)
                menor = ordenado_local[-1] if ordenado_local else (None, 0)

                txt_resultado.delete("1.0", "end")
                txt_resultado.insert("end", "=== AÇÃO COM MAIS CADASTROS ===\n")
                if maior[0]:
                    acao = acao_por_id_atual.get(maior[0], {})
                    txt_resultado.insert("end", f"  {acao.get('local', '')} - {acao.get('data', '')}: {maior[1]}\n")

                txt_resultado.insert("end", "\n=== AÇÃO COM MENOS CADASTROS ===\n")
                if menor[0]:
                    acao = acao_por_id_atual.get(menor[0], {})
                    txt_resultado.insert("end", f"  {acao.get('local', '')} - {acao.get('data', '')}: {menor[1]}\n")

                txt_resultado.insert("end", "\n=== RANKING COMPLETO ===\n")
                for acao_id, qtd in ordenado_local:
                    acao = acao_por_id_atual.get(acao_id, {})
                    txt_resultado.insert("end", f"  {acao.get('local', '')} - {acao.get('data', '')}: {qtd}\n")

                return ordenado_local, acao_por_id_atual

            ordenado, acao_por_id_atual = popular_aba3()

            # Exporta PDF
            def exportar_pdf():
                try:
                    ordenado_fresh, acao_por_id_fresh = popular_aba3()
                    caminho = caminho_pdf("relatorios", "ranking_acoes")
                    gerar_pdf_ranking(ordenado_fresh, caminho, acao_por_id_fresh)
                    messagebox.showinfo("PDF", f"Relatório salvo em:\n{caminho}", parent=frm)
                    self._abrir_arquivo(caminho)
                except Exception as e:
                    messagebox.showerror("Erro", str(e), parent=frm)

            btn_frame = tb.Frame(frm)
            btn_frame.pack(fill="x", pady=5)
            tb.Button(btn_frame, text="Atualizar", bootstyle=INFO, command=popular_aba3).pack(side="left", padx=5)
            tb.Button(btn_frame, text="Exportar PDF", bootstyle=SUCCESS, command=exportar_pdf).pack(side="left", padx=5)

        # Aba4 content
        def aba4_content():
            frm = aba4

            var_acao_nome = tk.StringVar()
            tb.Label(frm, text="Nome da Ação:").pack(anchor="w", pady=(5, 0))
            ent_acao = tb.Entry(frm, textvariable=var_acao_nome, width=40)
            ent_acao.pack(fill="x", pady=5)

            txt_resultado = tk.Text(frm, height=15, wrap="none")
            txt_resultado.pack(fill="both", expand=True, pady=10)

            # Calcula
            def calcular():
                nome = var_acao_nome.get().lower().strip()
                txt_resultado.delete("1.0", "end")

                if not nome:
                    return

                acoes_filtradas = [a for a in acoes if nome in a.get("local", "").lower().strip()]

                total = 0
                entregues = 0
                for a in acoes_filtradas:
                    dados = [c for c in cadastros if str(c.get("acao_id", "")) == str(a.get("id", ""))]
                    total += len(dados)
                    entregues += sum(1 for c in dados if c.get("entregue"))

                txt_resultado.insert("end", f"AÇÕES ENCONTRADAS: {len(acoes_filtradas)}\n")
                txt_resultado.insert("end", f"TOTAL CADASTROS: {total}\n")
                txt_resultado.insert("end", f"IDENTIDADES ENTREGUES: {entregues}\n\n")

                txt_resultado.insert("end", "=== POR AÇÃO ===\n")
                for a in acoes_filtradas:
                    dados = [c for c in cadastros if str(c.get("acao_id", "")) == str(a.get("id", ""))]
                    e = sum(1 for c in dados if c.get("entregue"))
                    txt_resultado.insert("end", f"  {a.get('data', '')} - {a.get('local', '')}: {len(dados)} (Entregues: {e})\n")

            # Exporta PDF
            def exportar_pdf():
                nome = var_acao_nome.get().lower().strip()
                acoes_filtradas = [a for a in acoes if nome in a.get("local", "").lower().strip()]

                all_dados = []
                for a in acoes_filtradas:
                    all_dados.extend([c for c in cadastros if str(c.get("acao_id", "")) == str(a.get("id", ""))])

                try:
                    caminho = caminho_pdf("relatorios", f"acao_nome_{var_acao_nome.get()}")
                    gerar_pdf_relatorio(all_dados, caminho, f"Ação: {var_acao_nome.get()}", "", "")
                    messagebox.showinfo("PDF", f"Relatório salvo em:\n{caminho}", parent=frm)
                    self._abrir_arquivo(caminho)
                except Exception as e:
                    messagebox.showerror("Erro", str(e), parent=frm)

            btn_frame = tb.Frame(frm)
            btn_frame.pack(fill="x", pady=5)
            tb.Button(btn_frame, text="Buscar", bootstyle=PRIMARY, command=calcular).pack(side="left", padx=5)
            tb.Button(btn_frame, text="Exportar PDF", bootstyle=SUCCESS, command=exportar_pdf).pack(side="left", padx=5)

        aba1_content()
        aba2_content()
        aba3_content()
        aba4_content()
        notebook.select(aba1)
        self.update_idletasks()

    # Abre acesso tela
    def _open_acesso_screen(self):
        top = self._open_inline_container("Acesso", voltar=self._build_home_screen)
        top.title("Acesso")
        top.geometry("850x560")
        top.grab_set()

        frm = tb.Frame(top, padding=12)
        frm.pack(fill="both", expand=True)

        botoes = tb.Frame(frm)
        botoes.pack(fill="x", pady=(0, 10))

        if self.perfil == "admin":

            control = tb.Frame(frm)
            control.pack(fill="x", pady=5)

            tb.Label(control, text="Pesquisar usuário:").pack(side="left", padx=5)

            self.ent_pesquisa_user = tb.Entry(control, width=30)
            self.ent_pesquisa_user.pack(side="left", padx=5)

            self.ent_pesquisa_user.bind("<KeyRelease>", lambda e: self._pesquisar_usuario()
            )

            tb.Button(
                control,
                text="Criar Login",
                bootstyle="success",
                command=self._open_criar_usuario
            ).pack(side="left", padx=5)

            tb.Button(
                control,
                text="Editar",
                bootstyle="primary",
                command=self._editar_usuario_selecionado
            ).pack(side="left", padx=5)

            tb.Button(
                control,
                text="Redefinir Senha",
                bootstyle="warning",
                command=self._redefinir_senha_usuario
            ).pack(side="left", padx=5)

            tb.Button(
                control,
                text="Excluir",
                bootstyle="danger",
                command=lambda: self._excluir_usuario_tree(self.tree_users)
            ).pack(side="left", padx=5)

            tb.Button(
                control,
                text="Ativar/Desativar",
                bootstyle="warning",
                command=self._alternar_status_usuario
            ).pack(side="left", padx=5)

            cols = ("cpf", "nome", "perfil", "data_nascimento", "email", "ativo")
            self.tree_users = tb.Treeview(frm, columns=cols, show="headings", height=16)
            for c in cols:
                self.tree_users.heading(c, text=c.upper())
            self.tree_users.column("cpf", width=110)
            self.tree_users.column("nome", width=180)
            self.tree_users.column("perfil", width=80)
            self.tree_users.column("data_nascimento", width=100)
            self.tree_users.column("email", width=180)
            self.tree_users.column("ativo", width=70)
            self.tree_users.pack(fill="both", expand=True, pady=8)
            self._preparar_tabela(self.tree_users, minimo=90, maximo=420)
            self.tree_users.bind("<Double-1>", lambda e: self._editar_usuario_selecionado())
            self._carregar_usuarios()
        else:

            tb.Label(
                frm,
                text=f"Usuário: {self._nome_usuario_logado()}",
                font=("Segoe UI", 12, "bold")
            ).pack(anchor="w", pady=10)

            botoes_usuario = tb.Frame(frm)
            botoes_usuario.pack(anchor="w", pady=10)

            tb.Button(
                botoes_usuario,
                text="Editar Meus Dados",
                bootstyle="primary",
                command=lambda: self._editar_usuario_por_cpf(self.usuario)
            ).pack(side="left", padx=5)

            tb.Button(
                botoes_usuario,
                text="Redefinir Minha Senha",
                bootstyle="warning",
                command=self._redefinir_minha_senha
            ).pack(side="left", padx=5)

    # Carrega usuarios
    def _carregar_usuarios(self):
        for i in self.tree_users.get_children():
            self.tree_users.delete(i)
        usuarios_ordenados = sorted(self.usuarios.items(), key=lambda item: normalizar_texto_ordenacao(item[1].get("nome", "")))
        for cpf, dados in usuarios_ordenados:
            nome = dados.get("nome", "")
            perfil = dados.get("perfil", "")
            data_nasc = dados.get("data_nascimento", "")
            email = dados.get("email", "")
            ativo = "Sim" if dados.get("ativo", True) else "Não"
            cpf_mask = mask_cpf_from_clean(cpf)
            self.tree_users.insert("", "end", values=(cpf_mask, nome, perfil, data_nasc, email, ativo))

    # Desfaz entrega selecionada
    def desfazer_entrega_selecionada(self, tree_h):

        selecionado = tree_h.focus()
        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione um registro para desfazer.")
            return

        if not self._autorizar_admin():
            return

        valores = tree_h.item(selecionado)["values"]
        protocolo_alvo = str(valores[1])

        if not messagebox.askyesno("Confirmar", f"Desfazer entrega do protocolo {protocolo_alvo}?"):
            return

        if os.path.exists(DB_ENTREGAS):
            with open(DB_ENTREGAS, "r", encoding="utf-8") as f:
                entregas = json.load(f)
            entregas_at = [e for e in entregas if str(e.get("protocolo")) != protocolo_alvo]
            with open(DB_ENTREGAS, "w", encoding="utf-8") as f:
                json.dump(entregas_at, f, indent=4, ensure_ascii=False)

        for c in self.db.get("cadastros", []):
            if str(c.get("protocolo")) == protocolo_alvo:
                c["entregue"] = False
                c.pop("data_entrega", None)
                c.pop("operador_entrega", None)

        save_db(self.db)
        tree_h.delete(selecionado)
        self._registrar_log("Desfez entrega", protocolo_alvo)
        self.atualizar_tabela()

        messagebox.showinfo("Sucesso", "Operação concluída!")

    # Desfaz todas entregas tela
    def desfazer_todas_entregas_tela(self, tree_h):

        if not self._autorizar_admin():
            return

        itens = tree_h.get_children()
        if not itens:
            messagebox.showwarning("Aviso", "Não há registros na tela.")
            return

        if not messagebox.askyesno("Confirmar", f"Desfazer TODAS as {len(itens)} entregas da tela?"):
            return

        protocolos_tela = [str(tree_h.item(i)["values"][1]) for i in itens]

        if os.path.exists(DB_ENTREGAS):
            with open(DB_ENTREGAS, "r", encoding="utf-8") as f:
                entregas = json.load(f)
            entregas_res = [e for e in entregas if str(e.get("protocolo")) not in protocolos_tela]
            with open(DB_ENTREGAS, "w", encoding="utf-8") as f:
                json.dump(entregas_res, f, indent=4, ensure_ascii=False)

        for c in self.db.get("cadastros", []):
            if str(c.get("protocolo")) in protocolos_tela:
                c["entregue"] = False
                c.pop("data_entrega", None)
                c.pop("operador_entrega", None)

        save_db(self.db)
        for i in itens: tree_h.delete(i)
        self._registrar_log("Desfez tudo na tela", f"{len(protocolos_tela)} itens")
        messagebox.showinfo("Sucesso", "Operação concluída!")

    # Alternar status usuário
    def _alternar_status_usuario(self):
        sel = self.tree_users.selection()
        if not sel:
            messagebox.showwarning("Aviso", "Selecione um usuário na tabela.")
            return
        cpf_mask = self.tree_users.item(sel[0], "values")[0]
        cpf = somente_digs(cpf_mask)
        if cpf not in self.usuarios:
            messagebox.showerror("Erro", "Usuário não encontrado.")
            return
        if self.usuarios[cpf].get("perfil") == "admin":
            messagebox.showerror("Erro", "Não é permitido desativar um administrador.")
            return
        atual = self.usuarios[cpf].get("ativo", True)
        self.usuarios[cpf]["ativo"] = not atual
        save_usuarios(self.usuarios)
        self._carregar_usuarios()
        status = "ativado" if not atual else "desativado"
        messagebox.showinfo("OK", f"Usuário {status} com sucesso.")

    # Pesquisa usuário
    def _pesquisar_usuario(self):
        termo = normalizar_texto_ordenacao(self.ent_pesquisa_user.get())
        for i in self.tree_users.get_children():
            self.tree_users.delete(i)
        for cpf, dados in self.usuarios.items():
            nome = normalizar_texto_ordenacao(dados.get("nome", ""))
            perfil = normalizar_texto_ordenacao(dados.get("perfil", ""))
            cpf_mask = mask_cpf_from_clean(cpf)
            dn = normalizar_texto_ordenacao(dados.get("data_nascimento", ""))
            email = normalizar_texto_ordenacao(dados.get("email", ""))
            ativo = normalizar_texto_ordenacao("sim" if dados.get("ativo", True) else "nao")
            if termo in cpf.lower() or termo in nome or termo in perfil or termo in dn or termo in email or termo in ativo:
                self.tree_users.insert("", "end", values=(cpf_mask, dados.get("nome", ""), dados.get("perfil", ""), dados.get("data_nascimento", ""), dados.get("email", ""), "Sim" if dados.get("ativo", True) else "Não"))

    # Ordena usuarios
    def _ordenar_usuarios(self, chave):
        for i in self.tree_users.get_children():
            self.tree_users.delete(i)
        if chave == "cpf":
            ordenado = sorted(self.usuarios.items(), key=lambda x: x[0])
        else:
            ordenado = sorted(self.usuarios.items(), key=lambda x: normalizar_texto_ordenacao(x[1].get(chave, "")))
        for cpf, dados in ordenado:
            cpf_mask = mask_cpf_from_clean(cpf)
            self.tree_users.insert("", "end", values=(cpf_mask, dados.get("nome", ""), dados.get("perfil", ""), dados.get("data_nascimento", ""), dados.get("email", ""), "Sim" if dados.get("ativo", True) else "Não"))

    # Abre gerenciar usuarios
    def _open_gerenciar_usuarios(self):

        top = self._open_inline_container("Gerenciar Usuários", voltar=self._build_home_screen)
        top.title("Gerenciar Usuários")
        top.geometry("800x520")
        top.grab_set()

        frm = ctk.CTkFrame(top, fg_color="transparent")
        frm.pack(fill="both", expand=True, padx=12, pady=12)

        control = ctk.CTkFrame(frm, fg_color="transparent")
        control.pack(fill="x")
        ctk.CTkLabel(control, text="Pesquisar:", text_color=("#334155", "#CBD5E1")).pack(side="left", padx=6)
        ent = ctk.CTkEntry(control, width=300)
        ent.pack(side="left", padx=6)
        ent.bind("<KeyRelease>", lambda e: self._pesquisar_usuario_geral(normalizar_texto_ordenacao(ent.get()), tree))

        ctk.CTkButton(control, text="Editar", command=lambda: self._editar_usuario_tree(tree), fg_color="#2563EB", hover_color="#1D4ED8", text_color="white", corner_radius=8, height=34, font=("Segoe UI", 12, "bold")).pack(side="left", padx=6)
        ctk.CTkButton(control, text="Excluir", command=lambda: self._excluir_usuario_tree(tree), fg_color="#E53935", hover_color="#C62828", text_color="white", corner_radius=8, height=34, font=("Segoe UI", 12, "bold")).pack(side="left", padx=6)

        cols = ("cpf", "nome", "perfil")
        tree = tb.Treeview(frm, columns=cols, show="headings", height=18)
        for c in cols:
            tree.heading(c, text=c.upper())
        tree.column("cpf", width=180)
        tree.column("nome", width=380)
        tree.column("perfil", width=120)
        tree.pack(fill="both", expand=True, pady=8)
        self._preparar_tabela(tree, minimo=90, maximo=420)

        for cpf, dados in self.usuarios.items():
            tree.insert("", "end", values=(mask_cpf_from_clean(cpf), dados.get("nome", ""), dados.get("perfil", "")))

    # Pesquisa usuário geral
    def _pesquisar_usuario_geral(self, termo, tree):
        for i in tree.get_children():
            tree.delete(i)
        for cpf, dados in self.usuarios.items():
            nome = normalizar_texto_ordenacao(dados.get("nome", ""))
            perfil = normalizar_texto_ordenacao(dados.get("perfil", ""))
            if termo in cpf.lower() or termo in nome or termo in perfil:
                tree.insert("", "end", values=(mask_cpf_from_clean(cpf), dados.get("nome", ""), dados.get("perfil", "")))

    # Edita usuário árvore
    def _editar_usuario_tree(self, tree):
        sel = tree.selection()
        if not sel:
            messagebox.showwarning("Aviso", "Selecione um usuário.")
            return
        cpf_mask = tree.item(sel[0], "values")[0]
        cpf = somente_digs(cpf_mask)
        self._editar_usuario_por_cpf(cpf)

    # Exclui usuário árvore
    def _excluir_usuario_tree(self, tree):

        if self.perfil != "admin":
            messagebox.showerror(
                "Acesso Negado",
                "Somente administradores podem excluir usuários."
            )
            return

        sel = tree.selection()

        if not sel:
            messagebox.showwarning(
                "Aviso",
                "Selecione um usuário."
            )
            return

        cpf_mask = tree.item(sel[0], "values")[0]
        cpf = somente_digs(cpf_mask)

        if cpf == self.usuario:
            messagebox.showwarning(
                "Aviso",
                "Não é possível excluir o usuário logado."
            )
            return

        nome = self.usuarios.get(cpf, {}).get("nome", "")

        confirmar = messagebox.askyesno(
            "Confirmar Exclusão",
            f"Deseja realmente excluir o usuário:\n\n"
            f"Nome: {nome}\n"
            f"CPF: {cpf_mask}"
        )

        if not confirmar:
            return

        self.usuarios.pop(cpf, None)

        save_usuarios(self.usuarios)

        self._registrar_log(
            "Usuário excluído",
            cpf
        )

        messagebox.showinfo(
            "Sucesso",
            "Usuário excluído com sucesso."
        )

        self._carregar_usuarios()

    # Edita usuário selecionado
    def _editar_usuario_selecionado(self):
        sel = self.tree_users.selection()
        if not sel:
            messagebox.showwarning("Aviso", "Selecione um usuário.")
            return
        cpf_mask = self.tree_users.item(sel[0], "values")[0]
        cpf = somente_digs(cpf_mask)
        self._editar_usuario_por_cpf(cpf)

    # Edita usuário por CPF
    def _editar_usuario_por_cpf(self, cpf):
        dados = self.usuarios.get(cpf)
        if not dados:
            messagebox.showerror("Erro", "Usuário não encontrado.")
            return
        top = self._open_inline_container("Editar Usuário", voltar=self._open_acesso_screen)
        top.title("Editar Usuário")
        top.geometry("460x320")
        top.grab_set()
        frm = ctk.CTkFrame(top, fg_color="transparent")
        frm.pack(fill="both", expand=True, padx=12, pady=12)
        ctk.CTkLabel(frm, text="Editar Usuário", font=("Segoe UI", 12, "bold"), text_color=("#0F172A", "#E2E8F0")).pack(pady=8)
        ctk.CTkLabel(frm, text="CPF (login):", text_color=("#334155", "#CBD5E1")).pack(anchor="w")
        ent_cpf = tb.Entry(frm)
        ent_cpf.insert(0, mask_cpf_from_clean(cpf))
        ent_cpf.configure(state="readonly")
        ent_cpf.pack(fill="x", pady=4)
        ctk.CTkLabel(frm, text="Nome:", text_color=("#334155", "#CBD5E1")).pack(anchor="w")
        ent_nome = ctk.CTkEntry(frm)
        ent_nome.insert(0, dados.get("nome", ""))
        ent_nome.pack(fill="x", pady=4)
        ent_nome.bind("<KeyRelease>", lambda e: self._forcar_maiusculo_entry(ent_nome))
        ctk.CTkLabel(frm, text="Data de Nascimento (DD/MM/AAAA):", text_color=("#334155", "#CBD5E1")).pack(anchor="w")
        ent_dn = ctk.CTkEntry(frm)
        ent_dn.insert(0, dados.get("data_nascimento", ""))
        ent_dn.pack(fill="x", pady=4)
        ent_dn.bind("<KeyRelease>", lambda e: mascara_data_entrada(ent_dn))
        ctk.CTkLabel(frm, text="E-mail:", text_color=("#334155", "#CBD5E1")).pack(anchor="w")
        ent_email = ctk.CTkEntry(frm)
        ent_email.insert(0, dados.get("email", ""))
        ent_email.pack(fill="x", pady=4)
        ctk.CTkLabel(frm, text="Perfil:", text_color=("#334155", "#CBD5E1")).pack(anchor="w")
        perfil_combo = tb.Combobox(frm, values=["admin", "usuario"])
        perfil_combo.set(dados.get("perfil", "usuario"))
        perfil_combo.pack(fill="x", pady=4)

        ctk.CTkLabel(frm, text="Vans Permitidas (Ctrl+Clique para múltiplos):", text_color=("#334155", "#CBD5E1")).pack(anchor="w", pady=(8, 0))
        listbox_vans = tk.Listbox(frm, selectmode="extended", height=4, exportselection=False)
        listbox_vans.pack(fill="x", pady=4)
        vans = load_vans()
        vans_por_nome = {}
        for v in vans:
            if v.get("ativa", True):
                listbox_vans.insert("end", v.get("nome", ""))
                vans_por_nome[v.get("nome", "")] = v.get("id", "")
        vans_permitidas = dados.get("vans_permitidas", [])
        for i in range(listbox_vans.size()):
            nome_van = listbox_vans.get(i)
            vid = vans_por_nome.get(nome_van, "")
            if vid in vans_permitidas:
                listbox_vans.selection_set(i)

        ctk.CTkLabel(frm, text="Van Ativa:", text_color=("#334155", "#CBD5E1")).pack(anchor="w", pady=(8, 0))
        combo_van_ativa = tb.Combobox(frm, values=[], state="readonly")
        combo_van_ativa.pack(fill="x", pady=4)

        def atualizar_combo_van_ativa_edicao(event=None):
            sel = listbox_vans.curselection()
            vans_selecionadas = [listbox_vans.get(i) for i in sel]
            combo_van_ativa.configure(values=vans_selecionadas)
            van_ativa_atual = dados.get("van_ativa", "")
            if van_ativa_atual and van_ativa_atual in vans_selecionadas:
                combo_van_ativa.set(van_ativa_atual)
            elif vans_selecionadas:
                combo_van_ativa.set(vans_selecionadas[0])
            else:
                combo_van_ativa.set("")

        listbox_vans.bind("<<ListboxSelect>>", atualizar_combo_van_ativa_edicao)
        atualizar_combo_van_ativa_edicao()

        ctk.CTkLabel(frm, text="Nova senha (opcional):", text_color=("#334155", "#CBD5E1")).pack(anchor="w")
        ent_new = ctk.CTkEntry(frm, show="*")
        ent_new.pack(fill="x", pady=4)
        # Salva
        def salvar():
            nome = ent_nome.get().strip().upper()
            if not nome:
                messagebox.showerror("Erro", "Informe o nome.")
                return
            dn = ent_dn.get().strip()
            if not validar_data_nascimento(dn):
                messagebox.showerror("Erro", "Informe uma data de nascimento válida (DD/MM/AAAA).")
                return
            email = ent_email.get().strip()
            if not validar_email(email):
                messagebox.showerror("Erro", "Informe um e-mail válido.")
                return
            perfil = perfil_combo.get()
            if perfil not in ("admin", "usuario"):
                messagebox.showerror("Erro", "Perfil inválido.")
                return
            sel = listbox_vans.curselection()
            vans_permitidas = []
            for idx in sel:
                nome_van = listbox_vans.get(idx)
                vid = vans_por_nome.get(nome_van, "")
                if vid:
                    vans_permitidas.append(vid)
            van_ativa = vans_por_nome.get(combo_van_ativa.get().strip(), "")
            self.usuarios[cpf]["nome"] = nome
            self.usuarios[cpf]["data_nascimento"] = dn
            self.usuarios[cpf]["email"] = email
            self.usuarios[cpf]["perfil"] = perfil
            self.usuarios[cpf]["vans_permitidas"] = vans_permitidas
            self.usuarios[cpf]["van_ativa"] = van_ativa
            nova = ent_new.get().strip()
            if nova:
                self.usuarios[cpf].update(hash_password(nova))
                self.usuarios[cpf]["ultima_troca_senha"] = datetime.now().strftime("%d/%m/%Y")
                self.usuarios[cpf]["senha_expirada"] = False
                self.usuarios[cpf]["tentativas_recuperacao"] = 0
                self.usuarios[cpf]["tentativas_login"] = 0
                self.usuarios[cpf]["bloqueado_ate"] = ""
                self._registrar_log("Senha redefinida", cpf)
            save_usuarios(self.usuarios)
            self._registrar_log("Usuário editado", cpf)
            messagebox.showinfo("OK", "Usuário atualizado.")
            self._fechar_tela_interna(top)
            if hasattr(self, "tree_users"):
                self._carregar_usuarios()
        ctk.CTkButton(frm, text="Salvar", command=salvar, fg_color="#16A34A", hover_color="#15803D", text_color="white", corner_radius=8, height=38, font=("Segoe UI", 12, "bold")).pack(pady=8)

    # Redefine senha usuário
    def _redefinir_senha_usuario(self):
        sel = self.tree_users.selection()
        if not sel:
            messagebox.showwarning("Aviso", "Selecione um usuário.")
            return
        cpf_mask = self.tree_users.item(sel[0], "values")[0]
        cpf = somente_digs(cpf_mask)
        top = self._open_inline_container("Redefinir Senha", voltar=self._open_acesso_screen)
        top.title("Redefinir Senha")
        top.geometry("380x220")
        top.grab_set()
        frm = ctk.CTkFrame(top, fg_color="transparent")
        frm.pack(fill="both", expand=True, padx=12, pady=12)
        ctk.CTkLabel(frm, text=f"Redefinir senha do usuário {mask_cpf_from_clean(cpf)}", font=("Segoe UI", 11, "bold"), text_color=("#0F172A", "#E2E8F0")).pack(pady=8)
        ctk.CTkLabel(frm, text="Nova senha:", text_color=("#334155", "#CBD5E1")).pack(anchor="w")
        ent1 = ctk.CTkEntry(frm, show="*")
        ent1.pack(fill="x", pady=5)
        ctk.CTkLabel(frm, text="Confirmar senha:", text_color=("#334155", "#CBD5E1")).pack(anchor="w")
        ent2 = ctk.CTkEntry(frm, show="*")
        ent2.pack(fill="x", pady=5)
        # Salva
        def salvar():
            p1 = ent1.get().strip()
            p2 = ent2.get().strip()
            if not p1:
                messagebox.showerror("Erro", "Digite a nova senha.")
                return
            if p1 != p2:
                messagebox.showerror("Erro", "As senhas não coincidem.")
                return
            self.usuarios[cpf].update(hash_password(p1))
            self.usuarios[cpf]["ultima_troca_senha"] = datetime.now().strftime("%d/%m/%Y")
            self.usuarios[cpf]["senha_expirada"] = False
            self.usuarios[cpf]["tentativas_recuperacao"] = 0
            self.usuarios[cpf]["tentativas_login"] = 0
            self.usuarios[cpf]["bloqueado_ate"] = ""
            save_usuarios(self.usuarios)
            self._registrar_log("Senha redefinida", cpf)
            messagebox.showinfo("OK", "Senha alterada.")
            self._fechar_tela_interna(top)
        ctk.CTkButton(frm, text="Salvar", command=salvar, fg_color="#16A34A", hover_color="#15803D", text_color="white", corner_radius=8, height=38, font=("Segoe UI", 12, "bold")).pack(pady=10)

    # Abre gerenciar servicos
    def _open_gerenciar_servicos(self):
        if self.perfil != "admin":
            messagebox.showerror("Acesso negado", "Somente administradores podem gerenciar serviços.")
            return

        top = self._open_inline_container("Gerenciar Serviços", voltar=self._build_home_screen)
        top.title("Gerenciar Serviços")
        top.geometry("700x420")
        top.grab_set()

        frm = ctk.CTkFrame(top, fg_color="transparent")
        frm.pack(fill="both", expand=True)

        control = ctk.CTkFrame(frm, fg_color="transparent")
        control.pack(fill="x")
        ctk.CTkLabel(control, text="Novo Serviço:", text_color=("#334155", "#CBD5E1")).pack(side="left", padx=6)
        ent_new = ctk.CTkEntry(control, width=380)
        ent_new.pack(side="left", padx=6)

        cols = ("id", "nome")
        tree = tb.Treeview(frm, columns=cols, show="headings", height=18)
        for c in cols:
            tree.heading(c, text=c.upper())
        tree.column("id", width=100)
        tree.column("nome", width=420)
        tree.pack(fill="both", expand=True, pady=8)
        self._preparar_tabela(tree, minimo=80, maximo=520)

        for s in self.db.get("servicos", []):
            tree.insert("", "end", values=(s["id"], s["nome"]))

        # Gera proximo id
        def gerar_proximo_id():
            existentes = [s for s in self.db.get("servicos", []) if str(s.get("id","")).isdigit()]
            if not existentes:
                return "1"
            try:
                nums = [int(s["id"]) for s in existentes]
                return str(max(nums) + 1)
            except Exception:
                return str(len(self.db.get("servicos", [])) + 1)

        # Adiciona
        def adicionar():
            nome = ent_new.get().strip()
            if not nome:
                messagebox.showwarning("Aviso", "Digite o nome do serviço.")
                return

            novo_id = gerar_proximo_id()

            self.db["servicos"].append({"id": novo_id, "nome": nome})
            save_db(self.db)

            tree.insert("", "end", values=(novo_id, nome))
            ent_new.delete(0, tk.END)

            self._registrar_log("Criou serviço", nome)
            messagebox.showinfo("OK", f"Serviço '{nome}' cadastrado com ID {novo_id}.")

        ctk.CTkButton(control, text="Adicionar", command=adicionar, fg_color="#16A34A", hover_color="#15803D", text_color="white", corner_radius=8, height=34, font=("Segoe UI", 12, "bold")).pack(side="left", padx=6)

        # Edita
        def editar():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Aviso", "Selecione um serviço.")
                return

            sid, nome_atual = tree.item(sel[0], "values")

            janela = self._open_inline_container("Editar Serviço", voltar=self._open_gerenciar_servicos)
            janela.title("Editar Serviço")
            janela.geometry("450x220")
            janela.resizable(False, False)
            janela.grab_set()

            frame_edit = ctk.CTkFrame(janela, fg_color="transparent")
            frame_edit.pack(fill="both", expand=True)

            ctk.CTkLabel(
                frame_edit,
                text="Editar Serviço",
                font=("Segoe UI", 14, "bold"),
                text_color=("#0F172A", "#E2E8F0")
            ).pack(pady=(0, 15))

            ctk.CTkLabel(frame_edit, text="Nome do serviço:", text_color=("#334155", "#CBD5E1")).pack(anchor="w")

            ent_nome_edit = ctk.CTkEntry(frame_edit, width=380)
            ent_nome_edit.pack(pady=8)
            ent_nome_edit.insert(0, nome_atual)
            ent_nome_edit.focus()

            # Salva edicao
            def salvar_edicao():
                novo_nome = ent_nome_edit.get().strip()

                if not novo_nome:
                    messagebox.showwarning("Aviso", "Digite o nome do serviço.")
                    return

                for s in self.db["servicos"]:
                    if s["nome"].lower() == novo_nome.lower() and s["id"] != sid:
                        messagebox.showerror(
                            "Erro",
                            "Já existe um serviço com esse nome.\n\n"
                            "Escolha um nome diferente."
                        )
                        return

                for s in self.db["servicos"]:
                    if s["id"] == sid:
                        s["nome"] = novo_nome
                        break

                save_db(self.db)
                self._registrar_log("Editou serviço", sid)

                messagebox.showinfo("OK", "Serviço atualizado com sucesso.")
                self._fechar_tela_interna(janela)

            botoes = ctk.CTkFrame(frame_edit, fg_color="transparent")
            botoes.pack(pady=15)

            ctk.CTkButton(
                botoes,
                text="Salvar",
                command=salvar_edicao,
                fg_color="#16A34A",
                hover_color="#15803D",
                text_color="white",
                corner_radius=8,
                height=34,
                font=("Segoe UI", 12, "bold")
            ).pack(side="left", padx=8)

        # Exclui
        def excluir():
            if not self._autorizar_admin():
                return

            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Aviso", "Selecione um serviço.")
                return
            sid, nome = tree.item(sel[0], "values")
            if messagebox.askyesno("Confirmar", f"Excluir serviço '{nome}'?"):
                self.db["servicos"] = [s for s in self.db["servicos"] if s["id"] != sid]
                save_db(self.db)
                tree.delete(sel[0])
                self._registrar_log("Excluiu serviço", sid)

        btns = ctk.CTkFrame(frm, fg_color="transparent")
        btns.pack(fill="x")
        ctk.CTkButton(btns, text="Editar", command=editar, fg_color="#2563EB", hover_color="#1D4ED8", text_color="white", corner_radius=8, height=34, font=("Segoe UI", 12, "bold")).pack(side="left", padx=6, pady=6)
        ctk.CTkButton(btns, text="Excluir", command=excluir, fg_color="#E53935", hover_color="#C62828", text_color="white", corner_radius=8, height=34, font=("Segoe UI", 12, "bold")).pack(side="left", padx=6, pady=6)

    # Abre gerenciar organizacoes
    def _open_gerenciar_organizacoes(self):
        if self.perfil != "admin":
            messagebox.showerror("Acesso negado", "Somente administradores podem gerenciar organizações.")
            return

        top = self._open_inline_container("Gerenciar Organizações", voltar=self._build_home_screen)
        top.title("Gerenciar Organizações")
        top.geometry("700x420")
        top.grab_set()

        frm = ctk.CTkFrame(top, fg_color="transparent")
        frm.pack(fill="both", expand=True)

        control = ctk.CTkFrame(frm, fg_color="transparent")
        control.pack(fill="x")
        ctk.CTkLabel(control, text="Nova Organização:", text_color=("#334155", "#CBD5E1")).pack(side="left", padx=6)
        ent_new = ctk.CTkEntry(control, width=380)
        ent_new.pack(side="left", padx=6)

        cols = ("id", "nome")
        tree = tb.Treeview(frm, columns=cols, show="headings", height=18)
        for c in cols:
            tree.heading(c, text=c.upper())
        tree.column("id", width=100)
        tree.column("nome", width=420)
        tree.pack(fill="both", expand=True, pady=8)
        self._preparar_tabela(tree, minimo=80, maximo=520)

        organizacoes = load_organizacoes()
        for idx, o in enumerate(organizacoes, start=1):
            tree.insert("", "end", values=(idx, o.get("nome", "")))

        # Gera proximo id
        def gerar_proximo_id():
            existentes = [o for o in organizacoes if str(o.get("id", "")).isdigit()]
            if not existentes:
                return "1"
            try:
                nums = [int(o["id"]) for o in existentes]
                return str(max(nums) + 1)
            except Exception:
                return str(len(organizacoes) + 1)

        # Adiciona
        def adicionar():
            nome = ent_new.get().strip()
            if not nome:
                messagebox.showwarning("Aviso", "Digite o nome da organização.")
                return

            novo_id = gerar_proximo_id()
            organizacoes.append({"id": novo_id, "nome": nome})
            save_organizacoes(organizacoes)

            tree.insert("", "end", values=(len(organizacoes), nome))
            ent_new.delete(0, tk.END)

            self._registrar_log("Criou organização", nome)
            messagebox.showinfo("OK", f"Organização '{nome}' cadastrada com ID {novo_id}.")

        ctk.CTkButton(control, text="Adicionar", command=adicionar, fg_color="#16A34A", hover_color="#15803D", text_color="white", corner_radius=8, height=34, font=("Segoe UI", 12, "bold")).pack(side="left", padx=6)

        # Edita
        def editar():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Aviso", "Selecione uma organização.")
                return

            _, nome_atual = tree.item(sel[0], "values")

            janela = self._open_inline_container("Editar Organização", voltar=self._open_gerenciar_organizacoes)
            janela.title("Editar Organização")
            janela.geometry("450x220")
            janela.resizable(False, False)
            janela.grab_set()

            frame_edit = ctk.CTkFrame(janela, fg_color="transparent")
            frame_edit.pack(fill="both", expand=True)

            ctk.CTkLabel(
                frame_edit,
                text="Editar Organização",
                font=("Segoe UI", 14, "bold"),
                text_color=("#0F172A", "#E2E8F0")
            ).pack(pady=(0, 15))

            ctk.CTkLabel(frame_edit, text="Nome da organização:", text_color=("#334155", "#CBD5E1")).pack(anchor="w")

            ent_nome_edit = ctk.CTkEntry(frame_edit, width=380)
            ent_nome_edit.pack(pady=8)
            ent_nome_edit.insert(0, nome_atual)
            ent_nome_edit.focus()

            # Salva edicao
            def salvar_edicao():
                novo_nome = ent_nome_edit.get().strip()

                if not novo_nome:
                    messagebox.showwarning("Aviso", "Digite o nome da organização.")
                    return

                for o in organizacoes:
                    if o["nome"].lower() == novo_nome.lower() and o["nome"] != nome_atual:
                        messagebox.showerror(
                            "Erro",
                            "Já existe uma organização com esse nome.\n\n"
                            "Escolha um nome diferente."
                        )
                        return

                for o in organizacoes:
                    if o["nome"] == nome_atual:
                        o["nome"] = novo_nome
                        break

                save_organizacoes(organizacoes)
                self._registrar_log("Editou organização", novo_nome)

                messagebox.showinfo("OK", "Organização atualizada com sucesso.")
                self._fechar_tela_interna(janela)

            botoes = ctk.CTkFrame(frame_edit, fg_color="transparent")
            botoes.pack(pady=15)

            ctk.CTkButton(
                botoes,
                text="Salvar",
                command=salvar_edicao,
                fg_color="#16A34A",
                hover_color="#15803D",
                text_color="white",
                corner_radius=8,
                height=34,
                font=("Segoe UI", 12, "bold")
            ).pack(side="left", padx=8)

        # Exclui
        def excluir():
            if not self._autorizar_admin():
                return

            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Aviso", "Selecione uma organização.")
                return
            _, nome = tree.item(sel[0], "values")
            if messagebox.askyesno("Confirmar", f"Excluir organização '{nome}'?"):
                indice = tree.index(sel[0])
                organizacoes.pop(indice)
                save_organizacoes(organizacoes)
                tree.delete(sel[0])
                self._registrar_log("Excluiu organização", nome)

        btns = ctk.CTkFrame(frm, fg_color="transparent")
        btns.pack(fill="x")
        ctk.CTkButton(btns, text="Editar", command=editar, fg_color="#2563EB", hover_color="#1D4ED8", text_color="white", corner_radius=8, height=34, font=("Segoe UI", 12, "bold")).pack(side="left", padx=6, pady=6)
        ctk.CTkButton(btns, text="Excluir", command=excluir, fg_color="#E53935", hover_color="#C62828", text_color="white", corner_radius=8, height=34, font=("Segoe UI", 12, "bold")).pack(side="left", padx=6, pady=6)

    # Abre gerenciar vans
    def _open_gerenciar_vans(self):
        if self.perfil != "admin":
            messagebox.showerror("Acesso negado", "Somente administradores podem gerenciar vans.")
            return

        top = self._open_inline_container("Gerenciar Vans", voltar=self._build_home_screen)
        top.title("Gerenciar Vans")
        top.geometry("700x420")
        top.grab_set()

        frm = ctk.CTkFrame(top, fg_color="transparent")
        frm.pack(fill="both", expand=True)

        control = ctk.CTkFrame(frm, fg_color="transparent")
        control.pack(fill="x")
        ctk.CTkLabel(control, text="Nome da Van:", text_color=("#334155", "#CBD5E1")).pack(side="left", padx=6)
        ent_new = ctk.CTkEntry(control, width=380)
        ent_new.pack(side="left", padx=6)

        cols = ("id", "nome", "ativa")
        tree = tb.Treeview(frm, columns=cols, show="headings", height=18)
        for c in cols:
            tree.heading(c, text=c.upper())
        tree.column("id", width=100)
        tree.column("nome", width=320)
        tree.column("ativa", width=100)
        tree.pack(fill="both", expand=True, pady=8)
        self._preparar_tabela(tree, minimo=80, maximo=520)

        def carregar_vans():
            if not tree.winfo_exists():
                return
            tree.delete(*tree.get_children())
            vans = load_vans()
            for idx, v in enumerate(vans, start=1):
                ativo = "Sim" if v.get("ativa", True) else "Não"
                tree.insert("", "end", values=(idx, v.get("nome", ""), ativo), tags=(v.get("id", ""),))

        carregar_vans()

        def gerar_proximo_id():
            vans = load_vans()
            existentes = [v for v in vans if str(v.get("id", "")).isdigit()]
            if not existentes:
                return "1"
            try:
                nums = [int(v["id"]) for v in existentes]
                return str(max(nums) + 1)
            except Exception:
                return str(len(vans) + 1)

        def adicionar():
            nome = ent_new.get().strip()
            if not nome:
                messagebox.showwarning("Aviso", "Digite o nome da van.")
                return

            vans = load_vans()
            novo_id = gerar_proximo_id()
            vans.append({"id": novo_id, "nome": nome, "descricao": "", "ativa": True})
            save_vans(vans)

            carregar_vans()
            ent_new.delete(0, tk.END)
            self._registrar_log("Criou van", nome)
            messagebox.showinfo("OK", f"Van '{nome}' cadastrada com ID {novo_id}.")

        ctk.CTkButton(control, text="Adicionar", command=adicionar, fg_color="#16A34A", hover_color="#15803D", text_color="white", corner_radius=8, height=34, font=("Segoe UI", 12, "bold")).pack(side="left", padx=6)

        def editar():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Aviso", "Selecione uma van.")
                return

            item = sel[0]
            van_id = tree.item(item, "tags")[0]
            vans = load_vans()
            van = next((v for v in vans if str(v.get("id")) == str(van_id)), None)
            if not van:
                return

            nome_atual = van.get("nome", "")

            janela = self._open_inline_container("Editar Van", voltar=self._open_gerenciar_vans)
            janela.title("Editar Van")
            janela.geometry("450x220")
            janela.resizable(False, False)
            janela.grab_set()

            frame_edit = ctk.CTkFrame(janela, fg_color="transparent")
            frame_edit.pack(fill="both", expand=True)

            ctk.CTkLabel(
                frame_edit,
                text="Editar Van",
                font=("Segoe UI", 14, "bold"),
                text_color=("#0F172A", "#E2E8F0")
            ).pack(pady=(0, 15))

            ctk.CTkLabel(frame_edit, text="Nome da van:", text_color=("#334155", "#CBD5E1")).pack(anchor="w")

            ent_nome_edit = ctk.CTkEntry(frame_edit, width=380)
            ent_nome_edit.pack(pady=8)
            ent_nome_edit.insert(0, nome_atual)
            ent_nome_edit.focus()

            def salvar_edicao():
                novo_nome = ent_nome_edit.get().strip()
                if not novo_nome:
                    messagebox.showwarning("Aviso", "Digite o nome da van.")
                    return

                for v in vans:
                    if v["nome"].lower() == novo_nome.lower() and v["nome"] != nome_atual:
                        messagebox.showerror(
                            "Erro",
                            "Já existe uma van com esse nome.\n\nEscolha um nome diferente."
                        )
                        return

                for v in vans:
                    if v["nome"] == nome_atual:
                        v["nome"] = novo_nome
                        break

                save_vans(vans)
                self._registrar_log("Editou van", novo_nome)
                messagebox.showinfo("OK", "Van atualizada com sucesso.")
                self._fechar_tela_interna(janela)
                carregar_vans()

            botoes = ctk.CTkFrame(frame_edit, fg_color="transparent")
            botoes.pack(pady=15)

            ctk.CTkButton(
                botoes,
                text="Salvar",
                command=salvar_edicao,
                fg_color="#16A34A",
                hover_color="#15803D",
                text_color="white",
                corner_radius=8,
                height=34,
                font=("Segoe UI", 12, "bold")
            ).pack(side="left", padx=8)

        def excluir():
            if not self._autorizar_admin():
                return

            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Aviso", "Selecione uma van.")
                return
            _, nome, _ = tree.item(sel[0], "values")
            if messagebox.askyesno("Confirmar", f"Excluir van '{nome}'?\n\nOs cadastros vinculados continuarão com o nome histórico."):
                van_id = tree.item(sel[0], "tags")[0]
                vans = load_vans()
                vans = [v for v in vans if str(v.get("id")) != str(van_id)]
                save_vans(vans)
                carregar_vans()
                self._registrar_log("Excluiu van", nome)

        def alternar_status():
            if not self._autorizar_admin():
                return

            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Aviso", "Selecione uma van.")
                return
            van_id = tree.item(sel[0], "tags")[0]
            vans = load_vans()
            van = next((v for v in vans if str(v.get("id")) == str(van_id)), None)
            if not van:
                return
            van["ativa"] = not van.get("ativa", True)
            save_vans(vans)
            carregar_vans()
            status = "ativada" if van["ativa"] else "desativada"
            self._registrar_log("Alterou status da van", f"{van.get('nome')} - {status}")

        btns = ctk.CTkFrame(frm, fg_color="transparent")
        btns.pack(fill="x")
        ctk.CTkButton(btns, text="Editar", command=editar, fg_color="#2563EB", hover_color="#1D4ED8", text_color="white", corner_radius=8, height=34, font=("Segoe UI", 12, "bold")).pack(side="left", padx=6, pady=6)
        ctk.CTkButton(btns, text="Excluir", command=excluir, fg_color="#E53935", hover_color="#C62828", text_color="white", corner_radius=8, height=34, font=("Segoe UI", 12, "bold")).pack(side="left", padx=6, pady=6)
        ctk.CTkButton(btns, text="Ativar/Desativar", command=alternar_status, fg_color="#D97706", hover_color="#B45309", text_color="white", corner_radius=8, height=34, font=("Segoe UI", 12, "bold")).pack(side="left", padx=6, pady=6)

    # Existe ação mesma data
    def _existe_acao_mesma_data(self, data, ignorar_id=None):
        data = (data or "").strip()
        for a in self.db.get("acoes", []):
            if ignorar_id is not None and str(a.get("id")) == str(ignorar_id):
                continue
            if (a.get("data", "") or "").strip() == data:
                return True
        return False

    # Ação duplicada mesma data local
    def _acao_duplicada_mesma_data_local(self, data, local, ignorar_id=None):
        data = (data or "").strip()
        local_norm = self._normalizar_texto_ordenacao(local)
        for a in self.db.get("acoes", []):
            if ignorar_id is not None and str(a.get("id")) == str(ignorar_id):
                continue
            if (a.get("data", "") or "").strip() == data:
                if self._normalizar_texto_ordenacao(a.get("local", "")) == local_norm:
                    return True
        return False

    # Abre gerenciar acoes
    def _open_gerenciar_acoes(self):
        top = self._open_inline_container("Gerenciar Ações", voltar=self._build_home_screen)
        top.title("Gerenciar Ações")
        top.geometry("1200x800")
        top.grab_set()
        frm = ctk.CTkFrame(top, fg_color="transparent")
        frm.pack(fill="both", expand=True)

        ctk.CTkLabel(frm, text="Criar nova Ação", font=("Segoe UI", 13, "bold"), text_color=("#0F172A", "#E2E8F0")).pack(anchor="w", pady=6)
        inp = ctk.CTkFrame(frm, fg_color="transparent")
        inp.pack(fill="x", pady=6)
        inp.columnconfigure(1, weight=1)
        inp.columnconfigure(3, weight=1)
        inp.columnconfigure(5, weight=3)

        ctk.CTkLabel(inp, text="Data (DD/MM/AAAA):", text_color=("#334155", "#CBD5E1")).grid(row=0, column=0, sticky="w", padx=6)
        ent_data = ctk.CTkEntry(inp, width=130)
        ent_data.grid(row=0, column=1, sticky="ew", padx=(0, 12))
        ent_data.insert(0, hoje_str())
        ent_data.bind("<KeyRelease>", lambda e: mascara_data_entrada(ent_data))

        organizacoes_cadastradas = load_organizacoes()
        ctk.CTkLabel(inp, text="Organizações:", text_color=("#334155", "#CBD5E1")).grid(row=0, column=2, sticky="w", padx=6)
        combobox_orgs = ctk.CTkComboBox(inp, values=[""] + [o.get("nome", "") for o in organizacoes_cadastradas], width=240)
        combobox_orgs.grid(row=0, column=3, sticky="ew", padx=(0, 12))

        ctk.CTkLabel(inp, text="Local:", text_color=("#334155", "#CBD5E1")).grid(row=0, column=4, sticky="w", padx=6)
        ent_local = ctk.CTkEntry(inp, width=280)
        ent_local.grid(row=0, column=5, sticky="ew", padx=(0, 12))
        ent_local.bind("<KeyRelease>", lambda e: self._forcar_maiusculo_entry(ent_local))

        self._sugestoes_acao_popup = None
        self._sugestoes_acao_listbox = None

        # Esconder sugestoes
        def esconder_sugestoes():
            if self._sugestoes_acao_popup and self._sugestoes_acao_popup.winfo_exists():
                self._sugestoes_acao_popup.withdraw()

        # Aplicar sugestao
        def aplicar_sugestao(event=None):
            lb = self._sugestoes_acao_listbox
            if not lb or not lb.winfo_exists():
                return
            sel = lb.curselection()
            if not sel:
                return
            valor = lb.get(sel[0])
            ent_local.delete(0, tk.END)
            ent_local.insert(0, valor)
            esconder_sugestoes()
            ent_local.focus_set()

        # Mostra sugestoes
        def mostrar_sugestoes(sugestoes):
            if self._sugestoes_acao_popup is None or not self._sugestoes_acao_popup.winfo_exists():
                self._sugestoes_acao_popup = tk.Toplevel(self)
                self._sugestoes_acao_popup.wm_overrideredirect(True)
                self._sugestoes_acao_popup.attributes("-topmost", True)
                self._sugestoes_acao_listbox = tk.Listbox(
                    self._sugestoes_acao_popup,
                    font=("Segoe UI", 10),
                    activestyle="none"
                )
                self._sugestoes_acao_listbox.pack(fill="both", expand=True)
                self._sugestoes_acao_listbox.bind("<ButtonRelease-1>", aplicar_sugestao)
                self._sugestoes_acao_listbox.bind("<Return>", aplicar_sugestao)

            lb = self._sugestoes_acao_listbox
            lb.delete(0, tk.END)
            for s in sugestoes:
                lb.insert(tk.END, s)
            lb.configure(height=min(6, len(sugestoes)))

            ent_local.update_idletasks()
            x = ent_local.winfo_rootx()
            y = ent_local.winfo_rooty() + ent_local.winfo_height()
            w = ent_local.winfo_width()
            self._sugestoes_acao_popup.geometry(f"{w}x{min(120, 20 * len(sugestoes) + 4)}+{x}+{y}")
            self._sugestoes_acao_popup.deiconify()
            self._sugestoes_acao_popup.lift()

        # Atualiza sugestoes
        def atualizar_sugestoes(event=None):
            texto = ent_local.get().strip().upper()
            if not texto:
                esconder_sugestoes()
                return
            locais_unicos = []
            for a in self.db.get("acoes", []):
                loc = (a.get("local", "") or "").strip()
                if loc and loc not in locais_unicos:
                    locais_unicos.append(loc)
            sugestoes = [loc for loc in locais_unicos if texto in loc.upper()]
            if not sugestoes:
                esconder_sugestoes()
                return
            mostrar_sugestoes(sugestoes)

        # Ao perder foco ação
        def ao_perder_foco_acao(event=None):
            self.after(150, esconder_sugestoes)

        ent_local.bind("<KeyRelease>", lambda e: atualizar_sugestoes(), add="+")
        ent_local.bind("<Escape>", lambda e: esconder_sugestoes())
        ent_local.bind("<FocusOut>", ao_perder_foco_acao)

        ctk.CTkLabel(inp, text="Endereço:", text_color=("#334155", "#CBD5E1")).grid(row=1, column=0, sticky="w", padx=6, pady=(6, 0))
        ent_endereco = ctk.CTkEntry(inp, width=420)
        ent_endereco.grid(row=1, column=1, columnspan=5, sticky="ew", padx=(0, 6), pady=(6, 0))

        svc_frame = ctk.CTkFrame(frm, fg_color="transparent")
        svc_frame.pack(fill="both", pady=6)
        ctk.CTkLabel(svc_frame, text="Serviços disponíveis (Ctrl+Clique para múltiplos):", text_color=("#334155", "#CBD5E1")).pack(anchor="w")
        listbox = tk.Listbox(svc_frame, selectmode="extended", height=6)
        listbox.pack(fill="x", padx=6, pady=6)

        servicos = self.db.get("servicos", [])
        for s in servicos:
            listbox.insert("end", f"{s['id']}|{s['nome']}")

        van_frame = ctk.CTkFrame(frm, fg_color="transparent")
        van_frame.pack(fill="both", pady=6)
        ctk.CTkLabel(van_frame, text="Van autorizada:", text_color=("#334155", "#CBD5E1")).pack(anchor="w")
        combo_vans = tb.Combobox(van_frame, values=[], width=50, state="readonly")
        combo_vans.pack(fill="x", padx=6, pady=6)

        vans = load_vans()
        vans_por_nome = {}
        for v in vans:
            if v.get("ativa", True):
                vans_por_nome[v.get("nome", "")] = v.get("id", "")
        combo_vans.configure(values=list(vans_por_nome.keys()))

        # Adiciona ação
        def adicionar_acao():
            data = ent_data.get().strip()
            organizacao = combobox_orgs.get().strip()
            local = ent_local.get().strip().upper()
            endereco = ent_endereco.get().strip()

            if not data or not local:
                messagebox.showwarning("Aviso", "Preencha os campos obrigatórios.")
                return

            if not organizacao:
                messagebox.showwarning("Aviso", "Selecione uma organização.")
                return

            if not endereco:
                messagebox.showwarning("Aviso", "Informe o endereço da ação.")
                return

            sel = listbox.curselection()
            if not sel:
                messagebox.showwarning("Aviso", "Selecione ao menos um serviço.")
                return
            serv_ids = []
            for idx in sel:
                v = listbox.get(idx)
                sid = v.split("|", 1)[0]
                serv_ids.append(sid)

            sel_vans = combo_vans.get().strip()
            if not sel_vans:
                messagebox.showwarning("Aviso", "Selecione uma van.")
                return
            van_ids = [vans_por_nome.get(sel_vans, "")] if vans_por_nome.get(sel_vans, "") else []

            if self._acao_duplicada_mesma_data_local(data, local):
                messagebox.showerror(
                    "Ação duplicada",
                    f"Já existe uma ação com o local '{local}' na data {data}.\n\n"
                    "Não é permitido criar ações com o mesmo nome na mesma data."
                )
                return

            if self._existe_acao_mesma_data(data):
                if not messagebox.askyesno(
                    "Ação na mesma data",
                    "Já existe uma ação criada para esta data.\n\n"
                    "Deseja criar outra ação assim mesmo?"
                ):
                    return

            existentes = [a for a in self.db.get("acoes", []) if str(a.get("id","")).isdigit()]
            if not existentes:
                aid = "1"
            else:
                try:
                    nums = [int(a["id"]) for a in existentes]
                    aid = str(max(nums) + 1)
                except Exception:
                    aid = str(len(self.db.get("acoes", [])) + 1)

            nova = {
            "id": aid,
            "data": data,
            "local": local,
            "organizacoes": [organizacao],
            "endereco": endereco,
            "servicos": serv_ids,
            "vans": van_ids
        }
            self.db["acoes"].append(nova)
            save_db(self.db)

            combobox_orgs.set("")
            listbox.selection_clear(0, tk.END)
            ent_endereco.delete(0, tk.END)

            numero_ordem = len(self.db["acoes"])

            item_id = tree.insert(
                "",
                "end",
                values=(
                    numero_ordem,
                    data,
                    local,
                    organizacao,
                    endereco,
                    ", ".join(serv_ids)
                ),
                tags=(aid,)
            )

            tree.selection_set(item_id)
            tree.focus(item_id)
            tree.see(item_id)
            self._registrar_log("Criou ação", f"{aid} - {data} - {local}")
            messagebox.showinfo("OK", "Ação criada com sucesso.")

        btns_acoes = ctk.CTkFrame(frm, fg_color="transparent")
        btns_acoes.pack(fill="x", pady=6)
        ctk.CTkButton(btns_acoes, text="Adicionar Ação", command=adicionar_acao, fg_color="#16A34A", hover_color="#15803D", text_color="white", corner_radius=8, height=38, font=("Segoe UI", 12, "bold")).pack(side="left", padx=6)

        tb.Separator(frm).pack(fill="x", pady=8)
        cols = ("ordem", "data", "local", "organizacao", "endereco", "servicos", "vans")
        tree = tb.Treeview(frm, columns=cols, show="headings", height=12)
        for c in cols:
            tree.heading(c, text=c.upper())
        tree.column("ordem", width=80)
        tree.column("data", width=120)
        tree.column("local", width=200)
        tree.column("organizacao", width=180)
        tree.column("endereco", width=220)
        tree.column("servicos", width=180)
        tree.column("vans", width=180)
        tree.pack(fill="both", expand=True, padx=6, pady=6)
        self._preparar_tabela(tree, minimo=80, maximo=520, auto_ajuste=False)

        # Chave data ação
        def chave_data_acao(a):
            data_str = a.get("data", "01/01/1900")
            try:
                d, m, y = data_str.split("/")
                return (int(y), int(m), int(d))
            except Exception:
                return (1900, 1, 1)

        acoes_ordenadas = sorted(
            self.db.get("acoes", []),
            key=chave_data_acao,
            reverse=True
        )

        total_acoes = len(acoes_ordenadas)

        for indice, a in enumerate(acoes_ordenadas):

            numero_ordem = total_acoes - indice

            servs = a.get("servicos", [])
            orgs = a.get("organizacoes", []) or []
            vans_ids = a.get("vans", []) or []

            if isinstance(servs, str):
                servs = [s.strip() for s in servs.split(",") if s.strip()]

            if isinstance(orgs, str):
                orgs = [o.strip() for o in orgs.split(",") if o.strip()]

            if isinstance(vans_ids, str):
                vans_ids = [v.strip() for v in vans_ids.split(",") if v.strip()]

            vans_nomes = []
            for vid in vans_ids:
                van = next((v for v in vans if v.get("id") == vid), None)
                if van:
                    vans_nomes.append(van.get("nome", ""))

            tree.insert(
                "",
                "end",
                values=(
                    numero_ordem,
                    a.get("data", ""),
                    a.get("local", ""),
                    ", ".join(orgs),
                    a.get("endereco", ""),
                    ", ".join(servs),
                    ", ".join(vans_nomes)
                ),
                tags=(a["id"],)
            )

        acao_em_edicao = {"id": None, "item": None}

        # Seleciona servicos ação
        def selecionar_servicos_acao(servicos_ids):
            listbox.selection_clear(0, tk.END)
            servicos_ids = [str(sid).strip() for sid in servicos_ids]
            for i in range(listbox.size()):
                sid = listbox.get(i).split("|", 1)[0]
                if sid in servicos_ids:
                    listbox.select_set(i)

        # Limpa modo edicao ação
        def limpar_modo_edicao_acao():
            acao_em_edicao["id"] = None
            acao_em_edicao["item"] = None
            ent_data.delete(0, tk.END)
            ent_data.insert(0, hoje_str())
            ent_local.delete(0, tk.END)
            combobox_orgs.set("")
            listbox.selection_clear(0, tk.END)
            btn_salvar_alteracao.pack_forget()

        # Edita ação
        def editar_acao():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Aviso", "Selecione uma ação na lista para editar.")
                return

            item = sel[0]
            aid = tree.item(item, "tags")[0]
            acao = next((a for a in self.db.get("acoes", []) if str(a.get("id")) == str(aid)), None)

            if not acao:
                messagebox.showerror("Erro", "Ação não encontrada.")
                return

            acao_em_edicao["id"] = aid
            acao_em_edicao["item"] = item
            ent_data.delete(0, tk.END)
            ent_data.insert(0, acao.get("data", ""))
            ent_local.delete(0, tk.END)
            ent_local.insert(0, acao.get("local", ""))
            ent_endereco.delete(0, tk.END)
            ent_endereco.insert(0, acao.get("endereco", ""))
            selecionar_servicos_acao(acao.get("servicos", []))
            orgs_sel = acao.get("organizacoes", []) or []
            if isinstance(orgs_sel, str):
                orgs_sel = [o.strip() for o in orgs_sel.split(",") if o.strip()]
            combobox_orgs.set(orgs_sel[0] if orgs_sel else "")
            if not btn_salvar_alteracao.winfo_manager():
                btn_salvar_alteracao.pack(side="left", padx=6)

        # Salva alteracao ação
        def salvar_alteracao_acao():
            aid = acao_em_edicao["id"]
            item = acao_em_edicao["item"]
            if not aid or not item:
                messagebox.showwarning("Aviso", "Clique em Editar Ação e selecione uma ação primeiro.")
                return

            new_data = ent_data.get().strip()
            new_local = ent_local.get().strip()
            new_endereco = ent_endereco.get().strip()
            if not new_data or not new_local:
                messagebox.showwarning("Aviso", "Informe data e local.")
                return

            new_orgs = [combobox_orgs.get().strip()] if combobox_orgs.get().strip() else []
            new_servs = [listbox.get(i).split("|", 1)[0] for i in listbox.curselection()]
            if not new_servs:
                messagebox.showwarning("Aviso", "Selecione ao menos um serviço.")
                return

            if self._acao_duplicada_mesma_data_local(new_data, new_local, ignorar_id=aid):
                messagebox.showerror(
                    "Ação duplicada",
                    f"Já existe uma ação com o local '{new_local}' na data {new_data}.\n\n"
                    "Não é permitido criar ações com o mesmo nome na mesma data."
                )
                return
            for a in self.db["acoes"]:
                if str(a.get("id")) == str(aid):
                    a["data"] = new_data
                    a["local"] = new_local
                    a["organizacoes"] = new_orgs
                    a["servicos"] = new_servs
                    break

            save_db(self.db)
            ordem = tree.item(item, "values")[0]
            tree.item(
                item,
                values=(
                    ordem,
                    new_data,
                    new_local,
                    ", ".join(new_orgs),
                    new_endereco,
                    ", ".join(new_servs)
                )
            )
            self._registrar_log("Editou ação", aid)
            messagebox.showinfo("OK", "Ação atualizada com sucesso.")
            limpar_modo_edicao_acao()

        # Exclui ação
        def excluir_acao():
            if not self._autorizar_admin():
                return

            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Aviso", "Selecione uma ação.")
                return
            item_data = tree.item(sel[0])
            aid = str(item_data["tags"][0])
            valores = item_data["values"]
            data_old = valores[1]
            local_old = valores[2]

            id_acao_ativa = self._id_acao_ativa_atual()
            if id_acao_ativa and id_acao_ativa == aid:
                messagebox.showwarning(
                    "Ação ativa",
                    "Esta ação está ativa no momento e não pode ser excluída.\n"
                    "Finalize ou troque a ação ativa antes de excluí-la."
                )
                return

            cadastros_vinculados = sum(
                1 for cadastro in self.db.get("cadastros", [])
                if str(cadastro.get("acao_id", "")).strip() == aid
            )
            if cadastros_vinculados:
                messagebox.showwarning(
                    "Ação vinculada",
                    f"Esta ação possui {cadastros_vinculados} cadastro(s) vinculado(s) e não pode ser excluída."
                )
                return

            if messagebox.askyesno("Confirmar", f"Excluir ação em {data_old} - {local_old}?"):
                acao_excluida = next(
                    (a for a in self.db["acoes"] if str(a.get("id")) == aid),
                    None
                )
                if acao_excluida:
                    salvar_excluido(
                        DB_EXCLUIDOS_ACOES,
                        "acao",
                        acao_excluida,
                        self.usuario
                    )
                    self.db["acoes"] = [a for a in self.db["acoes"] if str(a.get("id")) != aid]
                    save_db(self.db)
                    tree.delete(sel[0])
                    self._registrar_log("Excluiu ação", acao_excluida)
                else:
                    messagebox.showerror("Erro", "Ação não encontrada no banco de dados.")
        ctk.CTkButton(
            btns_acoes,
            text="Editar Ação",
            command=editar_acao,
            fg_color="#2563EB",
            hover_color="#1D4ED8",
            text_color="white",
            corner_radius=8,
            height=38,
            font=("Segoe UI", 12, "bold")
        ).pack(side="left", padx=6)

        ctk.CTkButton(
            btns_acoes,
            text="Excluir Ação",
            command=excluir_acao,
            fg_color="#E53935",
            hover_color="#C62828",
            text_color="white",
            corner_radius=8,
            height=38,
            font=("Segoe UI", 12, "bold")
        ).pack(side="left", padx=6)

        btn_salvar_alteracao = ctk.CTkButton(
            btns_acoes,
            text="Salvar Alteração",
            command=salvar_alteracao_acao,
            fg_color="#16A34A",
            hover_color="#15803D",
            text_color="white",
            corner_radius=8,
            height=38,
            font=("Segoe UI", 12, "bold")
        )

    # Abre seleção ação serviço
    def _abrir_acao_somente_leitura(self):
        if not self.db.get("acoes"):
            messagebox.showinfo("Aviso", "Não há ações cadastradas pelo administrador.")
            return
        top = self._open_inline_container("Selecionar Ação e Serviço", voltar=self._build_cadastro_screen)
        top.title("Selecionar Ação e Serviço")
        top.geometry("640x360")
        top.grab_set()
        frm = ctk.CTkFrame(top, fg_color="transparent")
        frm.pack(fill="both", expand=True)

        ctk.CTkLabel(frm, text="Selecione a data da ação:", font=("Segoe UI", 11), text_color=("#334155", "#CBD5E1")).pack(anchor="w", pady=6)
        datas = sorted(
            {a["data"] for a in self.db.get("acoes", [])},
            key=lambda d: datetime.strptime(d, "%d/%m/%Y"),
            reverse=True
        )
        self.sel_data_var = tk.StringVar()
        comb_data = tb.Combobox(frm, values=datas, textvariable=self.sel_data_var)
        comb_data.pack(fill="x", pady=6)

        ctk.CTkLabel(frm, text="Selecione a ação (local):", font=("Segoe UI", 11), text_color=("#334155", "#CBD5E1")).pack(anchor="w", pady=6)
        self.sel_acao_var = tk.StringVar()
        comb_acao = tb.Combobox(frm, values=[], textvariable=self.sel_acao_var)
        comb_acao.pack(fill="x", pady=6)

        ctk.CTkLabel(frm, text="Selecione o serviço:", font=("Segoe UI", 11), text_color=("#334155", "#CBD5E1")).pack(anchor="w", pady=6)
        self.sel_servico_var = tk.StringVar()
        comb_serv = tb.Combobox(frm, values=[], textvariable=self.sel_servico_var)
        comb_serv.pack(fill="x", pady=6)

        def on_data_change(event=None):
            d = self.sel_data_var.get()
            acoes = [a for a in self.db.get("acoes", []) if a["data"] == d]
            if not acoes:
                comb_acao.configure(values=[])
                comb_serv.configure(values=[])
                return
            try:
                locais_unicos = {}
                for a in acoes:
                    loc = a["local"]
                    if loc not in locais_unicos:
                        locais_unicos[loc] = []
                    locais_unicos[loc].append(a)
                self._acoes_por_local = locais_unicos

                values = list(locais_unicos.keys())
                comb_acao.configure(values=values)

                ultima_acao = acoes[-1]
                local_padrao = ultima_acao["local"]
                comb_acao.set(local_padrao)

                servs = []
                for sid in ultima_acao.get("servicos", []):
                    s = next(
                        (s for s in self.db.get("servicos", []) if s["id"] == sid),
                        None
                    )
                    if s:
                        servs.append(f"{s['id']}|{s['nome']}")

                comb_serv.configure(values=servs)

                servico_padrao = ""
                for s in servs:
                    nome = s.split("|", 1)[1].strip().lower()
                    if nome == "identidade - cin":
                        servico_padrao = s
                        break

                if not servico_padrao and servs:
                    servico_padrao = servs[0]

                comb_serv.set(servico_padrao)
            except Exception:
                pass

        def on_acao_change(event=None):
            val = self.sel_acao_var.get()
            if not val:
                comb_serv.configure(values=[])
                return
            local_sel = self.sel_acao_var.get()

            acoes_local = self._acoes_por_local.get(local_sel, [])
            servicos = []
            for ac in acoes_local:
                for sid in ac.get("servicos", []):
                    s = next((sv for sv in self.db.get("servicos", []) if sv["id"] == sid), None)
                    if s:
                        servicos.append(f"{s['id']}|{s['nome']}")
            comb_serv.configure(values=servicos)
            if servicos:
                comb_serv.current(0)

        def seguir():
            d = self.sel_data_var.get().strip()
            local_sel = self.sel_acao_var.get().strip()
            servico_val = self.sel_servico_var.get().strip()
            if not d or not local_sel or not servico_val:
                messagebox.showwarning("Aviso", "Selecione data, ação e serviço.")
                return
            acoes_local = self._acoes_por_local.get(local_sel, [])
            if not acoes_local:
                messagebox.showerror("Erro", "Ação inválida.")
                return
            sid = servico_val.split("|", 1)[0].strip()
            acao_escolhida = None
            for ac in acoes_local:
                if sid in ac["servicos"]:
                    acao_escolhida = ac
                    break
            if acao_escolhida is None:
                messagebox.showerror("Erro", "Serviço inválido para esta ação.")
                return
            servico = next((s for s in self.db["servicos"] if s["id"] == sid), None)
            if not servico:
                messagebox.showerror("Erro", "Serviço inválido.")
                return
            self._selected_acao = {
                "id": acao_escolhida["id"],
                "data": acao_escolhida["data"],
                "local": acao_escolhida["local"],
                "vans": acao_escolhida.get("vans", [])
            }
            self._selected_servico = {
                "id": servico["id"],
                "nome": servico["nome"]
            }
            self._fechar_tela_interna(top)
            self._build_cadastro_screen()

        comb_data.bind("<<ComboboxSelected>>", on_data_change)
        comb_acao.bind("<<ComboboxSelected>>", on_acao_change)
        if datas:
            comb_data.set(datas[0])
            on_data_change()

        ctk.CTkButton(frm, text="Abrir Ação", command=seguir, fg_color="#2563EB", hover_color="#1D4ED8", text_color="white", corner_radius=8, height=38, font=("Segoe UI", 12, "bold")).pack(pady=12)

    def _open_selecao_acao_servico(self):

        if not self.db.get("acoes"):
            messagebox.showinfo("Aviso", "Não há ações cadastradas pelo administrador.")
            return
        top = self._open_inline_container("Selecionar Ação e Serviço", voltar=self._build_cadastro_screen)
        top.title("Selecionar Ação e Serviço")
        top.geometry("640x360")
        top.grab_set()
        frm = ctk.CTkFrame(top, fg_color="transparent")
        frm.pack(fill="both", expand=True)

        ctk.CTkLabel(frm, text="Selecione a data da ação:", font=("Segoe UI", 11), text_color=("#334155", "#CBD5E1")).pack(anchor="w", pady=6)
        datas = sorted(
            {a["data"] for a in self.db.get("acoes", [])},
            key=lambda d: datetime.strptime(d, "%d/%m/%Y"),
            reverse=True
        )
        self.sel_data_var = tk.StringVar()
        comb_data = tb.Combobox(frm, values=datas, textvariable=self.sel_data_var)
        comb_data.pack(fill="x", pady=6)

        ctk.CTkLabel(frm, text="Selecione a ação (local):", font=("Segoe UI", 11), text_color=("#334155", "#CBD5E1")).pack(anchor="w", pady=6)
        self.sel_acao_var = tk.StringVar()
        comb_acao = tb.Combobox(frm, values=[], textvariable=self.sel_acao_var)
        comb_acao.pack(fill="x", pady=6)

        ctk.CTkLabel(frm, text="Selecione o serviço:", font=("Segoe UI", 11), text_color=("#334155", "#CBD5E1")).pack(anchor="w", pady=6)
        self.sel_servico_var = tk.StringVar()
        comb_serv = tb.Combobox(frm, values=[], textvariable=self.sel_servico_var)
        comb_serv.pack(fill="x", pady=6)

        # Ao data change
        def on_data_change(event=None):

            d = self.sel_data_var.get()

            acoes = [
                a for a in self.db.get("acoes", [])
                if a["data"] == d
            ]

            if not acoes:
                comb_acao.configure(values=[])
                comb_serv.configure(values=[])
                return

            try:
                acoes = sorted(
                    acoes,
                    key=lambda x: int(x.get("id", 0))
                )
            except:
                pass

            locais_unicos = {}

            for a in acoes:
                loc = a["local"]

                if loc not in locais_unicos:
                    locais_unicos[loc] = []

                locais_unicos[loc].append(a)

            self._acoes_por_local = locais_unicos

            values = list(locais_unicos.keys())

            comb_acao.configure(values=values)


            ultima_acao = acoes[-1]
            local_padrao = ultima_acao["local"]

            comb_acao.set(local_padrao)


            servs = []

            for sid in ultima_acao.get("servicos", []):

                s = next(
                    (s for s in self.db.get("servicos", []) if s["id"] == sid),
                    None
                )

                if s:
                    servs.append(f"{s['id']}|{s['nome']}")

            comb_serv.configure(values=servs)


            servico_padrao = ""

            for s in servs:

                nome = s.split("|", 1)[1].strip().lower()

                if nome == "identidade - cin":
                    servico_padrao = s
                    break

            if not servico_padrao and servs:
                servico_padrao = servs[0]

            comb_serv.set(servico_padrao)

        # Ao ação change
        def on_acao_change(event=None):
            val = self.sel_acao_var.get()
            if not val:
                comb_serv.configure(values=[])
                return
            local_sel = self.sel_acao_var.get()

            acoes_local = self._acoes_por_local.get(local_sel, [])

            servs_ids = set()
            for ac in acoes_local:
                for sid in ac.get("servicos", []):
                    servs_ids.add(sid)

            servs = []
            for sid in servs_ids:
                s = next((s for s in self.db.get("servicos", []) if s["id"] == sid), None)
                if s:
                    servs.append(f"{s['id']}|{s['nome']}")

            comb_serv.configure(values=servs)
            comb_serv.set("")

        comb_data.bind("<<ComboboxSelected>>", on_data_change)
        comb_acao.bind("<<ComboboxSelected>>", on_acao_change)

        # Seguir
        def seguir():
          d = self.sel_data_var.get().strip()
          local_sel = self.sel_acao_var.get().strip()
          servico_val = self.sel_servico_var.get().strip()

          if not d or not local_sel or not servico_val:
              messagebox.showwarning("Aviso", "Selecione data, ação e serviço.")
              return

          acoes_local = self._acoes_por_local.get(local_sel, [])

          if not acoes_local:
              messagebox.showerror("Erro", "Ação inválida.")
              return

          sid = servico_val.split("|", 1)[0]

          acao_escolhida = None
          for ac in acoes_local:
              if sid in ac["servicos"]:
                  acao_escolhida = ac
                  break

          if acao_escolhida is None:
              messagebox.showerror("Erro", "Serviço inválido para esta ação.")
              return

          servico = next((s for s in self.db["servicos"] if s["id"] == sid), None)
          if not servico:
              messagebox.showerror("Erro", "Serviço inválido.")
              return

          self._selected_acao = {
              "id": acao_escolhida["id"],
              "data": acao_escolhida["data"],
              "local": acao_escolhida["local"],
              "vans": acao_escolhida.get("vans", [])
          }

          self._selected_servico = {
              "id": servico["id"],
              "nome": servico["nome"]
          }

          if self.perfil != "admin":
              usuario = self.usuarios.get(self.usuario or "", {})
              vans_permitidas = usuario.get("vans_permitidas", [])
              van_ativa = usuario.get("van_ativa", "")
              vans_acao = acao_escolhida.get("vans", []) or []

              if vans_acao and van_ativa not in vans_acao:
                  nome_acao = f"{acao_escolhida.get('data', '')} - {acao_escolhida.get('local', '')}"
                  vans_acao_nomes = []
                  for vid in vans_acao:
                      van = next((v for v in self.vans if v.get("id") == vid), None)
                      if van:
                          vans_acao_nomes.append(van.get("nome", ""))
                  if not vans_acao_nomes:
                      vans_acao_nomes = ["Van não identificada"]

                  if not vans_permitidas:
                      messagebox.showerror(
                          "Acesso negado",
                          f"Você não está autorizado a trabalhar nesta van.\n\n"
                          f"Ação: {nome_acao}\n"
                          f"Van da ação: {', '.join(vans_acao_nomes)}\n\n"
                          f"Procure o administrador."
                      )
                      return

                  vans_permitidas_nomes = []
                  for vid in vans_permitidas:
                      van = next((v for v in self.vans if v.get("id") == vid), None)
                      if van:
                          vans_permitidas_nomes.append(van.get("nome", ""))

                  if not vans_permitidas_nomes:
                      messagebox.showerror(
                          "Acesso negado",
                          f"Você não está autorizado a trabalhar nesta van.\n\n"
                          f"Ação: {nome_acao}\n"
                          f"Van da ação: {', '.join(vans_acao_nomes)}\n\n"
                          f"Procure o administrador."
                      )
                      return

                  resp = messagebox.askyesno(
                      "Van diferente",
                      f"Você está na {', '.join(vans_permitidas_nomes)}, mas esta ação pertence à {', '.join(vans_acao_nomes)}.\n\n"
                      f"Ação: {nome_acao}\n"
                      f"Deseja trocar para esta van?"
                  )
                  if not resp:
                      return

                  self.usuarios[self.usuario]["van_ativa"] = vans_acao[0]
                  save_usuarios(self.usuarios)

          self._salvar_acao_ativa()

          self._fechar_tela_interna(top)
          self._build_cadastro_screen(preselect=True)
          self._atualizar_estado_botoes_cadastro()

        ctk.CTkButton(frm, text="Ativar Ação", command=seguir, fg_color="#2563EB", hover_color="#1D4ED8", text_color="white", corner_radius=8, height=38, font=("Segoe UI", 12, "bold")).pack(pady=12)

    # Constrói a tela cadastro tela
    def _build_cadastro_screen(self, preselect=False):
        if not preselect and (not self._selected_acao or not self._selected_servico):
            if self.db.get("acao_ativa"):
                self.db["acao_ativa"] = None
                save_db(self.db)

        self._clear_screen()
        self._clear()
        self._ajustar_janela_principal()
        self._protocolo_origem_selecao = ""

        acao_ativa = self.db.get("acao_ativa") or {}
        self._modo_leitura = bool(self._selected_acao and self._selected_servico and not acao_ativa.get("id"))

        wrapper = ctk.CTkFrame(self, fg_color="transparent")
        wrapper.pack(fill="both", expand=True, padx=10, pady=10)

        self._criar_cabecalho(
            wrapper,
            "CADASTRO DE IDENTIDADES",
            voltar=self._build_home_screen,
            mostrar_sair=True,
            texto_cancelar="Voltar"
        )

        frame = ctk.CTkFrame(wrapper, fg_color="transparent")
        frame.pack(fill="both", expand=True, padx=12, pady=(4, 8))

        ctk.CTkLabel(
            frame,
            text="Cadastro de Identidades",
            font=("Segoe UI", 16, "bold"),
            text_color=("#0F172A", "#E2E8F0")
        ).grid(row=0, column=0, columnspan=8, sticky="w", padx=6, pady=(0, 8))

        card_campos = ctk.CTkFrame(
            frame,
            height=118,
            corner_radius=12,
            border_width=1,
            border_color=("#E2E8F0", "#334155"),
            fg_color=("#FFFFFF", "#1E293B")
        )
        card_campos.grid(row=1, column=0, columnspan=8, sticky="ew", padx=6, pady=(0, 2))
        card_campos.grid_propagate(True)
        card_campos.columnconfigure(1, weight=1)

        frame_foto = ctk.CTkFrame(
            card_campos,
            width=128,
            height=180,
            corner_radius=0,
            border_width=0,
            fg_color="transparent"
        )
        frame_foto.grid(row=0, column=0, rowspan=3, sticky="nsw", padx=(0, 14))
        frame_foto.grid_propagate(False)
        self.lbl_foto_cadastro = ctk.CTkLabel(
            frame_foto,
            text="Sem\nfoto",
            width=116,
            height=126,
            fg_color=("#E5E7EB", "#334155"),
            text_color=("#64748B", "#CBD5E1"),
            corner_radius=0,
            font=("Segoe UI", 12, "bold")
        )
        self.lbl_foto_cadastro.pack(padx=6, pady=(6, 4))
        self.btn_capturar_foto_idnet = ctk.CTkButton(
            frame_foto,
            text="Capturar foto",
            command=self._capturar_foto_idnet_manual,
            fg_color="#0F766E",
            hover_color="#115E59",
            text_color="white",
            corner_radius=7,
            height=28,
            font=("Segoe UI", 10, "bold")
        )
        self.btn_capturar_foto_idnet.pack(fill="x", padx=6, pady=(0, 6))
        self._foto_cadastro_path = ""
        self._foto_cadastro_img = None

        linha_campos = ctk.CTkFrame(card_campos, fg_color="transparent")
        linha_campos.grid(row=0, column=1, sticky="ew")
        linha_campos.columnconfigure(0, weight=0)
        linha_campos.columnconfigure(1, weight=1)
        linha_campos.columnconfigure(2, weight=0)
        linha_campos.columnconfigure(3, weight=1)
        linha_campos.columnconfigure(4, weight=0)
        linha_campos.columnconfigure(5, weight=2)

        ctk.CTkLabel(linha_campos, text="Protocolo:", text_color=("#334155", "#CBD5E1")).grid(row=0, column=0, sticky="w", padx=(0, 4))
        self.ent_protocolo = ctk.CTkEntry(linha_campos, width=150, font=("Segoe UI", 12, "bold"))
        self.ent_protocolo.grid(row=1, column=0, sticky="ew", padx=(0, 10))
        self.ent_protocolo.bind("<KeyRelease>", self._agendar_pesquisa_cadastro)

        ctk.CTkLabel(linha_campos, text="CPF:", text_color=("#334155", "#CBD5E1")).grid(row=0, column=1, sticky="w", padx=(0, 4))
        self.ent_cpf = ctk.CTkEntry(linha_campos, width=160, font=("Segoe UI", 12, "bold"))
        self.ent_cpf.grid(row=1, column=1, sticky="ew", padx=(0, 10))
        self.ent_cpf.bind("<KeyRelease>", self._ao_digitar_cpf)

        ctk.CTkLabel(linha_campos, text="Nome:", text_color=("#334155", "#CBD5E1")).grid(row=0, column=2, sticky="w", padx=(0, 4))
        self.ent_nome = ctk.CTkEntry(linha_campos, width=360, font=("Segoe UI", 12, "bold"))
        self.ent_nome.grid(row=1, column=2, columnspan=4, sticky="ew", padx=(0, 0))
        self.ent_nome.bind("<KeyRelease>", self._agendar_pesquisa_cadastro)

        linha_campos2 = ctk.CTkFrame(card_campos, fg_color="transparent")
        linha_campos2.grid(row=1, column=1, sticky="ew")
        linha_campos2.columnconfigure(0, weight=2)

        ctk.CTkLabel(linha_campos2, text="Telefone:", text_color=("#334155", "#CBD5E1")).grid(row=0, column=0, sticky="w", padx=(0, 4))
        frame_tel = ctk.CTkFrame(linha_campos2, fg_color="transparent")
        frame_tel.grid(row=1, column=0, sticky="w", padx=(0, 10))

        self.ent_tel = ctk.CTkEntry(frame_tel, width=180)
        self.ent_tel.bind(
            "<KeyRelease>",
            lambda e: mascara_tel_entrada(self.ent_tel)
        )
        self.ent_tel.pack(side="left")
        self.ent_tel_extra = None

        # Adiciona telefone extra
        def adicionar_telefone_extra():
            if self.ent_tel_extra is not None:
                self.ent_tel_extra.destroy()
                self.ent_tel_extra = None
                btn_add_tel.configure(text="+", bootstyle="success-outline")
                self.ent_tel.focus_set()
                return

            self.ent_tel_extra = tb.Entry(frame_tel, width=18)
            self.ent_tel_extra.bind(
                "<KeyRelease>",
                lambda e: mascara_tel_entrada(self.ent_tel_extra)
            )
            self.ent_tel_extra.pack(side="left", padx=(5, 0), before=chk_sem_tel)
            btn_add_tel.configure(text="-", bootstyle="danger-outline")
            self.ent_tel_extra.focus_set()

        btn_add_tel = tb.Button(
            frame_tel,
            text="+",
            width=3,
            bootstyle="success-outline",
            command=adicionar_telefone_extra
        )
        btn_add_tel.pack(side="left", padx=(5, 0))
        self.btn_add_tel = btn_add_tel
        self._adicionar_telefone_extra = adicionar_telefone_extra
        self.var_sem_telefone = tk.BooleanVar()

        # Toggle telefone
        def toggle_telefone():
            if self.var_sem_telefone.get():
                self.ent_tel.delete(0, tk.END)
                self.ent_tel.configure(state="disabled")
                if self.ent_tel_extra is not None:
                    self.ent_tel_extra.destroy()
                    self.ent_tel_extra = None
                btn_add_tel.configure(text="+", bootstyle="success-outline")
                btn_add_tel.configure(state="disabled")
            else:
                self.ent_tel.configure(state="normal")
                btn_add_tel.configure(state="normal")

        self.chk_sem_tel = tb.Checkbutton(
            frame_tel,
            text="Não informado",
            variable=self.var_sem_telefone,
            command=toggle_telefone,
            bootstyle="secondary"
        )
        self.chk_sem_tel.pack(side="left", padx=(5, 0))

        linha_acao_servico = ctk.CTkFrame(card_campos, fg_color="transparent")
        linha_acao_servico.grid(row=2, column=1, sticky="ew")
        linha_acao_servico.columnconfigure(1, weight=1)
        linha_acao_servico.columnconfigure(3, weight=2)
        linha_acao_servico.columnconfigure(5, weight=1)

        ctk.CTkLabel(linha_acao_servico, text="Data:", text_color=("#334155", "#CBD5E1")).grid(row=0, column=0, sticky="w", padx=(0, 4))
        self.ent_data = ctk.CTkEntry(
            linha_acao_servico,
            width=86,
            height=28,
            fg_color=("#E5E7EB", "#334155"),
            text_color=("#334155", "#CBD5E1")
        )
        self.ent_data.grid(row=0, column=1, sticky="w", padx=(6, 18))
        self.ent_data.insert(0, hoje_str())
        self.ent_data.bind("<KeyRelease>", lambda e: mascara_data_entrada(self.ent_data))

        ctk.CTkLabel(linha_acao_servico, text="Ação:", text_color=("#334155", "#CBD5E1")).grid(row=0, column=2, sticky="w")
        self.ent_acao_info = ctk.CTkEntry(
            linha_acao_servico,
            height=28,
            fg_color=("#E5E7EB", "#334155"),
            text_color=("#334155", "#CBD5E1")
        )
        self.ent_acao_info.grid(row=0, column=3, sticky="ew", padx=(6, 18))

        ctk.CTkLabel(linha_acao_servico, text="Serviço:", text_color=("#334155", "#CBD5E1")).grid(row=0, column=4, sticky="w")
        self.ent_servico_info = ctk.CTkEntry(
            linha_acao_servico,
            height=28,
            fg_color=("#E5E7EB", "#334155"),
            text_color=("#334155", "#CBD5E1")
        )
        self.ent_servico_info.grid(row=0, column=5, sticky="ew", padx=(6, 18))

        self.btn_selecionar_acao_servico = ctk.CTkButton(
            linha_acao_servico,
            text="Selecionar Ação/Serviço",
            command=self._open_selecao_acao_servico,
            fg_color="#2563EB",
            hover_color="#1D4ED8",
            text_color="white",
            corner_radius=8,
            height=34,
            font=("Segoe UI", 11, "bold")
        )
        self.btn_selecionar_acao_servico.grid(row=0, column=6, sticky="e", padx=(10, 0))
        if preselect and hasattr(self, "_selected_acao") and hasattr(self, "_selected_servico"):
            a = self._selected_acao
            s = self._selected_servico
            self.db["acao_ativa"] = {
                "id": a["id"],
                "data": a["data"],
                "servico_id": s.get("id", "") if isinstance(s, dict) else s,
                "vans": a.get("vans", [])
            }
            save_db(self.db)

        self._preencher_campos_acao_ativa()
        btns = tb.Frame(frame)
        btns.grid(row=2, column=0, columnspan=8, sticky="ew", padx=6, pady=(6, 4))

        self.btn_abrir_acao = ctk.CTkButton(btns, text="Abrir Ação", command=self._abrir_acao_somente_leitura, fg_color="#2563EB", hover_color="#1D4ED8", text_color="white", corner_radius=8, height=36, font=("Segoe UI", 12, "bold"))
        self.btn_abrir_acao.pack(side="left", padx=4)

        self.btn_salvar = ctk.CTkButton(btns, text="Salvar", command=self._salvar_ou_atualizar, fg_color="#16A34A", hover_color="#15803D", text_color="white", corner_radius=8, height=36, font=("Segoe UI", 12, "bold"))
        self.btn_salvar.pack(side="left", padx=4)
        self.btn_capturar_idnet = ctk.CTkButton(btns, text="Capturar idNet", command=self._capturar_dados_idnet, fg_color="#0F766E", hover_color="#115E59", text_color="white", corner_radius=8, height=36, font=("Segoe UI", 12, "bold"))
        self.btn_capturar_idnet.pack(side="left", padx=4)
        ctk.CTkButton(btns, text="Pesquisar", command=self._pesquisar, fg_color="#2563EB", hover_color="#1D4ED8", text_color="white", corner_radius=8, height=36, font=("Segoe UI", 12, "bold")).pack(side="left", padx=4)
        ctk.CTkButton(btns, text="Limpar", command=self._limpar_campos, fg_color="#64748B", hover_color="#475569", text_color="white", corner_radius=8, height=36, font=("Segoe UI", 12, "bold")).pack(side="left", padx=4)
        self.btn_excluir = ctk.CTkButton(btns, text="Excluir", command=self._excluir, fg_color="#E53935", hover_color="#C62828", text_color="white", corner_radius=8, height=36, font=("Segoe UI", 12, "bold"))
        self.btn_excluir.pack(side="left", padx=4)
        ctk.CTkButton(btns, text="Pesquisar tudo", command=self._pesquisar_tudo, fg_color="#64748B", hover_color="#475569", text_color="white", corner_radius=8, height=36, font=("Segoe UI", 12, "bold")).pack(side="left", padx=4)
        self.btn_finalizar_acao = ctk.CTkButton(btns, text="Finalizar Ação", command=self._finalizar_acao, fg_color="#D97706", hover_color="#B45309", text_color="white", corner_radius=8, height=36, font=("Segoe UI", 12, "bold"))
        self.btn_finalizar_acao.pack(side="left", padx=4)
        tb.Frame(btns).pack(side="left", fill="x", expand=True)
        self._atualizar_estado_botoes_cadastro()

        tb.Separator(frame, orient="horizontal").grid(row=3, column=0, columnspan=8, sticky="ew", padx=6, pady=(2, 2))

        tree_frame = tb.Frame(frame, height=300)
        tree_frame.grid(row=4, column=0, columnspan=8, sticky="nsew", padx=6, pady=(0, 4))
        frame.rowconfigure(4, weight=1)
        frame.columnconfigure(0, weight=1)

        cols = ("ordem", "protocolo", "nome", "cpf", "telefone", "data", "acao", "servico", "van")
        self.tree = tb.Treeview(tree_frame, columns=cols, show="headings")

        titulos_colunas = {
            "ordem": "Nº",
            "protocolo": "PEDIDO",
            "nome": "REQUERENTE",
            "cpf": "CPF",
            "telefone": "TELEFONE",
            "data": "DATA",
            "acao": "AÇÃO",
            "servico": "SERVIÇO",
            "van": "VAN",
        }

        larguras_colunas = {
            "ordem": 30,
            "protocolo": 115,
            "nome": 215,
            "cpf": 125,
            "telefone": 150,
            "data": 100,
            "acao": 190,
            "servico": 165,
            "van": 180,
        }
        minwidth_colunas = {
            "ordem": 30,
            "protocolo": 100,
            "nome": 150,
            "cpf": 100,
            "telefone": 120,
            "data": 90,
            "acao": 150,
            "servico": 120,
            "van": 180,
        }
        ancoras_colunas = {
            "ordem": "center",
            "protocolo": "center",
            "cpf": "center",
            "telefone": "center",
            "data": "center",
            "nome": "w",
            "acao": "w",
            "servico": "w",
            "van": "w",
        }

        for c in cols:
            self.tree.heading(
                c,
                text=titulos_colunas[c],
                command=lambda coluna=c: self._ordenar_tabela_cadastro(coluna)
            )
            self.tree.column(
                c,
                width=larguras_colunas[c],
                minwidth=minwidth_colunas[c],
                stretch=False,
                anchor=ancoras_colunas[c],
            )

        self.tree.pack(side="top", fill="both", expand=True)
        self._preparar_tabela(self.tree, minimo=70, maximo=1000, auto_ajuste=True)
        self.tree.bind("<<TreeviewSelect>>", lambda e: self._carregar_selecionado())
        self.tree.bind("<Double-1>", lambda e: self._ao_duplo_clique_tabela())

        export = ctk.CTkFrame(frame, fg_color="transparent")
        export.grid(row=5, column=0, columnspan=7, sticky="w", padx=6, pady=(0, 4))
        ctk.CTkButton(export, text="Exportar Excel", command=self._exportar_excel, fg_color="#2563EB", hover_color="#1D4ED8", text_color="white", corner_radius=8, height=34, font=("Segoe UI", 12, "bold")).pack(side="left", padx=6)
        ctk.CTkButton(export, text="Exportar PDF", command=self._abrir_tela_ordenacao_pdf, fg_color="#2563EB", hover_color="#1D4ED8", text_color="white", corner_radius=8, height=34, font=("Segoe UI", 12, "bold")).pack(side="left", padx=6)

        if getattr(self, "_carregar_todos_job", None):
            self.after_cancel(self._carregar_todos_job)
        acao_ativa = self.db.get("acao_ativa") or {}
        if acao_ativa.get("id"):
            self._carregar_todos_job = self.after_idle(self._carregar_todos)
        else:
            self._carregar_todos_job = None


    # Salva ou atualiza
    def _salvar_ou_atualizar(self):
        if not self._selected_acao or not self._selected_servico:
            messagebox.showwarning("Aviso", "Selecione uma ação e um serviço antes de cadastrar.")
            return

        acao_ativa = self.db.get("acao_ativa")

        if not acao_ativa:
            messagebox.showwarning("Aviso", "Nenhuma ação ativa. Selecione uma ação antes de cadastrar.")
            return

        if acao_ativa.get("data") != hoje_str():

            resposta = messagebox.askyesno(
                "Ação anterior",
                "Esta ação não é da data de hoje.\n\nDeseja continuar mesmo assim?"
            )

            if not resposta:
                return

        proto = self.ent_protocolo.get().strip()
        nome = self.ent_nome.get().strip()
        cpf = self.ent_cpf.get().strip()
        if hasattr(self, "var_sem_telefone") and self.var_sem_telefone.get():
            tel = "Não informado"
        else:
            telefones = [self.ent_tel.get().strip()]
            if getattr(self, "ent_tel_extra", None) is not None:
                telefones.append(self.ent_tel_extra.get().strip())
            tel = " / ".join(t for t in telefones if t)

        data = self.ent_data.get().strip()

        if not proto:
            messagebox.showwarning("Aviso", "Informe o protocolo.")
            return

        proto_numeros = somente_digs(proto)
        cpf_numeros = somente_digs(cpf)

        if proto_numeros and proto_numeros == cpf_numeros:
            messagebox.showwarning(
                "Aviso",
                "O protocolo e o CPF não podem conter a mesma numeração."
            )
            return

        if validar_cpf(proto):
            messagebox.showwarning(
                "Aviso",
                "O valor digitado no campo protocolo é um CPF válido. Não insira CPF no campo protocolo."
            )
            return

        if not nome:
            messagebox.showwarning("Aviso", "Informe o nome.")
            return
        if not validar_cpf(cpf):
            messagebox.showwarning("Aviso", "CPF inválido.")
            return
        try:
            datetime.strptime(data, "%d/%m/%Y")
        except:
            messagebox.showwarning("Aviso", "Data inválida.")
            return

        acao_info = self._selected_acao["id"] if self._selected_acao else ""
        servico_info = self._selected_servico["id"] if self._selected_servico else ""

        acao_ativa = self.db.get("acao_ativa") or {}
        vans_ids = acao_ativa.get("vans", []) or []
        if vans_ids:
            vans = load_vans()
            van = next((v for v in vans if v.get("id") == vans_ids[0]), None)
            van_id = str(van.get("id", "")) if van else ""
            van_nome_str = van.get("nome", "") if van else ""
        else:
            van_id = ""
            van_nome_str = ""

        registros = self.db["cadastros"]
        protocolo_normalizado = somente_digs(proto).lstrip("0") or "0"
        encontrado = next((r for r in registros if (somente_digs(str(r.get("protocolo", ""))).lstrip("0") or "0") == protocolo_normalizado), None)
        if encontrado:
            protocolo_origem_normalizado = somente_digs(self._protocolo_origem_selecao).lstrip("0") or "0"
            if protocolo_origem_normalizado != protocolo_normalizado:
                messagebox.showwarning("Aviso", "Protocolo já cadastrado.")
                return
            foto_cadastro = getattr(self, "_foto_cadastro_path", "") or encontrado.get("foto", "")
            exige_foto_edicao = bool(str(tel).strip())
            if exige_foto_edicao and (not foto_cadastro or not os.path.exists(foto_cadastro)):
                messagebox.showwarning(
                    "Foto obrigatória",
                    "Para editar este cadastro com telefone ou marcar 'Não informado', capture a foto antes de salvar."
                )
                return

            encontrado.update({
            "nome": nome,
            "cpf": cpf,
            "telefone": tel,
            "data": data,
            "acao_id": acao_info,
            "servico_id": servico_info,
            "van_id": van_id,
            "van_nome": van_nome_str,
            "foto": foto_cadastro
        })

            self._registrar_log("Atualizou cadastro", proto)
            msg = "Cadastro atualizado."
        else:
            registros.append({"protocolo": proto, "nome": nome, "cpf": cpf, "telefone": tel, "data": data, "entregue": False,
                              "acao_id": acao_info, "servico_id": servico_info, "van_id": van_id, "van_nome": van_nome_str,
                              "foto": getattr(self, "_foto_cadastro_path", ""), "hora": datetime.now().strftime("%H:%M")})
            self._registrar_log("Criou cadastro", {
                "protocolo": proto,
                "nome": nome,
                "cpf": cpf,
                "telefone": tel
        })
            msg = "Cadastro criado."
        save_db(self.db)
        messagebox.showinfo("OK", msg)
        self._carregar_todos()
        self._preparar_novo_cadastro_apos_salvar(data)

    # Exclui cadastro por protocolo
    def _excluir_cadastro_por_protocolo(self, protocolo):
        protocolo_normalizado = somente_digs(str(protocolo)).lstrip("0") or "0"

        cadastro_excluido = next(
            (
                r for r in self.db.get("cadastros", [])
                if (somente_digs(str(r.get("protocolo", ""))).lstrip("0") or "0") == protocolo_normalizado
            ),
            None
        )

        if not cadastro_excluido:
            return False

        salvar_excluido(
            DB_EXCLUIDOS_CADASTROS,
            "cadastro",
            cadastro_excluido.copy(),
            self.usuario
        )

        self.db["cadastros"] = [
            r for r in self.db.get("cadastros", [])
            if (somente_digs(str(r.get("protocolo", ""))).lstrip("0") or "0") != protocolo_normalizado
        ]

        save_db(self.db)
        return True

    # Exclui
    def _excluir(self):

        if not self._autorizar_admin():
            return

        selecionado = self.tree.selection()

        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione um registro na tabela para excluir.")
            return

        valores = self.tree.item(selecionado[0])["values"]

        protocolo = valores[1]
        nome = valores[2]
        cpf = valores[3]

        confirmar = messagebox.askyesno(
            "Confirmar exclusão",
            f"Deseja realmente excluir este cadastro?\n\n"
            f"Protocolo: {protocolo}\n"
            f"Nome: {nome}\n"
            f"CPF: {cpf}"
        )

        if not confirmar:
            return

        excluiu = self._excluir_cadastro_por_protocolo(protocolo)

        if not excluiu:
            messagebox.showwarning("Aviso", "Cadastro não encontrado no banco.")
            return

        messagebox.showinfo("Sucesso", "Cadastro excluído com sucesso.")

        self._carregar_todos()
        self._limpar_campos()


    # Finalizar ação
    def _finalizar_acao(self):
        if not self.db.get("acao_ativa"):
            messagebox.showinfo("Aviso", "Nenhuma ação ativa.")
            return

        confirmar = messagebox.askyesno(
            "Finalizar Ação",
            "Deseja realmente finalizar a ação atual?"
        )

        if not confirmar:
            return

        self.db["acao_ativa"] = None
        save_db(self.db)
        try:
            if os.path.exists(ARQUIVO_ACAO_ATIVA):
                os.remove(ARQUIVO_ACAO_ATIVA)
        except Exception:
            pass

        self._selected_acao = None
        self._selected_servico = None

        messagebox.showinfo("OK", "Ação finalizada com sucesso.")

        self._build_cadastro_screen()
        self._atualizar_estado_botoes_cadastro()

    # Ao digitar CPF
    def _ao_digitar_cpf(self, event):
        mascara_cpf_entrada(self.ent_cpf)
        if getattr(self, "_pesquisa_cadastro_job", None):
            self.after_cancel(self._pesquisa_cadastro_job)
        if getattr(self, "_carregar_todos_job", None):
            self.after_cancel(self._carregar_todos_job)
            self._carregar_todos_job = None
        self._pesquisa_cadastro_job = self.after_idle(self._pesquisar_cadastro_automatico)

    # Agenda pesquisa cadastro
    def _agendar_pesquisa_cadastro(self, event=None):
        if getattr(self, "_pesquisa_cadastro_job", None):
            self.after_cancel(self._pesquisa_cadastro_job)
        if getattr(self, "_carregar_todos_job", None):
            self.after_cancel(self._carregar_todos_job)
            self._carregar_todos_job = None
        self._pesquisa_cadastro_job = self.after_idle(self._pesquisar_cadastro_automatico)

    # Pesquisa cadastro automatico
    def _pesquisar_cadastro_automatico(self):
        if not hasattr(self, "tree") or not self.tree.winfo_exists():
            return

        self._pausar_auto_ajuste_tabela(self.tree)
        try:
            protocolo = self.ent_protocolo.get().strip().lower()
            cpf = self.ent_cpf.get().strip().lower()
            nome = normalizar_texto_ordenacao(self.ent_nome.get())

            acao_ativa = self.db.get("acao_ativa") or {}
            acao_id = str(acao_ativa.get("id", "")).strip()
            data_ativa = str(acao_ativa.get("data", "")).strip()
            com_acao_ativa = bool(acao_id)

            if not any([not com_acao_ativa and protocolo, cpf, nome]):
                self._carregar_todos()
                return

            cache_acao = {}
            cache_servico = {}

            registros = []
            for r in self.db.get("cadastros", []):
                if not com_acao_ativa and protocolo and protocolo not in str(r.get("protocolo", "")).lower():
                    continue

                if cpf and somente_digs(cpf) and somente_digs(cpf) not in somente_digs(str(r.get("cpf", ""))):
                    continue

                if nome and nome not in normalizar_texto_ordenacao(r.get("nome", "")):
                    continue

                registros.append(r)

            for i in self.tree.get_children():
                self.tree.delete(i)

            for idx, r in enumerate(registros, start=1):
                acao_id_r = str(r.get("acao_id", "")).strip()
                servico_id_r = str(r.get("servico_id", "")).strip()

                if acao_id_r not in cache_acao:
                    cache_acao[acao_id_r] = get_nome_acao_por_id(self.db, acao_id_r)
                if servico_id_r not in cache_servico:
                    cache_servico[servico_id_r] = get_nome_servico_por_id(self.db, servico_id_r)

                self.tree.insert("", "end", values=(
                    idx,
                    r.get("protocolo", ""),
                    r.get("nome", ""),
                    r.get("cpf", ""),
                    r.get("telefone", ""),
                    r.get("data", ""),
                    cache_acao.get(acao_id_r, ""),
                    cache_servico.get(servico_id_r, ""),
                    r.get("van_nome", "")
                ))
        finally:
            self._retomar_auto_ajuste_tabela(self.tree)

    # Carrega todos
    def _carregar_todos(self):
        if not hasattr(self, "tree") or not self.tree.winfo_exists():
            return

        for i in self.tree.get_children():
            self.tree.delete(i)

        hoje = hoje_str()

        acao_ativa = self.db.get("acao_ativa") or {}
        acao_id = str(acao_ativa.get("id", "")).strip()
        data_ativa = str(acao_ativa.get("data", "")).strip()

        acao_filtro = acao_id
        data_filtro = data_ativa
        if getattr(self, "_modo_leitura", False) and self._selected_acao:
            acao_filtro = str(self._selected_acao.get("id", "")).strip()
            data_filtro = str(self._selected_acao.get("data", "")).strip()

        registros = []
        for r in self.db.get("cadastros", []):
            if str(r.get("data", "")).strip() != hoje:
                continue

            if acao_filtro:
                if str(r.get("acao_id", "")).strip() != acao_filtro:
                    continue
                if data_filtro and str(r.get("data", "")).strip() != data_filtro:
                    continue

            registros.append(r)

        ultimo_item = None
        for idx, r in enumerate(registros, start=1):
            item_id = self.tree.insert("", "end", values=(
                idx,
                r.get("protocolo", ""),
                r.get("nome", ""),
                r.get("cpf", ""),
                r.get("telefone", ""),
                r.get("data", ""),
                get_nome_acao_por_id(self.db, r.get("acao_id")),
                get_nome_servico_por_id(self.db, r.get("servico_id")),
                r.get("van_nome", "")
            ))
            ultimo_item = item_id

        if ultimo_item:
            self.tree.see(ultimo_item)

        if not registros:
            self.tree.column("van", width=180, minwidth=180)
        else:
            self.tree.column("van", width=180, minwidth=180)


    # Carrega selecionado
    def _carregar_selecionado(self):
        sel = self.tree.selection()
        if not sel:
            return
        vals = self.tree.item(sel[0], "values")
        _, proto, nome, cpf, tel, data, acao, servico, van = vals
        self._protocolo_origem_selecao = proto.strip()
        protocolo_normalizado = somente_digs(str(proto)).lstrip("0") or "0"
        registro_selecionado = next(
            (
                r for r in self.db.get("cadastros", [])
                if (somente_digs(str(r.get("protocolo", ""))).lstrip("0") or "0") == protocolo_normalizado
            ),
            {}
        )
        self._exibir_foto_cadastro(registro_selecionado.get("foto", ""))

        acao_ativa = self.db.get("acao_ativa") or {}
        acao_id = str(acao_ativa.get("id", "")).strip()
        data_ativa = str(acao_ativa.get("data", "")).strip()
        com_acao_ativa = bool(acao_id)

        if not com_acao_ativa:
            self.ent_protocolo.configure(state="normal")
            self.ent_protocolo.delete(0, tk.END)
            self.ent_protocolo.insert(0, proto)
            self.ent_nome.configure(state="normal")
            self.ent_nome.delete(0, tk.END)
            self.ent_nome.insert(0, nome)
            self.ent_cpf.configure(state="normal")
            self.ent_cpf.delete(0, tk.END)
            self.ent_cpf.insert(0, cpf)
            self.ent_tel.configure(state="normal")
            self.ent_tel.delete(0, tk.END)
            if getattr(self, "ent_tel_extra", None) is not None:
                self.ent_tel_extra.destroy()
                self.ent_tel_extra = None
            if getattr(self, "btn_add_tel", None) is not None:
                self.btn_add_tel.configure(text="+", bootstyle="success-outline", state="normal")
            if hasattr(self, "var_sem_telefone"):
                self.var_sem_telefone.set(False)
            if "informado" in str(tel).lower():
                if hasattr(self, "var_sem_telefone"):
                    self.var_sem_telefone.set(True)
                self.ent_tel.configure(state="disabled")
                if getattr(self, "btn_add_tel", None) is not None:
                    self.btn_add_tel.configure(text="+", bootstyle="success-outline", state="disabled")
            else:
                telefones = [t.strip() for t in str(tel).split(" / ") if t.strip()]
                if telefones:
                    self.ent_tel.insert(0, telefones[0])
                if len(telefones) > 1 and hasattr(self, "_adicionar_telefone_extra"):
                    self._adicionar_telefone_extra()
                    self.ent_tel_extra.insert(0, telefones[1])

            self.ent_data.configure(state="normal")
            self.ent_data.delete(0, tk.END)
            self.ent_data.insert(0, data)
            self.ent_acao_info.configure(state="normal")
            self.ent_acao_info.delete(0, tk.END)
            self.ent_acao_info.insert(0, acao)
            self.ent_acao_info.configure(state="disabled")
            self.ent_servico_info.configure(state="normal")
            self.ent_servico_info.delete(0, tk.END)
            self.ent_servico_info.insert(0, servico)
            self.ent_servico_info.configure(state="disabled")

            self.ent_protocolo.configure(state="readonly")
            self.ent_data.configure(state="disabled")
            self.ent_acao_info.configure(state="disabled")
            self.ent_servico_info.configure(state="disabled")
            self.ent_cpf.configure(state="readonly")
            self.ent_nome.configure(state="readonly")
            self.ent_tel.configure(state="readonly")
            if getattr(self, "btn_add_tel", None) is not None:
                self.btn_add_tel.configure(state="disabled")
            if getattr(self, "ent_tel_extra", None) is not None:
                self.ent_tel_extra.configure(state="readonly")
            if hasattr(self, "chk_sem_tel"):
                self.chk_sem_tel.configure(state="disabled")
        else:
            self.ent_protocolo.configure(state="normal")
            self.ent_nome.configure(state="normal")
            self.ent_cpf.configure(state="normal")
            self.ent_protocolo.delete(0, tk.END)
            self.ent_protocolo.insert(0, proto)
            self.ent_nome.delete(0, tk.END)
            self.ent_nome.insert(0, nome)
            self.ent_cpf.delete(0, tk.END)
            self.ent_cpf.insert(0, cpf)
            self.ent_tel.configure(state="normal")
            self.ent_tel.delete(0, tk.END)
            if getattr(self, "ent_tel_extra", None) is not None:
                self.ent_tel_extra.destroy()
                self.ent_tel_extra = None
            if getattr(self, "btn_add_tel", None) is not None:
                self.btn_add_tel.configure(text="+", bootstyle="success-outline", state="normal")
            if hasattr(self, "chk_sem_tel"):
                self.chk_sem_tel.configure(state="normal")
            if hasattr(self, "var_sem_telefone"):
                self.var_sem_telefone.set(False)
            if getattr(self, "ent_data", None) is not None:
                self.ent_data.delete(0, tk.END)
                self.ent_data.insert(0, data_ativa or hoje_str())

            if "informado" in str(tel).lower():
                if hasattr(self, "var_sem_telefone"):
                    self.var_sem_telefone.set(True)
                self.ent_tel.configure(state="disabled")
                if getattr(self, "btn_add_tel", None) is not None:
                    self.btn_add_tel.configure(text="+", bootstyle="success-outline", state="disabled")
            else:
                telefones = [t.strip() for t in str(tel).split(" / ") if t.strip()]
                if telefones:
                    self.ent_tel.insert(0, telefones[0])
            if len(telefones) > 1 and hasattr(self, "_adicionar_telefone_extra"):
                self._adicionar_telefone_extra()
                self.ent_tel_extra.insert(0, telefones[1])

    # Ao duplo clique tabela
    def _ao_duplo_clique_tabela(self):
        sel = self.tree.selection()
        if not sel:
            return
        self._carregar_selecionado()
        self._carregar_todos()

    # Ação ativa do dia
    def _acao_ativa_do_dia(self):
        acao_ativa = self.db.get("acao_ativa") or {}
        acao_id = str(acao_ativa.get("id", "")).strip()
        data_acao = str(acao_ativa.get("data", "")).strip()
        if not acao_id or data_acao != hoje_str():
            return None
        return acao_ativa

    # Id ação ativa atual
    def _id_acao_ativa_atual(self):
        acao_ativa = self._acao_ativa_do_dia()
        if not acao_ativa:
            return ""
        return str(acao_ativa.get("id", "")).strip()

    # Atualiza estado botões cadastro
    def _atualizar_estado_botoes_cadastro(self):
        acao_ativa = self.db.get("acao_ativa") or {}
        tem_acao_ativa = bool(acao_ativa.get("id"))

        if getattr(self, "btn_abrir_acao", None) and hasattr(self.btn_abrir_acao, "winfo_exists"):
            if self.btn_abrir_acao.winfo_exists():
                if tem_acao_ativa:
                    self.btn_abrir_acao.pack_forget()
                else:
                    self.btn_abrir_acao.pack(side="left", padx=4)

        botoes_acao = [self.btn_salvar, self.btn_capturar_idnet, self.btn_excluir, self.btn_finalizar_acao]
        for btn in botoes_acao:
            if getattr(btn, "winfo_exists", lambda: False)():
                if tem_acao_ativa:
                    btn.pack(side="left", padx=4)
                else:
                    btn.pack_forget()

        if getattr(self, "btn_selecionar_acao_servico", None) and hasattr(self.btn_selecionar_acao_servico, "winfo_exists"):
            if self.btn_selecionar_acao_servico.winfo_exists():
                self.btn_selecionar_acao_servico.configure(state="disabled" if tem_acao_ativa else "normal")

        if getattr(self, "btn_capturar_foto_idnet", None) and hasattr(self.btn_capturar_foto_idnet, "winfo_exists"):
            if self.btn_capturar_foto_idnet.winfo_exists():
                if tem_acao_ativa:
                    self.btn_capturar_foto_idnet.pack(fill="x", padx=6, pady=(0, 6))
                else:
                    self.btn_capturar_foto_idnet.pack_forget()

    # Prepara novo cadastro após salvar
    def _preparar_novo_cadastro_apos_salvar(self, data_padrao=None):
        if getattr(self, "tree", None):
            try:
                self.tree.selection_remove(self.tree.selection())
            except Exception:
                pass

        self._limpar_campos(data_padrao)
        for campo in ("ent_protocolo", "ent_nome", "ent_cpf", "ent_tel"):
            widget = getattr(self, campo, None)
            if widget:
                try:
                    widget.configure(state="normal")
                except Exception:
                    pass

        if getattr(self, "ent_tel_extra", None) is not None:
            try:
                self.ent_tel_extra.configure(state="normal")
            except Exception:
                pass
        if getattr(self, "btn_add_tel", None) is not None:
            self.btn_add_tel.configure(state="normal")
        if hasattr(self, "chk_sem_tel"):
            self.chk_sem_tel.configure(state="normal")
        self._protocolo_origem_selecao = ""
        self.ent_protocolo.focus_set()

    # Pasta fotos idNet
    def _pasta_fotos_idnet(self):
        pasta = os.path.join(PASTA_IMG, "idnet")
        os.makedirs(pasta, exist_ok=True)
        return pasta

    # Limpa foto cadastro
    def _limpar_foto_cadastro(self):
        self._foto_cadastro_path = ""
        self._foto_cadastro_img = None
        if getattr(self, "lbl_foto_cadastro", None) and self.lbl_foto_cadastro.winfo_exists():
            try:
                fundo = Image.new("RGB", (116, 126), (255, 255, 255))
                self._foto_cadastro_img = ctk.CTkImage(
                    light_image=fundo,
                    dark_image=fundo,
                    size=(116, 126)
                )
                self.lbl_foto_cadastro.configure(image=self._foto_cadastro_img, text="Sem\nfoto")
            except Exception:
                pass

    # Exibe foto cadastro
    def _exibir_foto_cadastro(self, caminho):
        if not caminho or not os.path.exists(caminho):
            self._limpar_foto_cadastro()
            return

        try:
            imagem = Image.open(caminho).convert("RGB")
            imagem.thumbnail((116, 126), Image.LANCZOS)
            fundo = Image.new("RGB", (116, 126), (255, 255, 255))
            x = (116 - imagem.width) // 2
            y = (126 - imagem.height) // 2
            fundo.paste(imagem, (x, y))
            if self._foto_cadastro_img is not None:
                try:
                    self._foto_cadastro_img.destroy()
                except Exception:
                    pass
            self._foto_cadastro_img = ctk.CTkImage(
                light_image=fundo,
                dark_image=fundo,
                size=(116, 126)
            )
            self._foto_cadastro_path = caminho
            if getattr(self, "lbl_foto_cadastro", None) and self.lbl_foto_cadastro.winfo_exists():
                self.lbl_foto_cadastro.configure(image=self._foto_cadastro_img, text="")
        except Exception:
            self._limpar_foto_cadastro()

    # Captura recorte foto idNet
    def _capturar_recorte_foto_idnet(self, hwnd):
        if not hwnd:
            return None

        try:
            rect = wintypes.RECT()
            ctypes.windll.user32.GetWindowRect(hwnd, ctypes.byref(rect))

            # Recorte calibrado para a foto do idNet antes do Ctrl+A destacar a pagina.
            esquerda = rect.left + 220
            topo = rect.top + 180
            direita = esquerda + 120
            baixo = topo + 150

            imagem = ImageGrab.grab(bbox=(esquerda, topo, direita, baixo))
            if imagem.mode != "RGBA":
                imagem = imagem.convert("RGBA")
            fundo = Image.new("RGB", imagem.size, (255, 255, 255))
            fundo.paste(imagem, mask=imagem.split()[3])
            return fundo
        except Exception:
            return None

    # Salva foto idNet capturada
    def _salvar_foto_idnet_capturada(self, imagem, protocolo):
        if imagem is None or not protocolo:
            return ""

        try:
            nome = nome_arquivo_seguro(somente_digs(protocolo) or protocolo)
            caminho = os.path.join(self._pasta_fotos_idnet(), f"{nome}.jpg")
            imagem.save(caminho, "JPEG", quality=92)
            return caminho
        except Exception:
            return ""

    # Captura foto idNet manualmente
    def _capturar_foto_idnet_manual(self):
        protocolo = self.ent_protocolo.get().strip() if getattr(self, "ent_protocolo", None) else ""
        if not protocolo:
            messagebox.showwarning("Capturar foto", "Preencha ou capture o protocolo antes de capturar a foto.")
            return

        hwnd = self._localizar_janela_idnet()
        if not hwnd:
            messagebox.showwarning("Capturar foto", "Não encontrei a janela do idNet aberta.")
            return

        user32 = ctypes.windll.user32
        try:
            user32.ShowWindow(hwnd, 9)
            user32.SetForegroundWindow(hwnd)
            time.sleep(0.45)
            imagem = self._capturar_recorte_foto_idnet(hwnd)
        finally:
            try:
                user32.SetForegroundWindow(self.winfo_id())
            except Exception:
                pass

        caminho = self._salvar_foto_idnet_capturada(imagem, protocolo)
        if not caminho:
            messagebox.showwarning("Capturar foto", "Não consegui capturar a foto do idNet.")
            return

        self._exibir_foto_cadastro(caminho)
        messagebox.showinfo("Capturar foto", "Foto capturada com sucesso.")

    # Localiza janela do idNet no navegador
    def _localizar_janela_idnet(self):
        user32 = ctypes.windll.user32
        encontrados = []

        @ctypes.WINFUNCTYPE(ctypes.c_bool, wintypes.HWND, wintypes.LPARAM)
        def enum_proc(hwnd, lparam):
            if not user32.IsWindowVisible(hwnd):
                return True

            tamanho = user32.GetWindowTextLengthW(hwnd)
            if tamanho <= 0:
                return True

            buffer = ctypes.create_unicode_buffer(tamanho + 1)
            user32.GetWindowTextW(hwnd, buffer, tamanho + 1)
            titulo = buffer.value.strip()
            titulo_lower = titulo.lower()

            termos = ("idnet", "montreal", "edge", "chrome")
            if any(termo in titulo_lower for termo in termos):
                encontrados.append((hwnd, titulo))

            return True

        user32.EnumWindows(enum_proc, 0)

        for hwnd, titulo in encontrados:
            titulo_lower = titulo.lower()
            if "idnet" in titulo_lower or "montreal" in titulo_lower:
                return hwnd

        return encontrados[0][0] if encontrados else None

    # Pressiona tecla no Windows
    def _pressionar_tecla_windows(self, vk):
        user32 = ctypes.windll.user32
        user32.keybd_event(vk, 0, 0, 0)
        time.sleep(0.04)
        user32.keybd_event(vk, 0, 2, 0)
        time.sleep(0.04)

    # Le texto atual da area de transferencia
    def _ler_texto_clipboard(self):
        try:
            self.update()
            return self.clipboard_get()
        except Exception:
            return ""

    # Copia texto da pagina idNet aberta
    def _copiar_texto_janela_idnet(self):
        hwnd = self._localizar_janela_idnet()
        if not hwnd:
            return ""

        user32 = ctypes.windll.user32
        antigo_clipboard = ""
        try:
            antigo_clipboard = self.clipboard_get()
        except Exception:
            antigo_clipboard = ""

        try:
            self.clipboard_clear()
            self.update()
            user32.ShowWindow(hwnd, 9)
            user32.SetForegroundWindow(hwnd)
            time.sleep(0.55)

            vk_control = 0x11
            vk_a = 0x41
            vk_c = 0x43

            user32.keybd_event(vk_control, 0, 0, 0)
            self._pressionar_tecla_windows(vk_a)
            user32.keybd_event(vk_control, 0, 2, 0)
            time.sleep(0.35)

            texto_copiado = ""
            for _ in range(2):
                user32.keybd_event(vk_control, 0, 0, 0)
                self._pressionar_tecla_windows(vk_c)
                user32.keybd_event(vk_control, 0, 2, 0)

                limite = time.time() + 2.0
                while time.time() < limite:
                    texto_copiado = self._ler_texto_clipboard()
                    if texto_copiado.strip():
                        return texto_copiado
                    time.sleep(0.15)

            return texto_copiado
        finally:
            try:
                user32.SetForegroundWindow(self.winfo_id())
            except Exception:
                pass
            try:
                self.clipboard_clear()
                if antigo_clipboard:
                    self.clipboard_append(antigo_clipboard)
            except Exception:
                pass

    # Extrai pedido cpf requerente texto idNet
    def _extrair_dados_idnet(self, texto):
        texto = texto or ""
        linhas = [l.strip() for l in re.split(r"[\r\n]+", texto) if l.strip()]
        dados = {"pedido": "", "cpf": "", "nome": ""}

        pedido = re.search(r"\bPedido\s*:?\s*([0-9]{5,})", texto, re.IGNORECASE)
        if pedido:
            dados["pedido"] = pedido.group(1).strip()

        cpf = re.search(r"\b(\d{3}\.?\d{3}\.?\d{3}-?\d{2})\b", texto)
        if cpf:
            dados["cpf"] = mask_cpf_from_clean(somente_digs(cpf.group(1)))

        def nome_valido_idnet(valor):
            valor = str(valor or "").strip()
            if not valor:
                return False
            if re.search(r"\d|[/\\]|:", valor):
                return False
            termos_invalidos = (
                r"^antropometria$",
                r"^endere[cç]o\s*/\s*contatos$",
                r"^documenta[cç][aã]o$",
                r"^biometria$",
                r"^observa[cç][oõ]es\s*/\s*resumo$",
                r"^ajustar nomes manualmente$",
                r"^nome social$",
                r"^filia[cç][aã]o\s+\d",
                r"documento",
                r"apresentad",
                r"certid[aã]o",
                r"certificado",
                r"expedidor",
                r"data de",
                r"nascimento",
                r"sexo",
                r"estado civil",
                r"profiss[aã]o",
                r"nacionalidade",
                r"pa[ií]s",
                r"pedido",
                r"modalidade",
                r"situa[cç][aã]o",
                r"status",
                r"listar atividades",
            )
            if any(re.search(padrao, valor, re.IGNORECASE) for padrao in termos_invalidos):
                return False
            if not re.fullmatch(r"[A-Za-zÀ-ÖØ-öø-ÿ' -]+", valor):
                return False

            partes = [p for p in valor.split() if p]
            partes_fortes = [p for p in partes if len(p) > 2]
            return 2 <= len(partes) <= 8 and len(partes_fortes) >= 2

        def limpar_linha_nome_idnet(valor):
            valor = str(valor or "").strip()
            valor = re.sub(r"Requerente\s*\*", "", valor, flags=re.IGNORECASE).strip()
            valor = re.sub(r"Ajustar Nomes Manualmente", "", valor, flags=re.IGNORECASE).strip()
            valor = re.sub(r"^\s*[:\-]\s*", "", valor).strip()
            return valor

        for i, linha in enumerate(linhas):
            if re.search(r"\bRequerente\s*\*", linha, re.IGNORECASE):
                candidatos = [limpar_linha_nome_idnet(linha)]

                for proxima in linhas[i + 1:i + 12]:
                    if re.fullmatch(r"Nome Social", proxima, re.IGNORECASE):
                        break
                    candidatos.append(limpar_linha_nome_idnet(proxima))

                for candidato in candidatos:
                    if nome_valido_idnet(candidato):
                        dados["nome"] = candidato
                        break
                break

            req = re.search(r"Requerente\s+\*\s*[:\-]?\s+(.+)$", linha, re.IGNORECASE)
            if req:
                candidato = limpar_linha_nome_idnet(req.group(1))
                if nome_valido_idnet(candidato):
                    dados["nome"] = candidato
                    break

        if not dados["nome"]:
            for i, linha in enumerate(linhas):
                if re.fullmatch(r"Requerente", linha, re.IGNORECASE):
                    for proxima in linhas[i + 1:i + 8]:
                        candidato = limpar_linha_nome_idnet(proxima)
                        if nome_valido_idnet(candidato):
                            dados["nome"] = candidato
                            break
                    break

        return dados

    # Captura dados do idNet
    def _capturar_dados_idnet(self):
        texto = self._copiar_texto_janela_idnet()
        if not texto.strip():
            messagebox.showwarning(
                "Capturar idNet",
                "Consegui acionar a página do idNet, mas o texto não chegou à área de transferência. Clique uma vez dentro da página do idNet e tente novamente."
            )
            return

        dados = self._extrair_dados_idnet(texto)
        faltando = [nome for nome, valor in (("Pedido", dados["pedido"]), ("CPF", dados["cpf"]), ("Requerente", dados["nome"])) if not valor]
        if faltando:
            messagebox.showwarning(
                "Capturar idNet",
                "Não consegui localizar: " + ", ".join(faltando) + ". Confira se a aba Requerente do idNet está aberta."
            )
            return

        self.ent_protocolo.configure(state="normal")
        self.ent_protocolo.delete(0, tk.END)
        self.ent_protocolo.insert(0, dados["pedido"])

        self.ent_cpf.configure(state="normal")
        self.ent_cpf.delete(0, tk.END)
        self.ent_cpf.insert(0, dados["cpf"])

        self.ent_nome.configure(state="normal")
        self.ent_nome.delete(0, tk.END)
        self.ent_nome.insert(0, dados["nome"])
        self.ent_tel.focus_set()

        messagebox.showinfo("Capturar idNet", "Pedido, CPF e Requerente preenchidos com sucesso.")

    # Limpa campos
    def _limpar_campos(self, data_padrao=None):
      self._protocolo_origem_selecao = ""
      self._limpar_foto_cadastro()
      self.ent_protocolo.configure(state="normal")
      self.ent_protocolo.delete(0, tk.END)
      self.ent_nome.configure(state="normal")
      self.ent_nome.delete(0, tk.END)
      self.ent_cpf.configure(state="normal")
      self.ent_cpf.delete(0, tk.END)
      self.ent_tel.configure(state="normal")
      self.ent_tel.delete(0, tk.END)
      if getattr(self, "ent_tel_extra", None) is not None:
          self.ent_tel_extra.destroy()
          self.ent_tel_extra = None
      if getattr(self, "btn_add_tel", None) is not None:
          self.btn_add_tel.configure(text="+", bootstyle="success-outline", state="normal")
      if hasattr(self, "chk_sem_tel"):
          self.chk_sem_tel.configure(state="normal")
      if hasattr(self, "var_sem_telefone"):
          self.var_sem_telefone.set(False)

      self.ent_data.configure(state="normal")
      self.ent_data.delete(0, tk.END)
      if self._selected_acao:
          self.ent_data.insert(0, self._selected_acao.get("data") or hoje_str())
      else:
          self.ent_data.insert(0, data_padrao or "")

      acao_ativa = self.db.get("acao_ativa") or {}
      com_acao_ativa = bool(str(acao_ativa.get("id", "")).strip())

      self.ent_acao_info.configure(state="normal")
      self.ent_acao_info.delete(0, tk.END)
      self.ent_acao_info.insert(0, "")
      self.ent_acao_info.configure(state="disabled")
      self.ent_servico_info.configure(state="normal")
      self.ent_servico_info.delete(0, tk.END)
      self.ent_servico_info.insert(0, "")
      self.ent_servico_info.configure(state="disabled")

      self._preencher_campos_acao_ativa()

      self._atualizar_estado_botoes_cadastro()

    # Pesquisa
    def _pesquisar(self):
        self._preencher_campos_acao_ativa()
        protocolo = self.ent_protocolo.get().strip().lower()
        nome = normalizar_texto_ordenacao(self.ent_nome.get())
        cpf = self.ent_cpf.get().strip().lower()
        data = self.ent_data.get().strip().lower()

        acao = normalizar_texto_ordenacao(self.ent_acao_info.get())
        servico = normalizar_texto_ordenacao(self.ent_servico_info.get())

        for i in self.tree.get_children():
            self.tree.delete(i)

        encontrados = []

        campos_principais_vazios = not any([protocolo, nome, cpf, data])

        if campos_principais_vazios:
            if not acao or "nenhuma" in acao:
                messagebox.showwarning(
                    "Pesquisa",
                    "Informe Protocolo, Nome, CPF ou Data.\n\n"
                    "Ou selecione uma Ação e um Serviço."
                )
                return

            encontrados = [
                r for r in self.db["cadastros"]
                if acao in r.get("acao","").lower()
                and r.get("servico_id") == self._selected_servico["id"]
            ]
        else:
            for r in self.db["cadastros"]:
                if (
                    (protocolo and protocolo in r.get("protocolo","").lower()) or
                    (nome and nome in normalizar_texto_ordenacao(r.get("nome", ""))) or
                    (cpf and cpf in r.get("cpf","").lower()) or
                    (data and data in r.get("data","").lower())
                ):
                    encontrados.append(r)

        if not encontrados:
            messagebox.showinfo("Pesquisa", "Nenhum registro encontrado.")
            return
        primeiro = encontrados[0]

        acao_id = primeiro.get("acao_id")
        servico_id = primeiro.get("servico_id")

        acao_obj = next((a for a in self.db.get("acoes", []) if a["id"] == acao_id), None)
        servico_obj = next((s for s in self.db.get("servicos", []) if s["id"] == servico_id), None)

        if False and acao_obj:
            self._selected_acao = {
                "id": acao_obj["id"],
                "data": acao_obj["data"],
                "local": acao_obj["local"]
            }

            self.ent_acao_info.configure(state="normal")
            self.ent_acao_info.delete(0, tk.END)
            self.ent_acao_info.insert(0, f"{acao_obj['data']} - {acao_obj['local']}")
            self.ent_acao_info.configure(state="disabled")

        if False and servico_obj:
            self._selected_servico = {
                "id": servico_obj["id"],
                "nome": servico_obj["nome"]
            }

            self.ent_servico_info.configure(state="normal")
            self.ent_servico_info.delete(0, tk.END)
            self.ent_servico_info.insert(0, servico_obj["nome"])
            self.ent_servico_info.configure(state="disabled")

        for idx, r in enumerate(encontrados, start=1):
            numero = f"{idx:03d}"
            nome = f"{numero} - {r.get('nome','')}"

            item_id = self.tree.insert("", "end", values=(
                idx,
                r.get("protocolo", ""),
                r.get("nome", ""),
                r.get("cpf", ""),
                r.get("telefone", ""),
                r.get("data", ""),
                get_nome_acao_por_id(self.db, r.get("acao_id")),
                get_nome_servico_por_id(self.db, r.get("servico_id")),
                r.get("van_nome", "")
            ))


            self.tree.see(item_id)

    # Pesquisa tudo
    def _pesquisar_tudo(self):
        if getattr(self, "_modo_leitura", False):
            messagebox.showwarning("Aviso", "Não é permitido limpar a seleção no modo de leitura.")
            return

        for i in self.tree.get_children():
            self.tree.delete(i)

        for idx, r in enumerate(self.db["cadastros"], start=1):
            numero = f"{idx:03d}"
            nome = f"{numero} - {r.get('nome','')}"

            item_id = self.tree.insert("", "end", values=(
                idx,
                r.get("protocolo", ""),
                r.get("nome", ""),
                r.get("cpf", ""),
                r.get("telefone", ""),
                r.get("data", ""),
                get_nome_acao_por_id(self.db, r.get("acao_id")),
                get_nome_servico_por_id(self.db, r.get("servico_id")),
                r.get("van_nome", "")
            ))

            self.tree.selection_set(item_id)
            self.tree.focus(item_id)
            self.tree.see(item_id)
        self._selected_acao = None
        self._selected_servico = None

        self.ent_acao_info.configure(state="normal")
        self.ent_acao_info.delete(0, tk.END)
        self.ent_acao_info.insert(0, "Todas as ações")
        self.ent_acao_info.configure(state="disabled")

        self.ent_servico_info.configure(state="normal")
        self.ent_servico_info.delete(0, tk.END)
        self.ent_servico_info.insert(0, "Todos os serviços")
        self.ent_servico_info.configure(state="disabled")
    # Exporta excel
    def _exportar_excel(self):

        if not self.db["cadastros"]:
            messagebox.showwarning("Aviso", "Não há dados para exportar.")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório Geral"

        headers = ["Protocolo", "Nome", "CPF", "Telefone", "Data", "Ação", "Serviço"]
        ws.append(headers)

        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        bold_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = center

        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for r in self.db["cadastros"]:
            acao_nome = get_nome_acao_por_id(self.db, r.get("acao_id"))
            serv_nome = get_nome_servico_por_id(self.db, r.get("servico_id"))

            ws.append([
                r.get("protocolo"),
                r.get("nome"),
                r.get("cpf"),
                r.get("telefone"),
                r.get("data"),
                acao_nome,
                serv_nome
            ])

        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 3

        caminho = "relatorio_geral.xlsx"
        wb.save(caminho)

        os.startfile(caminho)

        messagebox.showinfo("OK", "Relatório Excel gerado com sucesso.")

    # Abre tela ordenação PDF
    def _abrir_tela_ordenacao_pdf(self):
        self._preencher_campos_acao_ativa()
        self._registros_pdf_ordenacao = []
        for item in self.tree.get_children():
            self._registros_pdf_ordenacao.append(self.tree.item(item)["values"])

        top = self._open_inline_container("Ordenar PDF", voltar=self._build_cadastro_screen)
        top.title("Ordenar PDF")
        top.geometry("300x260")
        top.grab_set()

        tb.Label(
            top,
            text="Ordenar PDF por:",
            font=("Segoe UI", 12, "bold")
        ).pack(pady=10)

        self._ordem_pdf = tk.StringVar(value="nome")

        opcoes = [
            ("Número", "ordem"),
            ("Nome", "nome"),
            ("CPF", "cpf"),
            ("Protocolo", "protocolo"),
            ("Telefone", "telefone"),
        ]

        for texto, valor in opcoes:
            tb.Radiobutton(
                top,
                text=texto,
                variable=self._ordem_pdf,
                value=valor
            ).pack(anchor="w", padx=20)

        tb.Button(
            top,
            text="Gerar PDF",
            bootstyle=SUCCESS,
            command=lambda: self._exportar_pdf_ordenado(top)
        ).pack(pady=15)

    # Exporta PDF ordenado
    def _exportar_pdf_ordenado(self, janela):
        registros = list(getattr(self, "_registros_pdf_ordenacao", []))
        registros_sem_numero = []
        for r in registros:
            registros_sem_numero.append(r[1:])

            coluna = self._ordem_pdf.get()

            if not coluna:
                coluna = "nome"

            mapa_colunas = {
                "protocolo": 0,
                "nome": 1,
                "cpf": 2,
                "telefone": 3,
            }

            indice = mapa_colunas[coluna]

            if coluna == "nome":
                registros_sem_numero.sort(
                    key=lambda x: normalizar_texto_ordenacao(x[indice])
                )
            else:
                registros_sem_numero.sort(
                    key=lambda x: str(x[indice]).upper()
                )

            registros = []
        for i, r in enumerate(registros_sem_numero, start=1):
            registros.append((i, *r))
        if self._selected_acao:
            self._pdf_acao = self._selected_acao.get("local", "")
            self._pdf_data = self._selected_acao.get("data", "")
            vans_ids = self._selected_acao.get("vans", []) or []
            if vans_ids:
                van = next((v for v in self.vans if v.get("id") == vans_ids[0]), None)
                self._pdf_van = van.get("nome", "") if van else ""
            else:
                self._pdf_van = ""
        else:
            self._pdf_acao = ""
            self._pdf_data = ""
            self._pdf_van = ""

        if self._selected_servico:
            if isinstance(self._selected_servico, dict):
                self._pdf_servico = self._selected_servico.get("nome", "")
            else:
                self._pdf_servico = get_nome_servico_por_id(self.db, self._selected_servico)
        else:
            self._pdf_servico = ""

        if not registros:
            messagebox.showwarning("Aviso", "Nenhum dado para exportar.")
            return

        coluna = self._ordem_pdf.get()

        mapa_colunas = {
            "ordem": 0,
            "protocolo": 1,
            "nome": 2,
            "cpf": 3,
            "telefone": 4,
        }

        indice = mapa_colunas[coluna]
        if coluna == "nome":
            registros.sort(key=lambda x: normalizar_texto_ordenacao(x[indice]))
        else:
            registros.sort(key=lambda x: str(x[indice]).upper())
        self._gerar_pdf_final(registros)
        self._fechar_tela_interna(janela)

    # Gera PDF final
    def _gerar_pdf_final(self, registros):
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from datetime import datetime
        import os

        filename = caminho_pdf("cadastro", "cadastro")
        c = canvas.Canvas(filename, pagesize=A4)
        largura, altura = A4

        margem_x = 40
        margem_y = 50

        y = self._desenhar_cabecalho_pdf(
        c,
        largura,
        altura,
        margem_x,
        self._pdf_data,
        registros
    )
        c.setFont("Helvetica-Bold", 10)
        colunas = [
            ("Nº", 40),
            ("Protocolo", 80),
            ("Nome", 200),
            ("CPF", 110),
            ("Telefone", 110),
        ]

        x = margem_x
        for titulo, largura_col in colunas:
            c.drawString(x, y, titulo)
            x += largura_col
        y -= 15

        c.setFont("Helvetica", 9)

        for r in registros:
            if y < margem_y:
                c.showPage()

                y = self._desenhar_cabecalho_pdf(
                    c,
                    largura,
                    altura,
                    margem_x,
                    self._pdf_data,
                    registros
                )
                c.setFont("Helvetica", 9)

            x = margem_x
            c.drawString(x, y, str(r[0]))
            x += 40

            c.drawString(x, y, str(r[1]))
            x += 80

            c.drawString(x, y, str(r[2]))
            x += 200

            c.drawString(x, y, mascarar_cpf_pdf(str(r[3])))
            x += 110

            c.drawString(x, y, str(r[4]))

            y -= 14

        from datetime import datetime
        data_emissao = datetime.now().strftime("%d/%m/%Y %H:%M")
        c.setFont("Helvetica", 9)
        c.drawString(margem_x, 30, f"Data de Emissão: {data_emissao}")

        c.save()

        messagebox.showinfo("OK", f"PDF gerado com sucesso: {filename}")
        try:
            os.startfile(filename)
        except:
            pass

    # Desenha cabeçalho PDF
    def _desenhar_cabecalho_pdf(self, c, largura, altura, margem_x, data_sel, registros):
        from datetime import datetime

        y = altura - 50

        try:
            brasao_w = 90
            brasao_h = 90
            c.drawImage(
                "img/Logo_Recife.png",
                (largura - brasao_w) / 2,
                y - brasao_h,
                width=brasao_w,
                height=brasao_h,
                preserveAspectRatio=True,
                mask="auto"
            )
        except:
            pass

        y -= 100

        c.setFont("Helvetica-Bold", 12)
        c.drawCentredString(largura / 2, y, "PREFEITURA DO RECIFE")
        y -= 16

        c.setFont("Helvetica-Bold", 11)
        c.drawCentredString(largura / 2, y, "EMPRESA MUNICIPAL DE INFORMATICA - EMPREL")
        y -= 14

        c.setFont("Helvetica", 10)
        c.drawCentredString(largura / 2, y, "Van Conecta Itinerante")
        y -= 18

        c.line(margem_x, y, largura - margem_x, y)
        y -= 20


        acao_nome = getattr(self, "_pdf_acao", "")
        servico_nome = getattr(self, "_pdf_servico", "")

        c.setFont("Helvetica-Bold", 14)
        c.drawCentredString(largura / 2, y, "RELATÓRIO DE IDENTIDADES CADASTRADAS")
        y -= 20

        c.setFont("Helvetica", 11)

        modo_geral = getattr(self, "_pdf_modo_geral", False)

        if not modo_geral:
            c.drawString(margem_x, y, f"Ação: {self._pdf_acao}")
            y -= 14

            c.drawString(margem_x, y, f"Serviço: {servico_nome}")
            y -= 15

            c.drawString(margem_x, y, f"Data do Atendimento: {self._pdf_data}")
            y -= 15

            if getattr(self, "_pdf_van", ""):
                c.drawString(margem_x, y, f"Van: {self._pdf_van}")
                y -= 20

            c.line(margem_x, y, largura - margem_x, y)
            y -= 25

            return y

    # Exporta PDF
    def exportar_pdf(self):
        data_sel = self.ent_data.get().strip()
        registros = [
            r for r in self.db["cadastros"]
            if r.get("data") == data_sel
        ]
        if not registros:
            messagebox.showwarning("Aviso", "Nada para exportar.")
            return

        from datetime import datetime
        import hashlib

        filename = caminho_pdf("cadastro", "cadastro")
        c = canvas.Canvas(filename, pagesize=A4)
        largura, altura = A4

        margem_x = 40
        margem_y = 50
        y = altura - margem_y

        c.setFont("Helvetica-Bold", 12)
        c.drawCentredString(largura / 2, y, "PREFEITURA DO RECIFE")
        y -= 16

        c.setFont("Helvetica-Bold", 11)
        c.drawCentredString(largura / 2, y, "EMPRESA MUNICIPAL DE INFORMATICA - EMPREL")
        y -= 14

        c.setFont("Helvetica", 10)
        c.drawCentredString(largura / 2, y, "Van Conecta Itinerante")

        y -= 18
        c.line(margem_x, y, largura - margem_x, y)
        y -= 20

        y = self._desenhar_cabecalho_pdf(
            c,
            largura,
            altura,
            margem_x,
            data_sel,
            registros
        )

        c.setFont("Helvetica-Bold", 10)
        modo_geral = getattr(self, "_pdf_modo_geral", False)

        if modo_geral:
            colunas = [
                ("Nº", 40),
                ("Protocolo", 80),
                ("Nome", 160),
                ("CPF", 100),
                ("Telefone", 100),
                ("Ação", 140),
                ("Data", 90),
            ]
        else:
            colunas = [
                ("Nº", 40),
                ("Protocolo", 80),
                ("Nome", 200),
                ("CPF", 110),
                ("Telefone", 110),
            ]

        # Cabeçalho
        def cabecalho(y_pos):
            x = margem_x
            for titulo, largura_col in colunas:
                c.drawString(x, y_pos, titulo)
                x += largura_col
            return y_pos - 15

        y = cabecalho(y)
        c.setFont("Helvetica", 9)

        for idx, r in enumerate(registros, start=1):
            if y < margem_y:
                c.showPage()

                y = self._desenhar_cabecalho_pdf(
                    c,
                    largura,
                    altura,
                    margem_x,
                    data_sel,
                    registros
                )

                c.setFont("Helvetica-Bold", 10)
                y = cabecalho(y)
                c.setFont("Helvetica", 9)


            x = margem_x
            c.drawString(x, y, str(idx))
            x += 40

            c.drawString(x, y, r.get("protocolo", ""))
            x += 80

            c.drawString(x, y, r.get("nome", ""))
            x += 220

            c.drawString(x, y, r.get("cpf", ""))
            x += 110

            c.drawString(x, y, r.get("telefone", ""))

            y -= 15

        c.drawRightString(
            largura - margem_x,
            margem_y + 10,
            f"Página 1"
        )
        rodape_y = 40

        c.setFont("Helvetica", 9)
        c.drawString(
            margem_x,
            rodape_y,
            f"Data de Emissão: {data_emissao}"
)
        c.save()
        messagebox.showinfo("OK", f"PDF institucional com validação gerado: {filename}")
        try:
            os.startfile(filename)
        except:
            pass

    # Exporta PDF por ação serviço
    def exportar_pdf_por_acao_servico(self):
        registros = self.db["cadastros"]
        if not registros:
            messagebox.showwarning("Aviso", "Nada para exportar.")
            return

        grupos = {}
        for r in registros:
            chave = (r.get("acao", "Sem ação"), r.get("servico", "Sem serviço"))
            grupos.setdefault(chave, []).append(r)

        for (acao, servico), itens in grupos.items():
            nome_arquivo = caminho_pdf("cadastro", "cadastro")

            c = canvas.Canvas(nome_arquivo, pagesize=A4)
            largura, altura = A4
            margem_x = 40
            margem_y = 50
            y = altura - margem_y

            c.setFont("Helvetica-Bold", 16)

            y -= 20

            c.setFont("Helvetica", 11)
            c.drawString(margem_x, y, f"Ação: {acao}")
            y -= 15
            c.drawString(margem_x, y, f"Serviço: {servico}")
            y -= 20

            c.line(margem_x, y, largura - margem_x, y)
            y -= 20

            c.setFont("Helvetica-Bold", 10)
            colunas = [
                ("Protocolo", 80),
                ("Nome", 200),
                ("CPF", 110),
                ("Telefone", 110),
            ]

            x = margem_x
            for t, w in colunas:
                c.drawString(x, y, t)
                x += w
            y -= 15

            c.setFont("Helvetica", 9)
            for r in itens:
                if y < margem_y:
                    c.showPage()
                    y = altura - margem_y

                x = margem_x
                c.drawString(x, y, r.get("protocolo", ""))
                c.drawString(x + 80, y, r.get("nome", ""))
                c.drawString(x + 280, y, r.get("cpf", ""))
                c.drawString(x + 390, y, r.get("telefone", ""))
                y -= 14

            c.save()

        messagebox.showinfo("OK", "Relatórios separados por ação e serviço gerados com sucesso.")

    # Carrega usuarios login
    def _carregar_usuarios_login(self):
            self.tree_login_users.delete(*self.tree_login_users.get_children())

            cur = self.db.cursor()
            cur.execute("""
                SELECT nome, cpf
                FROM usuarios
                ORDER BY nome
            """)

            for nome, cpf in cur.fetchall():
                self.tree_login_users.insert("", "end", values=(nome, cpf))

    # Tela entrega identidades
    def tela_entrega_identidades(self):
        for widget in self.winfo_children():
            widget.destroy()

        self.update_idletasks()
        self._ajustar_janela_principal()

        main = ctk.CTkFrame(self, fg_color="transparent")
        main.pack(fill="both", expand=True, padx=10, pady=10)

        self._criar_cabecalho(
            main,
            "ENTREGA DE IDENTIDADES",
            voltar=self._build_home_screen,
            mostrar_sair=True,
            texto_cancelar="Voltar"
        )

        container = ctk.CTkFrame(main, fg_color="transparent")
        container.pack(fill="both", expand=True, padx=15, pady=15)

        self._criar_titulo_pagina(
            container,
            "Gerencie as identidades prontas para entrega",
            subtitulo="Pesquise, monte a lista final e gere a entrega"
        )

        area_rolavel = ctk.CTkScrollableFrame(container, fg_color="transparent")
        area_rolavel.pack(fill="both", expand=True, pady=(10, 0))

        frame_top = ctk.CTkFrame(area_rolavel, fg_color="transparent")
        frame_top.pack(fill="both", expand=False)

        frame_top.columnconfigure(0, weight=0, minsize=230)
        frame_top.columnconfigure(1, weight=1)

        frame_pesquisa = tb.Labelframe(frame_top, text="Pesquisar Usuários", padding=10)
        frame_pesquisa.grid(row=0, column=0, sticky="nsew", padx=(0,10))

        ctk.CTkLabel(frame_pesquisa, text="Pesquisar por:", text_color=("#334155", "#CBD5E1")).pack(anchor="w", pady=(0, 5))
        self.tipo_pesquisa = tk.StringVar(value="protocolo")
        for texto, valor in [
            ("Protocolo", "protocolo"),
            ("CPF", "cpf"),
            ("Nome", "nome"),
            ("Data da Ação", "data"),
            ("Nome da Ação", "acao"),
            ("Todos", "todos")
        ]:
            tb.Radiobutton(
                frame_pesquisa,
                text=texto,
                variable=self.tipo_pesquisa,
                value=valor,
                command=self.configurar_campo_pesquisa
            ).pack(anchor="w")

        ctk.CTkLabel(frame_pesquisa, text="Valor:", text_color=("#334155", "#CBD5E1")).pack(anchor="w", pady=(10, 2))
        self.entry_pesquisa = ctk.CTkEntry(frame_pesquisa)
        self.entry_pesquisa.pack(fill="x", pady=(0, 8))

        ctk.CTkButton(frame_pesquisa, text="Pesquisar", command=lambda: self.executar_pesquisa(None), fg_color="#2563EB", hover_color="#1D4ED8", text_color="white", corner_radius=8, height=34, font=("Segoe UI", 12, "bold")).pack(fill="x", pady=5)

        ctk.CTkButton(frame_pesquisa, text="Limpar pesquisa", command=self.limpar_resultado_pesquisa, fg_color="#2563EB", hover_color="#1D4ED8", text_color="white", corner_radius=8, height=34, font=("Segoe UI", 12, "bold")).pack(fill="x", pady=5)

        ctk.CTkButton(frame_pesquisa, text="Limpar lista final", command=self.limpar_lista_final, fg_color="#2563EB", hover_color="#1D4ED8", text_color="white", corner_radius=8, height=34, font=("Segoe UI", 12, "bold")).pack(fill="x", pady=5)

        ctk.CTkButton(frame_pesquisa, text="Carregar lista final", command=self.carregar_lista_entrega_salva, fg_color="#2563EB", hover_color="#1D4ED8", text_color="white", corner_radius=8, height=34, font=("Segoe UI", 12, "bold")).pack(fill="x", pady=5)

        ctk.CTkButton(frame_pesquisa, text="Visualizar identidades entregues", command=self.tela_identidades_entregues, fg_color="#2563EB", hover_color="#1D4ED8", text_color="white", corner_radius=8, height=34, font=("Segoe UI", 12, "bold")).pack(fill="x", pady=5)

        frame_resultado = tb.Labelframe(frame_top, text="Resultados da Pesquisa", padding=10)
        frame_resultado.grid(row=0, column=1, sticky="nsew")

        self._resultados_pesquisa_entrega = []
        self._filtro_resultado_entrega = tk.StringVar()
        self._enter_resultado_pendente = False

        ctk.CTkLabel(frame_resultado, text="Filtrar resultados:", text_color=("#334155", "#CBD5E1")).pack(anchor="w", pady=(10, 2))
        self.entry_filtro_resultado_entrega = ctk.CTkEntry(
            frame_resultado,
            textvariable=self._filtro_resultado_entrega
        )
        self.entry_filtro_resultado_entrega.pack(fill="x", pady=(0, 8))
        self.entry_filtro_resultado_entrega.bind("<KeyRelease>", self._filtrar_resultados_entrega)
        self.entry_filtro_resultado_entrega.bind("<Return>", self._entrada_enter_filtro_resultado_entrega)

        self.tree_resultado = tb.Treeview(
            frame_resultado,
            columns=("protocolo", "nome", "cpf", "telefone", "acao", "situacao"),
            show="headings",
            height=10
        )

        self.tree_resultado.heading("protocolo", text="Protocolo")
        self.tree_resultado.heading("nome", text="Nome")
        self.tree_resultado.heading("cpf", text="CPF")
        self.tree_resultado.heading("telefone", text="Telefone")
        self.tree_resultado.heading("acao", text="Ação")
        self.tree_resultado.heading("situacao", text="Situação")

        self.tree_resultado.column("protocolo", width=90, minwidth=80, anchor="w", stretch=False)
        self.tree_resultado.column("nome", width=260, minwidth=180, anchor="w")
        self.tree_resultado.column("cpf", width=125, minwidth=110, anchor="w", stretch=False)
        self.tree_resultado.column("telefone", width=135, minwidth=120, anchor="w", stretch=False)
        self.tree_resultado.column("acao", width=220, minwidth=160, anchor="w")
        self.tree_resultado.column("situacao", width=130, minwidth=120, anchor="w", stretch=False)
        self.tree_resultado.tag_configure("entregue", background="#ffe5e5")
        self.tree_resultado.pack(fill="both", expand=True)
        self._preparar_tabela(self.tree_resultado, minimo=80, maximo=520)


        botoes_lista = ctk.CTkFrame(frame_resultado, fg_color="transparent")
        botoes_lista.pack(anchor="w", pady=10)

        ctk.CTkButton(botoes_lista, text="Adicionar à Lista", command=self.adicionar_lista_entrega, fg_color="#16A34A", hover_color="#15803D", text_color="white", corner_radius=8, height=34, font=("Segoe UI", 12, "bold")).pack(side="left", padx=(0, 6))

        ctk.CTkButton(botoes_lista, text="Adicionar Todos", command=self.adicionar_todos_entrega, fg_color="#D97706", hover_color="#B45309", text_color="white", corner_radius=8, height=34, font=("Segoe UI", 12, "bold")).pack(side="left", padx=6)

        frame_lista_container = ctk.CTkFrame(area_rolavel, fg_color="transparent")
        frame_lista_container.pack(fill="both", expand=True, pady=8)
        frame_lista_container.columnconfigure(0, weight=1)
        frame_lista_container.columnconfigure(1, weight=0, minsize=140)

        frame_lista = tb.Labelframe(frame_lista_container, text="Lista Final para Entrega", padding=10)
        frame_lista.grid(row=0, column=0, sticky="nsew")
        frame_lista.columnconfigure(0, weight=1)
        frame_lista.rowconfigure(0, weight=1)

        self.tree_lista = tb.Treeview(
            frame_lista,
            columns=("protocolo", "nome", "cpf", "telefone"),
            show="headings",
            height=10
        )

        self.tree_lista.heading("protocolo", text="Protocolo")
        self.tree_lista.heading("nome", text="Nome")
        self.tree_lista.heading("cpf", text="CPF")
        self.tree_lista.heading("telefone", text="Telefone")

        self.tree_lista.column("protocolo", width=90, minwidth=80, anchor="w", stretch=False)
        self.tree_lista.column("nome", width=280, minwidth=180, anchor="w")
        self.tree_lista.column("cpf", width=125, minwidth=110, anchor="w", stretch=False)
        self.tree_lista.column("telefone", width=135, minwidth=120, anchor="w", stretch=False)

        self.tree_lista.grid(row=0, column=0, sticky="nsew")

        scroll_y = tb.Scrollbar(frame_lista, orient="vertical", command=self.tree_lista.yview)
        scroll_y.grid(row=0, column=1, sticky="ns")
        self.tree_lista.configure(yscrollcommand=scroll_y.set)

        self._preparar_tabela(self.tree_lista, minimo=80, maximo=420)
        carregar_lista_tree(self.tree_lista)

        btn_pdf = ctk.CTkFrame(frame_lista_container, fg_color="transparent")
        btn_pdf.grid(row=0, column=1, sticky="ns", padx=(10, 0), pady=10)

        ctk.CTkButton(btn_pdf, text="Gerar PDF", command=self.gerar_pdf_entregas, fg_color="#2563EB", hover_color="#1D4ED8", text_color="white", corner_radius=8, height=34, font=("Segoe UI", 12, "bold")).pack(padx=10, pady=5, anchor="e")

        self.tree_resultado.bind("<Double-1>", self.duplo_clique_resultado)
        self.tree_resultado.bind("<Return>", self._entrada_enter_arvore_resultado_entrega)
        self.tree_resultado.bind("<<TreeviewSelect>>", self._marcar_selecao_resultado_feita)
        self.tree_resultado.bind("<FocusIn>", self._marcar_selecao_resultado_feita)
        self.entry_pesquisa.bind("<Return>", self._entrada_enter_pesquisa_entrega)
        self.entry_pesquisa.bind("<Down>", self._mover_selecao_resultado_entrega)
        self.entry_pesquisa.bind("<Up>", self._mover_selecao_resultado_entrega)

        self.tree_lista.bind("<Double-1>", self.clique_lista_final)


    # Abre pesquisa entrega
    def abrir_pesquisa_entrega(self):
        janela = self._open_inline_container("Pesquisar Entrega", voltar=self.tela_entrega_identidades)
        janela.title("Pesquisar")
        janela.geometry("350x250")

        ctk.CTkLabel(janela, text="Pesquisar por:", text_color=("#334155", "#CBD5E1")).pack(pady=10)

        self.tipo_pesquisa = tk.StringVar()

        opcoes = [
            ("Protocolo", "protocolo"),
            ("CPF", "cpf"),
            ("Nome", "nome"),
            ("Data da Ação", "data"),
            ("Nome da Ação", "acao")
        ]

        for texto, valor in opcoes:
            tb.Radiobutton(
                janela,
                text=texto,
                variable=self.tipo_pesquisa,
                value=valor,
                command=self.configurar_campo_pesquisa
            ).pack(anchor="w", padx=20)

        ctk.CTkLabel(janela, text="Digite o valor:", text_color=("#334155", "#CBD5E1")).pack(pady=5)

        self.entry_pesquisa = ctk.CTkEntry(janela)
        self.entry_pesquisa.pack(pady=5)

        ctk.CTkButton(janela, text="Pesquisar", command=lambda: self.executar_pesquisa(janela), fg_color="#16A34A", hover_color="#15803D", text_color="white", corner_radius=8, height=34, font=("Segoe UI", 12, "bold")).pack(pady=10)

    # Configura campo pesquisa
    def configurar_campo_pesquisa(self):

        tipo = self.tipo_pesquisa.get()

        self.entry_pesquisa.unbind("<KeyRelease>")
        self.entry_pesquisa.configure(state="normal")
        self.entry_pesquisa.delete(0, "end")

        if tipo == "protocolo":

            vcmd = (self.register(self.validar_numeros), "%P")
            self.entry_pesquisa.configure(validate="key", validatecommand=vcmd)
            self.entry_pesquisa.bind("<KeyRelease>", self.pesquisa_automatica_entrega)

        elif tipo == "cpf":

            self.entry_pesquisa.configure(validate="none")
            self.entry_pesquisa.bind("<KeyRelease>", self.mascara_cpf_e_pesquisa_automatica)

        elif tipo == "nome":

            self.entry_pesquisa.configure(validate="none")
            self.entry_pesquisa.bind("<KeyRelease>", self.pesquisa_automatica_entrega)

        elif tipo in ("data", "data_entrega"):

            self.entry_pesquisa.configure(validate="none")
            self.entry_pesquisa.bind("<KeyRelease>", self.mascara_data)

        elif tipo == "todos":

            self.entry_pesquisa.configure(validate="none", state="disabled")

        else:

            self.entry_pesquisa.configure(validate="none")

        if self.tipo_pesquisa.get() != "todos":
            self.entry_pesquisa.delete(0, tk.END)
            self.entry_pesquisa.focus_set()

    # Configura campo pesquisa entregues
    def configurar_campo_pesquisa_entregues(self):

        tipo = self.tipo_pesquisa.get()

        self.entry_pesquisa.unbind("<KeyRelease>")
        self.entry_pesquisa.configure(state="normal")
        self.entry_pesquisa.delete(0, tk.END)

        if tipo == "protocolo":

            vcmd = (self.register(self.validar_numeros), "%P")
            self.entry_pesquisa.configure(
                validate="key",
                validatecommand=vcmd
            )

        elif tipo == "todos":

            self.entry_pesquisa.configure(validate="none", state="disabled")

            if hasattr(self, "tree_entregues"):
                self.executar_pesquisa_entregues()
            return

        else:

            self.entry_pesquisa.configure(validate="none")

        self.entry_pesquisa.bind(
            "<KeyRelease>",
            lambda e: self.executar_pesquisa_entregues()
        )

        self.entry_pesquisa.focus_set()

        if hasattr(self, "tree_entregues"):
            for item in self.tree_entregues.get_children():
                self.tree_entregues.delete(item)

    # Valida numeros
    def validar_numeros(self, valor):
        return valor.isdigit() or valor == ""

    # Mascara CPF
    def mascara_cpf(self, event):

        texto = self.entry_pesquisa.get()
        numeros = "".join(filter(str.isdigit, texto))

        novo = ""

        if len(numeros) > 0:
            novo += numeros[:3]

        if len(numeros) >= 4:
            novo += "." + numeros[3:6]

        if len(numeros) >= 7:
            novo += "." + numeros[6:9]

        if len(numeros) >= 10:
            novo += "-" + numeros[9:11]

        self.entry_pesquisa.delete(0, "end")
        self.entry_pesquisa.insert(0, novo)

    # Mascara CPF e pesquisa automatica
    def mascara_cpf_e_pesquisa_automatica(self, event):
        self.mascara_cpf(event)
        self.pesquisa_automatica_entrega(event)

    # Pesquisa automatica entrega
    def pesquisa_automatica_entrega(self, event=None):
        tipo = self.tipo_pesquisa.get()
        if tipo not in ("protocolo", "cpf", "nome"):
            return
        if not hasattr(self, "tree_resultado") or not self.tree_resultado.winfo_exists():
            return
        self.executar_pesquisa(None)

    # Mascara data
    def mascara_data(self, event):

        texto = self.entry_pesquisa.get()
        numeros = "".join(filter(str.isdigit, texto))

        novo = ""

        if len(numeros) > 0:
            novo += numeros[:2]

        if len(numeros) >= 3:
            novo += "/" + numeros[2:4]

        if len(numeros) >= 5:
            novo += "/" + numeros[4:8]

        self.entry_pesquisa.delete(0, "end")
        self.entry_pesquisa.insert(0, novo)

    # Executar pesquisa
    def executar_pesquisa(self, janela):

        tipo = self.tipo_pesquisa.get()

        valor = "" if tipo == "todos" else normalizar_texto_ordenacao(self.entry_pesquisa.get())

        if not tipo:
            messagebox.showwarning("Aviso", "Selecione o tipo de pesquisa.")
            return

        self._pausar_auto_ajuste_tabela(self.tree_resultado)
        try:
            for item in self.tree_resultado.get_children():
                self.tree_resultado.delete(item)

            resultados = []

            for registro in self.db.get("cadastros", []):
                if tipo == "todos":
                    resultados.append(registro)
                    continue

                if tipo == "acao":
                    campo = self._nome_acao_registro(registro)
                else:
                    campo = str(registro.get(tipo, ""))

                if valor in normalizar_texto_ordenacao(campo):
                    resultados.append(registro)

            resultados = sorted(
                resultados,
                key=lambda x: normalizar_texto_ordenacao(x.get("nome"))
            )

            self._resultados_pesquisa_entrega = resultados
            self._filtro_resultado_entrega.set("")
            self._enter_resultado_pendente = bool(resultados)
            self._renderizar_resultados_entrega()
        finally:
            self._retomar_auto_ajuste_tabela(self.tree_resultado)

        if janela is not None:
            self._fechar_tela_interna(janela)

    # Renderizar resultados entrega
    def _renderizar_resultados_entrega(self):
        self._pausar_auto_ajuste_tabela(self.tree_resultado)
        try:
            for item in self.tree_resultado.get_children():
                self.tree_resultado.delete(item)

            termo = normalizar_texto_ordenacao(self._filtro_resultado_entrega.get())

            for registro in self._resultados_pesquisa_entrega:
                texto_busca = " ".join([
                    str(registro.get("protocolo", "")),
                    str(registro.get("nome", "")),
                    str(registro.get("cpf", "")),
                    str(registro.get("telefone", "")),
                    str(self._nome_acao_registro(registro))
                ])

                if termo and termo not in normalizar_texto_ordenacao(texto_busca):
                    continue

                entregue = bool(registro.get("entregue", False))
                situacao = "Entregue" if entregue else "Aguardando entrega"
                tags = ("entregue",) if entregue else ()
                self.tree_resultado.insert(
                    "",
                    "end",
                    values=(
                        registro.get("protocolo", ""),
                        registro.get("nome", ""),
                        registro.get("cpf", ""),
                        registro.get("telefone", ""),
                        self._nome_acao_registro(registro),
                        situacao
                    ),
                    tags=tags
                )
        finally:
            self._retomar_auto_ajuste_tabela(self.tree_resultado)

    # Filtrar resultados entrega
    def _filtrar_resultados_entrega(self, event=None):
        self._renderizar_resultados_entrega()
        self._enter_resultado_pendente = bool(self.tree_resultado.get_children())

    # Entrada enter pesquisa entrega
    def _entrada_enter_pesquisa_entrega(self, event=None):
        tipo = self.tipo_pesquisa.get()

        if tipo in ("protocolo", "cpf", "nome"):
            self._selecionar_ou_adicionar_resultado_entrega()
        elif tipo in ("data", "todos"):
            self.executar_pesquisa(None)
            self._selecionar_primeiro_resultado_entrega()

        return "break"

    # Entrada enter filtro resultado entrega
    def _entrada_enter_filtro_resultado_entrega(self, event=None):
        self._selecionar_ou_adicionar_resultado_entrega()
        return "break"

    # Entrada enter árvore resultado entrega
    def _entrada_enter_arvore_resultado_entrega(self, event=None):
        self.adicionar_lista_entrega()
        return "break"

    # Seleciona ou adiciona resultado entrega
    def _selecionar_ou_adicionar_resultado_entrega(self):
        if getattr(self, "_enter_resultado_pendente", False):
            self._selecionar_primeiro_resultado_entrega()
            return

        self.adicionar_lista_entrega()

    # Seleciona primeiro resultado entrega
    def _selecionar_primeiro_resultado_entrega(self):
        itens = self.tree_resultado.get_children()

        if not itens:
            return

        primeiro = itens[0]
        self.tree_resultado.selection_set(primeiro)
        self.tree_resultado.focus(primeiro)
        self.tree_resultado.see(primeiro)
        self.tree_resultado.focus_set()
        self._enter_resultado_pendente = False

    # Mover seleção resultado entrega
    def _mover_selecao_resultado_entrega(self, event=None):
        itens = self.tree_resultado.get_children()

        if not itens:
            return "break"

        itens_lista = list(itens)
        atual = self.tree_resultado.focus()
        indice_atual = itens_lista.index(atual) if atual in itens_lista else -1

        if event.keysym == "Down":
            novo_indice = (indice_atual + 1) % len(itens_lista)
        else:
            novo_indice = (indice_atual - 1) % len(itens_lista)

        item = itens_lista[novo_indice]
        self.tree_resultado.selection_set(item)
        self.tree_resultado.focus(item)
        self.tree_resultado.see(item)
        self.tree_resultado.focus_set()
        self._enter_resultado_pendente = False

        return "break"

    # Marcar seleção resultado feita
    def _marcar_selecao_resultado_feita(self, event=None):
        self._enter_resultado_pendente = False

    # Ação base lista entrega
    def _acao_base_lista_entrega(self, registros):
        grupos = []

        for registro in registros:
            acao = get_acao_por_id(self.db, registro.get("acao_id"))
            if acao:
                chave = (acao.get("data", "Sem data"), acao.get("local", "Sem local"))
            else:
                chave = (registro.get("data", "Sem data"), self._nome_acao_registro(registro) or "Sem local")

            if chave not in grupos:
                grupos.append(chave)

        if len(grupos) == 1:
            return grupos[0]

        return ("multiplas_datas", "multiplas_acoes")

    # Nome base lista entrega
    def _nome_base_lista_entrega(self, registros):
        data_acao, nome_acao = self._acao_base_lista_entrega(registros)
        data_acao = nome_arquivo_seguro(str(data_acao).replace("/", "-"))
        nome_acao = nome_arquivo_seguro(nome_acao)
        data_hora = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
        return f"{data_acao}_{nome_acao}_{data_hora}"

    # Caminho PDF lista entrega
    def _caminho_pdf_lista_entrega(self, registros):
        pasta = os.path.join(APP_DIR, "pdf", "entrega")
        os.makedirs(pasta, exist_ok=True)
        base = self._nome_base_lista_entrega(registros)
        caminho = os.path.join(pasta, f"{base}.pdf")
        contador = 2

        while os.path.exists(caminho):
            caminho = os.path.join(pasta, f"{base}_{contador}.pdf")
            contador += 1

        return caminho

    # Salva lista entrega arquivada
    def _salvar_lista_entrega_arquivada(self, registros, caminho_pdf):
        nome_json = f"{os.path.splitext(os.path.basename(caminho_pdf))[0]}.json"
        metadados = {
            "pdf": caminho_pdf,
            "gerado_em": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            "total": len(registros)
        }
        salvar_lista_tree(self.tree_lista, nome_json, metadados=metadados)

    # Arquivos lista entrega
    def _arquivos_lista_entrega(self):
        os.makedirs(PASTA_LISTAS, exist_ok=True)
        arquivos = []

        for nome in os.listdir(PASTA_LISTAS):
            if nome.lower().endswith(".json"):
                caminho = os.path.join(PASTA_LISTAS, nome)
                if os.path.isfile(caminho):
                    arquivos.append(nome)

        arquivos.sort(key=lambda nome: os.path.getmtime(os.path.join(PASTA_LISTAS, nome)), reverse=True)
        return arquivos

    # Carrega lista entrega salva
    def carregar_lista_entrega_salva(self):
        arquivos = self._arquivos_lista_entrega()
        if not arquivos:
            messagebox.showinfo("Listas", "Nenhuma lista salva encontrada.")
            return

        janela = tk.Toplevel(self)
        janela.title("Carregar lista final")
        janela.geometry("520x360")
        janela.transient(self)
        janela.grab_set()

        wrapper = tb.Frame(janela, padding=12)
        wrapper.pack(fill="both", expand=True)
        wrapper.columnconfigure(0, weight=1)
        wrapper.rowconfigure(1, weight=1)

        header = tb.Frame(wrapper, bootstyle="primary", padding=(12, 10))
        header.grid(row=0, column=0, sticky="ew")
        header.columnconfigure(0, weight=1)
        tb.Label(header, text="Carregar lista final", font=("Segoe UI", 16, "bold"), bootstyle="inverse-primary").grid(row=0, column=0, sticky="w")
        tb.Button(header, text="Cancelar", bootstyle="light", width=12, command=janela.destroy).grid(row=0, column=1, sticky="e")

        frame = tb.Frame(wrapper, padding=12)
        frame.grid(row=1, column=0, sticky="nsew")

        tb.Label(frame, text="Selecione uma lista salva:", font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=(0, 8))

        lista = tk.Listbox(frame, height=12)
        lista.pack(fill="both", expand=True)

        for nome in arquivos:
            lista.insert(tk.END, nome)

        # Carrega selecionada
        def carregar_selecionada():
            selecao = lista.curselection()
            if not selecao:
                messagebox.showwarning("Aviso", "Selecione uma lista.")
                return

            nome = arquivos[selecao[0]]
            for item in self.tree_lista.get_children():
                self.tree_lista.delete(item)

            carregar_lista_tree(self.tree_lista, nome)
            salvar_lista_tree(self.tree_lista)
            janela.destroy()

        botoes = tb.Frame(frame)
        botoes.pack(fill="x", pady=(10, 0))

        tb.Button(
            botoes,
            text="Carregar",
            bootstyle="success",
            command=carregar_selecionada
        ).pack(side="left", padx=(0, 8))

    # Adiciona lista entrega
    def adicionar_lista_entrega(self):
        selecionado = self.tree_resultado.focus()

        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione um registro.")
            return

        dados = self.tree_resultado.item(selecionado)["values"]
        protocolo_novo = str(dados[0]).strip()

        for item in self.tree_lista.get_children():
            valores = self.tree_lista.item(item)["values"]
            protocolo_existente = str(valores[0]).strip()

            if protocolo_existente == protocolo_novo:
                messagebox.showwarning("Aviso", "Este registro já está na lista.")
                return

        self.tree_lista.insert("", "end", values=dados[:4])
        self.tree_lista.yview_moveto(1)
        salvar_lista_tree(self.tree_lista)

        self.entry_pesquisa.delete(0, tk.END)
        self.entry_pesquisa.focus_set()

    # Adiciona todos entrega
    def adicionar_todos_entrega(self):

        itens = self.tree_resultado.get_children()

        if not itens:
            messagebox.showwarning("Aviso", "Não há registros na pesquisa.")
            return

        adicionados = 0

        for item in itens:

            dados = self.tree_resultado.item(item)["values"]
            protocolo_novo = str(dados[0]).strip()

            duplicado = False

            for item_lista in self.tree_lista.get_children():
                valores = self.tree_lista.item(item_lista)["values"]
                protocolo_existente = str(valores[0]).strip()

                if protocolo_existente == protocolo_novo:
                    duplicado = True
                    break

            if not duplicado:
                self.tree_lista.insert("", "end", values=dados[:4])
                adicionados += 1

        if adicionados:
            salvar_lista_tree(self.tree_lista)

        messagebox.showinfo(
            "Concluído",
            f"{adicionados} registros adicionados à lista de entrega."
        )
        self.entry_pesquisa.delete(0, tk.END)
        self.entry_pesquisa.focus_set()

    # Duplo clique resultado
    def duplo_clique_resultado(self, event):

            item = self.tree_resultado.identify_row(event.y)

            if item:
                self.tree_resultado.selection_set(item)
                self.tree_resultado.focus(item)

                self.adicionar_lista_entrega()

                self.entry_pesquisa.delete(0, tk.END)
                self.entry_pesquisa.focus_set()

    # Limpa resultado pesquisa
    def limpar_resultado_pesquisa(self):

        self._resultados_pesquisa_entrega = []
        self._filtro_resultado_entrega.set("")
        self._enter_resultado_pendente = False

        for item in self.tree_resultado.get_children():
            self.tree_resultado.delete(item)

    # Limpa lista final
    def limpar_lista_final(self):
        for item in self.tree_lista.get_children():
            self.tree_lista.delete(item)

        salvar_lista_tree(self.tree_lista)

    # Clique lista final
    def clique_lista_final(self, event):

        item = self.tree_lista.identify_row(event.y)

        if not item:
            return

        valores = self.tree_lista.item(item)["values"]
        nome = valores[1]

        confirmar = messagebox.askyesno(
            "Confirmar exclusão",
            f"Deseja remover {nome} da lista de entrega?"
        )

        if confirmar:
            self.tree_lista.delete(item)
            salvar_lista_tree(self.tree_lista)

    # Cadastro por protocolo
    def _cadastro_por_protocolo(self, protocolo):
        protocolo = str(protocolo).strip()
        for registro in self.db.get("cadastros", []):
            if str(registro.get("protocolo", "")).strip() == protocolo:
                return registro
        return None

    # Protocolos entregues confirmados
    def _protocolos_entregues_confirmados(self):
        protocolos = set()
        for entrega in self._carregar_entregas_confirmadas():
            protocolo = str(entrega.get("protocolo", "")).strip()
            if protocolo:
                protocolos.add(protocolo)

        return protocolos

    # Registro esta entregue
    def _registro_esta_entregue(self, registro):
        return bool(registro.get("entregue", False))

    # Coletar registros lista entrega
    def _coletar_registros_lista_entrega(self, tree_alvo=None):
        tree = tree_alvo or self.tree_lista
        registros = []
        for item in tree.get_children():
            valores = tree.item(item)["values"]
            if not valores:
                continue
            protocolo = str(valores[0]).strip() if valores else ""
            registro = self._cadastro_por_protocolo(protocolo)
            if registro:
                registros.append(registro)
        return registros

    # Remove registros lista entrega
    def _remover_registros_lista_entrega(self, registros, tree_alvo=None):
        tree = tree_alvo or self.tree_lista
        protocolos = {str(r.get("protocolo", "")).strip() for r in registros}
        removeu = False
        for item in list(tree.get_children()):
            valores = tree.item(item)["values"]
            if valores and str(valores[0]).strip() in protocolos:
                tree.delete(item)
                removeu = True

        if removeu and tree is self.tree_lista:
            salvar_lista_tree(self.tree_lista)

    # Registra entrega confirmada
    def _registrar_entrega_confirmada(self, registros):
        timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        for registro in registros:
            registro["entregue"] = True
            registro["data_entrega"] = timestamp
            registro["operador_entrega"] = self.usuario or "N/A"

        save_db(self.db)
        registrar_entrega_json_separado(registros, self.usuario or "N/A")

    # Nome ação registro
    def _nome_acao_registro(self, registro):
        return (
            registro.get("acao")
            or get_nome_acao_por_id(self.db, registro.get("acao_id"))
            or ""
        )

    # Carrega entregas confirmadas
    def _carregar_entregas_confirmadas(self):
        if not os.path.exists(DB_ENTREGAS):
            return []

        try:
            with open(DB_ENTREGAS, "r", encoding="utf-8") as f:
                entregas = json.load(f)
        except Exception:
            return []

        return entregas if isinstance(entregas, list) else []

    # Normaliza entrega confirmada
    def _normalizar_entrega_confirmada(self, entrega, cadastros_por_protocolo=None):
        protocolo = str(entrega.get("protocolo", "")).strip()
        if cadastros_por_protocolo is None:
            registro_db = self._cadastro_por_protocolo(protocolo)
        else:
            registro_db = cadastros_por_protocolo.get(protocolo)

        registro = {
            "protocolo": entrega.get("protocolo", ""),
            "nome": entrega.get("nome", ""),
            "cpf": entrega.get("cpf", ""),
            "telefone": entrega.get("telefone", ""),
            "data": entrega.get("data_acao", ""),
            "acao": entrega.get("acao", ""),
            "acao_id": entrega.get("acao_id", ""),
            "servico_id": entrega.get("servico_id", "")
        }

        if registro_db:
            for campo in ("nome", "cpf", "telefone", "data", "acao", "acao_id", "servico_id"):
                if not registro.get(campo):
                    registro[campo] = registro_db.get(campo, "")

        registro["entregue"] = True
        registro["data_entrega"] = entrega.get("data_entrega_registro", "")
        registro["operador_entrega"] = entrega.get("operador", "")
        return registro

    # Coletar identidades entregues
    def _coletar_identidades_entregues(self):
        cadastros_por_protocolo = {
            str(registro.get("protocolo", "")).strip(): registro
            for registro in self.db.get("cadastros", [])
            if str(registro.get("protocolo", "")).strip()
        }

        return [
            r for r in self.db.get("cadastros", [])
            if r.get("entregue", False)
    ]

    # Tela identidades entregues
    def tela_identidades_entregues(self):
        for widget in self.winfo_children():
            widget.destroy()

        self.update_idletasks()
        self._ajustar_janela_principal()

        main = tb.Frame(self, padding=10)
        main.pack(fill="both", expand=True)

        self._criar_cabecalho(
            main,
            "IDENTIDADES ENTREGUES",
            voltar=self.tela_entrega_identidades,
            mostrar_sair=True,
            texto_cancelar="Voltar"
        )

        container = tb.Frame(main, padding=15)
        container.pack(fill="both", expand=True)

        self._criar_titulo_pagina(
            container,
            "Consulte identidades entregues",

        )

        frame_top = tb.Frame(container)
        frame_top.pack(fill="x")
        frame_top.columnconfigure(0, weight=0, minsize=230)
        frame_top.columnconfigure(1, weight=1)

        frame_pesquisa = tb.Labelframe(frame_top, text="Pesquisar Entregues", padding=10)
        frame_pesquisa.grid(row=0, column=0, sticky="nsew", padx=(0, 10))

        tb.Label(frame_pesquisa, text="Pesquisar por:").pack(anchor="w", pady=(0, 5))
        self.tipo_pesquisa = tk.StringVar(value="todos")
        for texto, valor in [
            ("Todos", "todos"),
            ("Protocolo", "protocolo"),
            ("CPF", "cpf"),
            ("Nome", "nome"),
            ("Data da Ação", "data"),
            ("Data da Entrega", "data_entrega"),
            ("Nome da Ação", "acao")
        ]:
            tb.Radiobutton(
                frame_pesquisa,
                text=texto,
                variable=self.tipo_pesquisa,
                value=valor,
                command=self.configurar_campo_pesquisa_entregues
            ).pack(anchor="w")

        tb.Label(frame_pesquisa, text="Valor:").pack(anchor="w", pady=(10, 2))
        self.entry_pesquisa = tb.Entry(frame_pesquisa)
        self.entry_pesquisa.pack(fill="x", pady=(0, 8))

        self.entry_pesquisa.bind(
            "<KeyRelease>",
            lambda e: self.executar_pesquisa_entregues()
        )

        tb.Button(
            frame_pesquisa,
            text="Pesquisar",
            bootstyle="info",
            command=self.executar_pesquisa_entregues
        ).pack(fill="x", pady=5)

        tb.Button(
            frame_pesquisa,
            text="Limpar",
            bootstyle="warning",
            command=self.limpar_identidades_entregues
        ).pack(fill="x", pady=5)

        tb.Button(
            frame_pesquisa,
            text="Gerar PDF",
            bootstyle="primary",
            command=lambda: self.gerar_pdf_entregas(self.tree_entregues)
        ).pack(fill="x", pady=5)

        frame_resultado = tb.Labelframe(frame_top, text="Identidades Entregues", padding=10)
        frame_resultado.grid(row=0, column=1, sticky="nsew")

        colunas = ("protocolo", "nome", "cpf", "telefone", "acao", "data_entrega", "operador")
        self.tree_entregues = tb.Treeview(frame_resultado, columns=colunas, show="headings", height=16)
        for coluna, titulo in [
            ("protocolo", "Protocolo"),
            ("nome", "Nome"),
            ("cpf", "CPF"),
            ("telefone", "Telefone"),
            ("acao", "Ação"),
            ("data_entrega", "Data da Entrega"),
            ("operador", "Operador")
        ]:
            self.tree_entregues.heading(coluna, text=titulo)

        self.tree_entregues.column("protocolo", width=90, minwidth=80, anchor="w", stretch=False)
        self.tree_entregues.column("nome", width=250, minwidth=180, anchor="w")
        self.tree_entregues.column("cpf", width=125, minwidth=110, anchor="w", stretch=False)
        self.tree_entregues.column("telefone", width=135, minwidth=120, anchor="w", stretch=False)
        self.tree_entregues.column("acao", width=210, minwidth=160, anchor="w")
        self.tree_entregues.column("data_entrega", width=145, minwidth=130, anchor="w", stretch=False)
        self.tree_entregues.column("operador", width=120, minwidth=100, anchor="w", stretch=False)
        self.tree_entregues.pack(fill="both", expand=True)
        self._preparar_tabela(self.tree_entregues, minimo=80, maximo=520)
        self.tree_entregues.bind("<Double-1>", self.desfazer_entrega_identidade_selecionada)
        self.configurar_campo_pesquisa_entregues()

    # Limpa identidades entregues
    def limpar_identidades_entregues(self):
        self._pausar_auto_ajuste_tabela(self.tree_entregues)
        for item in self.tree_entregues.get_children():
            self.tree_entregues.delete(item)
        self._retomar_auto_ajuste_tabela(self.tree_entregues)

    # Executar pesquisa entregues
    def executar_pesquisa_entregues(self):
        tipo = self.tipo_pesquisa.get()

        if tipo != "todos":
            valor = normalizar_texto_ordenacao(self.entry_pesquisa.get())

            if not valor:
                for item in self.tree_entregues.get_children():
                    self.tree_entregues.delete(item)
                return
        else:
            valor = ""

        self._pausar_auto_ajuste_tabela(self.tree_entregues)
        try:
            for item in self.tree_entregues.get_children():
                self.tree_entregues.delete(item)

            resultados = []
            for registro in self._coletar_identidades_entregues():
                if tipo == "todos":
                    resultados.append(registro)
                    continue

                if tipo == "acao":
                    campo = self._nome_acao_registro(registro)
                else:
                    campo = str(registro.get(tipo, ""))

                if valor in normalizar_texto_ordenacao(campo):
                    resultados.append(registro)

            resultados = sorted(resultados, key=lambda x: normalizar_texto_ordenacao(x.get("nome")))

            for r in resultados:
                self.tree_entregues.insert(
                    "",
                    "end",
                    values=(
                        r.get("protocolo", ""),
                        r.get("nome", ""),
                        r.get("cpf", ""),
                        r.get("telefone", ""),
                        self._nome_acao_registro(r),
                        r.get("data_entrega", ""),
                        r.get("operador_entrega", "")
                    )
                )
        finally:
            self._retomar_auto_ajuste_tabela(self.tree_entregues)

    # Desfaz entrega identidade selecionada
    def desfazer_entrega_identidade_selecionada(self, event=None):

        selecionado = self.tree_entregues.focus()

        if not selecionado:
            return

        if not self._autorizar_admin():
            return

        valores = self.tree_entregues.item(selecionado)["values"]
        if not valores:
            return

        protocolo_alvo = str(valores[0]).strip()
        nome = str(valores[1]).strip()
        cpf = str(valores[2]).strip()

        confirmar = messagebox.askyesno(
            "Confirmar desfazer entrega",
            "Tem certeza que deseja desfazer a entrega deste cadastro?\n\n"
            f"CPF: {cpf}\n"
            f"Nome: {nome}"
        )
        if not confirmar:
            return

        registro = self._cadastro_por_protocolo(protocolo_alvo)
        if not registro:
            messagebox.showerror("Erro", "Cadastro nao encontrado no db.json.")
            return

        registro["entregue"] = False
        registro.pop("data_entrega", None)
        registro.pop("operador_entrega", None)
        save_db(self.db)

        if os.path.exists(DB_ENTREGAS):
            try:
                with open(DB_ENTREGAS, "r", encoding="utf-8") as f:
                    entregas = json.load(f)
            except Exception:
                entregas = []

            entregas = [
                entrega for entrega in entregas
                if str(entrega.get("protocolo", "")).strip() != protocolo_alvo
            ]
            with open(DB_ENTREGAS, "w", encoding="utf-8") as f:
                json.dump(entregas, f, indent=4, ensure_ascii=False)

        self._registrar_log("Desfez entrega", protocolo_alvo)
        self.executar_pesquisa_entregues()
        messagebox.showinfo("Sucesso", "Entrega desfeita. O cadastro voltou para a lista de pendentes.")




    # Gera PDF entregas
    def gerar_pdf_entregas(self, tree_alvo=None):
        tree = tree_alvo or getattr(self, "tree_lista", None)
        if tree is None or not tree.winfo_exists():
            messagebox.showwarning("Aviso", "Nenhuma tabela disponível.")
            return

        itens = tree.get_children()
        if not itens:
            messagebox.showwarning("Aviso", "Nenhum registro na lista.")
            return

        registros = self._coletar_registros_lista_entrega(tree_alvo=tree)
        if not registros:
            messagebox.showwarning("Aviso", "Nenhum registro encontrado no banco.")
            return

        if tree is self.tree_lista:
            deseja_entregar = messagebox.askyesno(
                "Entrega de identidade",
                "Deseja registrar a lista em tela como identidade entregue?"
            )

            if not deseja_entregar:
                registros = sorted(
                    registros,
                    key=lambda x: normalizar_texto_ordenacao(x.get("nome"))
                )
                nome_arquivo = self._caminho_pdf_lista_entrega(registros)
                self._gerar_pdf_entrega(registros, nome_arquivo)
                return

            ja_entregues = [r for r in registros if r.get("entregue")]

            if len(ja_entregues) == 1:
                r = ja_entregues[0]
                messagebox.showwarning(
                    "Carteira já entregue",
                    "Esta carteira já foi entregue.\n\n"
                    f"Protocolo: {r.get('protocolo', '')}\n"
                    f"Nome: {r.get('nome', '')}\n\n"
                    "Retire este registro da lista para continuar."
                )
                return

            if len(ja_entregues) > 1:
                remover = messagebox.askyesno(
                    "Carteiras já entregues",
                    f"{len(ja_entregues)} carteiras desta lista já foram entregues.\n\n"
                    "Deseja retirar automaticamente esses registros da lista e continuar?"
                )
                if not remover:
                    return

                self._remover_registros_lista_entrega(ja_entregues, tree_alvo=tree)
                protocolos_entregues = {str(r.get("protocolo", "")).strip() for r in ja_entregues}
                registros = [
                    r for r in registros
                    if str(r.get("protocolo", "")).strip() not in protocolos_entregues
                ]

                if not registros:
                    messagebox.showwarning(
                        "Aviso",
                        "Todos os registros da lista já estavam entregues."
                    )
                    return

            self._registrar_entrega_confirmada(registros)
            messagebox.showinfo(
                "Entrega registrada",
                f"{len(registros)} identidade(s) registrada(s) como entregue(s)."
)

        registros = sorted(
            registros,
            key=lambda x: normalizar_texto_ordenacao(x.get("nome"))
        )

        nome_arquivo = self._caminho_pdf_lista_entrega(registros)
        if tree is getattr(self, "tree_entregues", None):
            self._gerar_pdf_entregas_historico(registros, nome_arquivo)
        else:
            self._gerar_pdf_entrega(registros, nome_arquivo)
            self._salvar_lista_entrega_arquivada(registros, nome_arquivo)
            for item in self.tree_lista.get_children():
                self.tree_lista.delete(item)
            salvar_lista_tree(self.tree_lista)

    # Gera PDF entregas histórico
    def _gerar_pdf_entregas_historico(self, registros, nome_arquivo):
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        import os
        from collections import defaultdict

        c = canvas.Canvas(nome_arquivo, pagesize=A4)
        largura, altura = A4

        grupos = defaultdict(list)
        for r in registros:
            acao_id = r.get("acao_id")
            acao = get_acao_por_id(self.db, acao_id)

            if acao:
                chave = (
                    acao.get("data", "Sem data"),
                    acao.get("local", "Sem local")
                )
            elif r.get("acao") or r.get("data"):
                chave = (
                    r.get("data", "Sem data"),
                    r.get("acao", "Sem local")
                )
            else:
                chave = ("Sem data", "Sem local")

            grupos[chave].append(r)

        y = altura - 160

        # Desenha cabeçalho
        def desenhar_cabecalho(c, texto_acao):
            caminho_logo = os.path.join(
                APP_DIR,
                "img",
                "Logo_Recife.png"
            )

            try:
                c.drawImage(
                    caminho_logo,
                    40,
                    altura - 95,
                    width=70,
                    height=70,
                    preserveAspectRatio=True,
                    mask='auto'
                )
            except Exception:
                pass

            c.setFont("Helvetica-Bold", 13)
            c.drawString(130, altura - 50, "PREFEITURA MUNICIPAL DO RECIFE")

            c.setFont("Helvetica", 11)
            c.drawString(130, altura - 65, "EMPRESA MUNICIPAL DE INFORMATICA - EMPREL")

            titulo = "RELAÇÃO DE IDENTIDADES ENTREGUES"
            c.setFont("Helvetica-Bold", 14)
            largura_texto = c.stringWidth(titulo, "Helvetica-Bold", 14)
            c.drawString((largura - largura_texto) / 2, altura - 100, titulo)

            c.line(30, altura - 110, largura - 30, altura - 110)

            c.drawString(30, altura - 130, texto_acao)

            c.line(30, altura - 140, largura - 30, altura - 140)

        # Desenha tabela cabeçalho
        def desenhar_tabela_header():
            nonlocal y
            c.setFont("Helvetica-Bold", 12)
            c.drawString(30, y, "Protocolo")
            c.drawString(100, y, "Nome")
            c.drawString(280, y, "CPF")
            c.drawString(380, y, "Data Entrega")
            y -= 15
            c.setFont("Helvetica", 10)

        for (data_acao, local_acao), lista in grupos.items():
            texto_acao = f"{data_acao} - {local_acao}"

            desenhar_cabecalho(c, texto_acao)

            y = altura - 160
            desenhar_tabela_header()

            for r in lista:
                if y < 60:
                    c.showPage()
                    desenhar_cabecalho(c, texto_acao)
                    y = altura - 160
                    desenhar_tabela_header()

                c.drawString(30, y, str(r.get("protocolo", "")))
                c.drawString(100, y, str(r.get("nome", "")))
                c.drawString(280, y, mascarar_cpf_pdf(r.get("cpf", "")))
                c.drawString(380, y, str(r.get("data_entrega", "")))

                y -= 20

            c.showPage()

        c.save()

        messagebox.showinfo("Sucesso", f"PDF gerado: {nome_arquivo}")

        try:
            os.startfile(nome_arquivo)
        except Exception as e:
            print("Erro ao abrir PDF:", e)

    # Gera PDF entrega
    def _gerar_pdf_entrega(self, registros, nome_arquivo):
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from datetime import datetime
        import os
        from collections import defaultdict

        c = canvas.Canvas(nome_arquivo, pagesize=A4)
        largura, altura = A4

        grupos = defaultdict(list)

        for r in registros:
            acao_id = r.get("acao_id")
            acao = get_acao_por_id(self.db, acao_id)

            if acao:
                chave = (
                    acao.get("data", "Sem data"),
                    acao.get("local", "Sem local")
                )
            elif r.get("acao") or r.get("data"):
                chave = (
                    r.get("data", "Sem data"),
                    r.get("acao", "Sem local")
                )
            else:
                chave = ("Sem data", "Sem local")

            grupos[chave].append(r)

        # Desenha cabeçalho
        def desenhar_cabecalho(c, texto_acao):
            caminho_logo = os.path.join(
                APP_DIR,
                "img",
                "Logo_Recife.png"
            )

            c.drawImage(
                caminho_logo,
                40,
                altura - 95,
                width=70,
                height=70,
                preserveAspectRatio=True,
                mask='auto'
            )

            c.setFont("Helvetica-Bold", 13)
            c.drawString(130, altura - 50, "PREFEITURA MUNICIPAL DO RECIFE")

            c.setFont("Helvetica", 11)
            c.drawString(130, altura - 65, "EMPRESA MUNICIPAL DE INFORMATICA - EMPREL")

            titulo = "RELAÇÃO DE IDENTIDADE PARA ENTREGA"
            c.setFont("Helvetica-Bold", 14)
            largura_texto = c.stringWidth(titulo, "Helvetica-Bold", 14)
            c.drawString((largura - largura_texto) / 2, altura - 100, titulo)

            c.line(30, altura - 110, largura - 30, altura - 110)

            c.drawString(30, altura - 130, texto_acao)

            c.line(30, altura - 140, largura - 30, altura - 140)

        # Desenha tabela cabeçalho
        def desenhar_tabela_header():
            nonlocal y
            c.setFont("Helvetica-Bold", 12)
            c.drawString(30, y, "Protocolo")
            c.drawString(95, y, "Nome")
            c.drawString(270, y, "CPF")
            c.drawString(350, y, "Telefone")
            c.drawString(440, y, "Assinatura")
            y -= 15
            c.setFont("Helvetica", 10)

        for (data_acao, local_acao), lista in grupos.items():

            texto_acao = f"{data_acao} - {local_acao}"

            desenhar_cabecalho(c, texto_acao)

            y = altura - 160
            desenhar_tabela_header()

            for r in lista:

                if y < 80:
                    c.showPage()
                    desenhar_cabecalho(c, texto_acao)
                    y = altura - 160
                    desenhar_tabela_header()

                c.drawString(30, y, str(r.get("protocolo", "")))
                c.drawString(95, y, str(r.get("nome", "")))
                c.drawString(270, y, mascarar_cpf_pdf(r.get("cpf", "")))
                c.drawString(350, y, str(r.get("telefone", "")))

                c.line(440, y, largura - 30, y)

                y -= 20

            if y < 200:
                c.showPage()
                desenhar_cabecalho(c, texto_acao)
                y = altura - 220

            c.setFont("Helvetica", 11)

            c.drawCentredString(
                largura / 2,
                y,
                "Declaro que recebi todas as identidades acima listadas"
            )

            y -= 25

            box_width = 230
            box_height = 90
            gap = 20

            x1 = (largura / 2) - box_width - (gap / 2)
            x2 = (largura / 2) + (gap / 2)

            y_box = y - box_height

            # Desenha caixa
            def desenhar_box(x, y_base):
                c.rect(x, y_base, box_width, box_height)

                c.setFont("Helvetica", 9)

                c.drawString(x + 10, y_base + box_height - 20, "Nome:")
                c.line(x + 50, y_base + box_height - 20, x + box_width - 10, y_base + box_height - 20)

                c.drawString(x + 10, y_base + box_height - 40, "CPF:")
                c.line(x + 40, y_base + box_height - 40, x + box_width - 10, y_base + box_height - 40)

                c.drawString(x + 10, y_base + box_height - 60, "Data:")
                c.line(x + 50, y_base + box_height - 60, x + box_width - 10, y_base + box_height - 60)

                c.line(x + 20, y_base + 15, x + box_width - 20, y_base + 15)
                c.drawCentredString(x + box_width / 2, y_base + 2, "Assinatura")

            desenhar_box(x1, y_box)
            desenhar_box(x2, y_box)

            c.showPage()

        c.save()

        messagebox.showinfo("Sucesso", f"PDF gerado: {nome_arquivo}")
        self._atualizar_pdfs_fichas_hoje()

        try:
            os.startfile(nome_arquivo)
        except Exception as e:
            print("Erro ao abrir PDF:", e)

    # Tela gera fichas
    def tela_gerar_fichas(self):
        self._acao_ficha_pdf_id = ""

        for widget in self.winfo_children():
            widget.destroy()

        self._ajustar_janela_principal()

        main = ctk.CTkFrame(self, fg_color="transparent")
        main.pack(fill="both", expand=True, padx=10, pady=10)

        self._criar_cabecalho(
            main,
            "GERAR FICHAS DE ATENDIMENTO",
            voltar=self._build_home_screen,
            mostrar_sair=True,
            texto_cancelar="Voltar"
        )

        container = ctk.CTkFrame(main, fg_color="transparent")
        container.pack(fill="both", expand=True, padx=20, pady=20)

        self._criar_titulo_pagina(
            container,
            "Emissão de fichas para organização da ação",
        )
        ctk.CTkLabel(container, text="Serviço:", text_color=("#334155", "#CBD5E1")).pack(anchor="w", pady=(15,0))

        self.combo_servico = tb.Combobox(
            container,
            values=[s["nome"] for s in self.db["servicos"]],
            state="readonly"
        )
        self.combo_servico.pack(fill="x")

        ctk.CTkLabel(container, text="Intervalo de Fichas:", text_color=("#334155", "#CBD5E1")).pack(anchor="w", pady=(10,0))

        frame_intervalo = ctk.CTkFrame(container, fg_color="transparent")
        frame_intervalo.pack(fill="x")

        self.ent_inicio_ficha = tb.Entry(frame_intervalo)
        self.ent_inicio_ficha.pack(side="left", padx=(0,5))
        self.ent_inicio_ficha.configure(width=8)
        self.ent_inicio_ficha.insert(0, "1")

        self.ent_fim_ficha = tb.Entry(frame_intervalo)
        self.ent_fim_ficha.pack(side="left", padx=(5,0))
        self.ent_fim_ficha.configure(width=8)

        ctk.CTkLabel(container, text="Data da Ação:", text_color=("#334155", "#CBD5E1")).pack(anchor="w", pady=(15,0))

        self.var_data_acao_ficha = tk.StringVar()

        datas_acoes = sorted(
            {str(a.get("data", "")).strip() for a in self.db.get("acoes", []) if str(a.get("data", "")).strip()},
            key=lambda d: datetime.strptime(d, "%d/%m/%Y"),
            reverse=True
        )

        self.combo_data_acao = tb.Combobox(
            container,
            textvariable=self.var_data_acao_ficha,
            values=datas_acoes,
            state="readonly"
        )
        self.combo_data_acao.pack(fill="x")

        ctk.CTkLabel(container, text="Nome da Ação:", text_color=("#334155", "#CBD5E1")).pack(anchor="w", pady=(10,0))

        self.var_nome_acao_ficha = tk.StringVar()

        self.ent_nome_acao_ficha = tb.Entry(
            container,
            textvariable=self.var_nome_acao_ficha,
            state="readonly"
        )
        self.ent_nome_acao_ficha.pack(fill="x")

        # Atualiza nome ação
        def atualizar_nome_acao(event=None):
            data_sel = self.var_data_acao_ficha.get().strip()
            acao = next(
                (
                    a for a in self.db.get("acoes", [])
                    if str(a.get("data", "")).strip() == data_sel
                ),
                None
            )

            if acao:
                self._acao_ficha_pdf_id = str(acao.get("id", "")).strip()
                self.var_nome_acao_ficha.set(acao.get("local", ""))
            else:
                self._acao_ficha_pdf_id = ""
                self.var_nome_acao_ficha.set("")

        self.combo_data_acao.bind(
            "<<ComboboxSelected>>",
            atualizar_nome_acao
        )

        self.var_data_acao_ficha.set("")
        self.var_nome_acao_ficha.set("")

        ctk.CTkLabel(container, text="Período:", text_color=("#334155", "#CBD5E1")).pack(anchor="w", pady=(15,0))
        self.combo_periodo = tb.Combobox(
            container,
            values=["Manhã", "Tarde", "Noite", "TRIAGEM"],
            state="readonly"
        )
        self.combo_periodo.current(0)
        self.combo_periodo.pack(fill="x")

        btns = ctk.CTkFrame(container, fg_color="transparent")
        btns.pack(pady=25)

        ctk.CTkButton(
            btns,
            text="Gerar PDF",
            command=self._gerar_pdf_fichas,
            fg_color="#16A34A",
            hover_color="#15803D",
            text_color="white",
            corner_radius=8,
            height=38,
            font=("Segoe UI", 12, "bold")
        ).pack(side="left", padx=10)

        self.frame_pdfs_fichas_hoje = tb.Labelframe(
            container,
            text="PDFs de fichas gerados hoje",
            padding=10
        )
        self.frame_pdfs_fichas_hoje.pack(fill="x", pady=(0, 10))
        self._atualizar_pdfs_fichas_hoje()

    # Pasta PDF fichas
    def _pasta_pdf_fichas(self):
        pasta = os.path.join(APP_DIR, "pdf", "fichas")
        os.makedirs(pasta, exist_ok=True)
        return pasta

    # Pdfs fichas de hoje
    def _pdfs_fichas_de_hoje(self):
        pasta = self._pasta_pdf_fichas()
        hoje = datetime.now().date()
        arquivos = []

        for nome in os.listdir(pasta):
            if not nome.lower().endswith(".pdf"):
                continue
            caminho = os.path.join(pasta, nome)
            if os.path.isfile(caminho) and datetime.fromtimestamp(os.path.getmtime(caminho)).date() == hoje:
                arquivos.append(caminho)

        arquivos.sort(key=lambda p: os.path.getmtime(p), reverse=True)
        return arquivos

    # Abre PDF ficha
    def _abrir_pdf_ficha(self, caminho):
        try:
            os.startfile(caminho)
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível abrir o PDF.\n\n{e}")

    # Atualiza pdfs fichas hoje
    def _atualizar_pdfs_fichas_hoje(self):
        frame = getattr(self, "frame_pdfs_fichas_hoje", None)
        if not frame or not frame.winfo_exists():
            return

        for widget in frame.winfo_children():
            widget.destroy()

        arquivos = self._pdfs_fichas_de_hoje()

        if not arquivos:
            tb.Label(
                frame,
                text="Nenhum PDF de fichas gerado hoje."
            ).pack(anchor="w")
            return

        max_por_linha = 5

        for i, caminho in enumerate(arquivos):

            if i % max_por_linha == 0:
                linha = tb.Frame(frame)
                linha.pack(anchor="w", pady=2)

            nome = os.path.basename(caminho)

            partes = nome.replace(".pdf", "").split("_")

            servico_pdf = (
                partes[0].replace("PDF", "", 1)
                if partes else "FICHAS"
            )

            horario = (
                partes[2].replace("-", ":")
                if len(partes) >= 3
                else ""
            )

            texto = f"[PDF] {servico_pdf} {horario}".strip()

            tb.Button(
                linha,
                text=texto,
                bootstyle="secondary-outline",
                command=lambda c=caminho: self._abrir_pdf_ficha(c)
            ).pack(side="left", padx=5, pady=2)

    # Gera PDF fichas
    def _gerar_pdf_fichas(self):

        servico = self.combo_servico.get().strip()
        inicio = self.ent_inicio_ficha.get().strip()
        fim = self.ent_fim_ficha.get().strip()

        data = self.combo_data_acao.get().strip()
        nome_acao = self.var_nome_acao_ficha.get().strip()

        periodo = self.combo_periodo.get()

        acao_ficha_id = getattr(self, "_acao_ficha_pdf_id", "")
        acao_ficha = get_acao_por_id(self.db, acao_ficha_id) if acao_ficha_id else None

        if not servico or not inicio or not fim or not periodo:
            messagebox.showwarning("Aviso", "Preencha todos os campos.")
            return

        if not data or not nome_acao or not acao_ficha:
            messagebox.showwarning(
                "Selecione uma ação",
                "Selecione a data da ação para preencher automaticamente o nome da ação antes de gerar o PDF."
            )
            return

        data_acao = str(acao_ficha.get("data", "")).strip()
        nome_acao_automatico = str(acao_ficha.get("local", "")).strip()

        if data != data_acao or nome_acao != nome_acao_automatico:
            messagebox.showwarning(
                "Selecione uma ação",
                "O nome da ação deve ser preenchido automaticamente pela data escolhida antes de gerar o PDF."
            )
            return

        try:
            inicio = int(inicio)
            fim = int(fim)
        except:
            messagebox.showerror("Erro", "Intervalo inválido.")
            return

        if inicio > fim:
            messagebox.showerror("Erro", "O início não pode ser maior que o fim.")
            return

        nome_arquivo = caminho_pdf("fichas", f"PDF{servico.upper()}")

        c = canvas.Canvas(nome_arquivo, pagesize=A4)
        largura_pagina, altura_pagina = A4

        largura_ficha = 4.8 * cm
        altura_ficha = 5.5 * cm

        espaco_horizontal = 0.3 * cm
        espaco_vertical = 0.3 * cm

        colunas = 4

        margem_x = 0.5 * cm
        margem_y = 0.5 * cm

        colunas = int(
            (largura_pagina - margem_x)
            // (largura_ficha + espaco_horizontal)
        )

        linhas = int(
            (altura_pagina - margem_y)
            // (altura_ficha + espaco_vertical)
        )

        ficha_num = inicio

        total_fichas = fim - inicio + 1

        for i in range(total_fichas):

            coluna = i % colunas
            linha = (i // colunas) % linhas

            x = margem_x + coluna * (largura_ficha + espaco_horizontal)
            y = altura_pagina - margem_y - altura_ficha - linha * (altura_ficha + espaco_vertical)

            if i != 0 and linha == 0 and coluna == 0:
                c.showPage()

            c.setLineWidth(1)
            c.setDash()
            c.rect(x, y, largura_ficha, altura_ficha)

            espaco_corte = 0.1 * cm

            c.setDash(3, 3)
            c.setLineWidth(0.8)

            c.rect(
                x - espaco_corte,
                y - espaco_corte,
                largura_ficha + (espaco_corte * 2),
                altura_ficha + (espaco_corte * 2)
            )

            c.setDash()

            centro_x = x + largura_ficha / 2

            altura_faixa_topo = 0.8 * cm
            c.setFillColorRGB(0, 0, 0)
            c.rect(x, y + altura_ficha - altura_faixa_topo, largura_ficha, altura_faixa_topo, fill=1)

            c.setFillColorRGB(1, 1, 1)
            c.setFont("Helvetica-Bold", 9)
            c.drawCentredString(
                centro_x,
                y + altura_ficha - 0.55 * cm,
                "CONECTA ITINERANTE"
            )

            c.setFillColorRGB(0, 0, 0)
            c.setFont("Helvetica-Bold", 14)
            c.drawCentredString(
                centro_x,
                y + altura_ficha - 1.5 * cm,
                servico.upper()
            )

            c.setFont("Helvetica-Bold", 30)
            c.drawCentredString(
                centro_x,
                y + altura_ficha - 3.0 * cm,
                str(ficha_num).zfill(2)
            )

            altura_faixa_periodo = 0.7 * cm
            posicao_faixa = y + 1.3 * cm

            c.setFillColorRGB(0, 0, 0)
            c.rect(x, posicao_faixa, largura_ficha, altura_faixa_periodo, fill=1)

            c.setFillColorRGB(1, 1, 1)
            c.setFont("Helvetica-Bold", 13)
            c.drawCentredString(
                centro_x,
                posicao_faixa + 0.23 * cm,
                periodo.upper()
            )

            c.setFillColorRGB(0, 0, 0)
            c.setFont("Helvetica-Bold", 12)
            c.drawCentredString(
                centro_x,
                y + 0.6 * cm,
                data
            )

            c.setFont("Helvetica", 8)

            c.drawCentredString(
                centro_x,
                y + 0.25 * cm,
                nome_acao[:35]
            )

            ficha_num += 1

        c.save()

        self._atualizar_pdfs_fichas_hoje()
        self.update_idletasks()

        messagebox.showinfo("Sucesso", f"PDF gerado: {nome_arquivo}")

        try:
            os.startfile(nome_arquivo)
        except:
            pass

if __name__ == "__main__":
    app = App()
    app.mainloop()
