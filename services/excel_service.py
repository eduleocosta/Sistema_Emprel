import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import pandas as pd
from config import APP_DIR

# Cria planilha fichas
def criar_planilha_fichas(dados, nome_arquivo):
    wb = Workbook()
    ws = wb.active
    ws.title = "Fichas"

    headers = ["Protocolo", "Nome", "CPF", "Telefone", "Serviço", "Data Ação"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for row_idx, reg in enumerate(dados, 2):
        ws.cell(row=row_idx, column=1, value=reg.get("protocolo", ""))
        ws.cell(row=row_idx, column=2, value=reg.get("nome", ""))
        ws.cell(row=row_idx, column=3, value=reg.get("cpf", ""))
        ws.cell(row=row_idx, column=4, value=reg.get("telefone", ""))
        ws.cell(row=row_idx, column=5, value=reg.get("servico", ""))
        ws.cell(row=row_idx, column=6, value=reg.get("data", ""))

    wb.save(nome_arquivo)
    return nome_arquivo

# Pasta PDF fichas
def pasta_pdf_fichas():
    pasta = os.path.join(APP_DIR, "pdf", "fichas")
    os.makedirs(pasta, exist_ok=True)
    return pasta

# Pdfs fichas de hoje
def pdfs_fichas_de_hoje():
    pasta = pasta_pdf_fichas()
    hoje = datetime.now().date()
    arquivos = []

    for nome in os.listdir(pasta):
        if not nome.lower().endswith(".pdf"):
            continue
        caminho = os.path.join(pasta, nome)
        if os.path.isfile(caminho) and datetime.fromtimestamp(os.path.getmtime(caminho)).date() == hoje:
            arquivos.append(caminho)

    arquivos.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return arquivos